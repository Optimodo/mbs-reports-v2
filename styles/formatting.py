"""Formatting and styling configurations for Excel reports."""

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


# Overall Summary Sheet Style Configuration
OVERALL_SUMMARY_STYLES = {
    'title': {
        'font': Font(name='Calibri', size=14, bold=True, color='000000'),
        'alignment': Alignment(horizontal='left', vertical='center'),
        'fill': PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    },
    'timestamp': {
        'font': Font(name='Calibri', size=11, italic=True, color='000000'),
        'alignment': Alignment(horizontal='left', vertical='center'),
        'fill': PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    },
    'section_header': {
        'font': Font(name='Calibri', size=12, bold=True, color='000000'),
        'alignment': Alignment(horizontal='left', vertical='center'),
        'fill': PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
    },
    'column_header': {
        'font': Font(name='Calibri', size=11, bold=True, color='000000'),
        'alignment': Alignment(horizontal='left', vertical='center'),
        'fill': PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
    },
    'data_cell': {
        'font': Font(name='Calibri', size=11, color='000000'),
        'alignment': Alignment(horizontal='left', vertical='center'),
        'fill': PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    },
    'total_cell': {
        'font': Font(name='Calibri', size=11, bold=True, color='000000'),
        'alignment': Alignment(horizontal='left', vertical='center'),
        'fill': PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    },
    'border': Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
}

# Progression Report Status Configuration
PROGRESSION_STATUS_ORDER = {
    'Status A': {
        'display_name': 'Status A',
        'status_terms': [
            'A - Authorized and Accepted',
            'Accepted',
            'A',
            'A - Proceed',
            'A - Proceed (Lead Reviewer)'
        ]
    },
    'Status B': {
        'display_name': 'Status B',
        'status_terms': [
            'B - Partial Sign Off (with comment)',
            'Accepted with Comments',
            'B',
            'B - Proceed with Comments',
            'B - Proceed with Comments (Lead Reviewer)'
        ]
    },
    'Status C': {
        'display_name': 'Status C',
        'status_terms': [
            'Rejected',
            'QA - Rejected',
            'C - Rejected',
            'C',
            'C - Rejected (Lead Reviewer)',
            'C-Rejected'
        ]
    },
    'QC Rejected': {
        'display_name': 'QA/QC Rejected',
        'status_terms': [
            'QC Rejected',
            'QA Rejected'
        ]
    },
    'QC Checked': {
        'display_name': 'QC Checked',
        'status_terms': [
            'QC Checked',
            'QC Accepted'
        ]
    },
    'Under Review': {
        'display_name': 'Under Review/For Commenting',
        'status_terms': [
            'Under Review',
            'For Status Change',
            'For Commenting',
            'Awaiting QC Check',
            'Shared'
        ]
    },
    'Preliminary': {
        'display_name': 'Preliminary',
        'status_terms': [
            'Preliminary'
        ]
    },
    'Other': {
        'display_name': 'Other',
        'status_terms': [
            'Other'
        ]
    }
}

# Status-based conditional formatting
STATUS_STYLES = {
    'STATUS A': {
        'search_terms': ['A - Authorized and Accepted', 'Accepted', 'A', 'A - Proceed', 'A - Proceed (Lead Reviewer)', 'Status A'],
        'style': {
            'font': Font(name='Calibri', size=11, bold=True, color='000000'),
            'fill': PatternFill(start_color='25E82C', end_color='25E82C', fill_type='solid')
        }
    },
    'STATUS B': {
        'search_terms': ['B - Partial Sign Off (with comment)', 'Accepted with Comments', 'B', 'B - Proceed with Comments', 'B - Proceed with Comments (Lead Reviewer)', 'Status B'],
        'style': {
            'font': Font(name='Calibri', size=11, bold=True, color='000000'),
            'fill': PatternFill(start_color='EDDDA1', end_color='EDDDA1', fill_type='solid')
        }
    },
    'STATUS C': {
        'search_terms': ['Rejected', 'QA - Rejected', 'C - Rejected', 'QA Rejected', 'C', 'C - Rejected (Lead Reviewer)', 'C-Rejected', 'Status C'],
        'style': {
            'font': Font(name='Calibri', size=11, bold=True, color='000000'),
            'fill': PatternFill(start_color='ED1111', end_color='ED1111', fill_type='solid')
        }
    },
    'INFORMATION': {
        'search_terms': ['Information', 'Withdrawn-Obsolete', 'QA Passed', 'QA - Passed', 'For Status Change', 'For Commenting', 'Awaiting QC Check', 'QC Accepted', 'QC Checked', 'Under Review'],
        'style': {
            'font': Font(name='Calibri', size=11, bold=True, color='000000'),
            'fill': PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        }
    },
    'REVIEWED': {
        'search_terms': ['Reviewed'],
        'style': {
            'font': Font(name='Calibri', size=11, bold=True, color='000000'),
            'fill': PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        }
    },
    'PUBLISHED': {
        'search_terms': ['Published'],
        'style': {
            'font': Font(name='Calibri', size=11, bold=True, color='000000'),
            'fill': PatternFill(start_color='67DBB5', end_color='67DBB5', fill_type='solid')
        }
    },
    'SHARED': {
        'search_terms': ['Shared'],
        'style': {
            'font': Font(name='Calibri', size=11, bold=True, color='000000'),
            'fill': PatternFill(start_color='E0F090', end_color='E0F090', fill_type='solid')
        }
    },
    'PRELIMINARY': {
        'search_terms': ['Preliminary'],
        'style': {
            'font': Font(name='Calibri', size=11, bold=True, color='000000'),
            'fill': PatternFill(start_color='87CEEB', end_color='87CEEB', fill_type='solid')  # Light blue
        }
    },
    'OTHER': {
        'search_terms': ['Other'],
        'style': {
            'font': Font(name='Calibri', size=11, bold=True, color='000000'),
            'fill': PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')  # Light gray
        }
    }
}


def apply_status_style(cell, status_name, config=None):
    """Apply conditional formatting based on status name.
    
    Args:
        cell: openpyxl cell object
        status_name: Status string to match against
        config: Project configuration dictionary (optional)
        
    Returns:
        dict: The style configuration applied
    """
    # Try to use config-based styling first
    if config and 'STATUS_MAPPINGS' in config:
        status_mappings = config['STATUS_MAPPINGS']
        if status_name in status_mappings:
            color = status_mappings[status_name].get('color')
            if color:
                cell.font = OVERALL_SUMMARY_STYLES['data_cell']['font']
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                return OVERALL_SUMMARY_STYLES['data_cell']
    
    # Fallback to old hardcoded styling
    for style_config in STATUS_STYLES.values():
        if any(term == status_name for term in style_config['search_terms']):
            cell.font = style_config['style']['font']
            cell.fill = style_config['style']['fill']
            return style_config['style']
    
    # If no matching style found, use default data cell style
    cell.font = OVERALL_SUMMARY_STYLES['data_cell']['font']
    cell.fill = OVERALL_SUMMARY_STYLES['data_cell']['fill']
    return OVERALL_SUMMARY_STYLES['data_cell']

