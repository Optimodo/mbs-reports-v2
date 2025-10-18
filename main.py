"""Main orchestration script for MBS Document Reporter.

Database-driven document register processing and reporting.
"""

import warnings
import pandas as pd
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from config import load_project_config

# Suppress warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.filterwarnings('ignore', category=FutureWarning)

# Import from modular structure
from analyzers import create_summary_row
from utils.document_filters import get_main_report_data
from reports import (
    save_excel_with_retry,
    generate_progression_report,
    fill_empty_cells_with_zeros_in_file,
    save_certificate_report_with_retry
)
from utils import slugify
from utils.document_filters import filter_certificates, get_document_type_summary
from data import DocumentDatabase
from scripts.db_manager import update_database_with_new_files


def show_menu():
    """Display the main menu and get user choice."""
    print("\n" + "="*60)
    print("MBS Reports v2 - Document Register Reporter")
    print("="*60)
    print("1. Generate ALL reports for ALL projects")
    print("2. Generate ALL reports for SINGLE project")
    print("3. Generate SPECIFIC report type")
    print("4. Exit")
    print("="*60)
    
    choice = input("\nEnter your choice (1-4): ").strip()
    return choice


def get_project_selection(include_all_option=False):
    """Get project selection from user.
    
    Args:
        include_all_option: If True, includes "All Projects" as first option
        
    Returns:
        str: Project name, 'ALL' for all projects, or None if cancelled
    """
    print("\n" + "="*60)
    print("Select Project:")
    print("="*60)
    
    if include_all_option:
        print("1. All Projects")
        print("2. Oval Block B")
        print("3. New Malden")
        print("4. Greenwich Peninsula")
        print("5. Holloway Park")
        print("6. West Cromwell Road")
        print("7. Cancel")
        print("="*60)
        
        choice = input("\nEnter your choice (1-7): ").strip()
        
        if choice == '7':
            return None
        elif choice == '1':
            return 'ALL'
        else:
            project_map = {
                '2': 'OvalBlockB',
                '3': 'NewMalden',
                '4': 'GreenwichPeninsula',
                '5': 'HollowayPark',
                '6': 'WestCromwellRoad'
            }
            return project_map.get(choice)
    else:
        print("1. Oval Block B")
        print("2. New Malden")
        print("3. Greenwich Peninsula")
        print("4. Holloway Park")
        print("5. West Cromwell Road")
        print("6. Cancel")
        print("="*60)
        
        choice = input("\nEnter your choice (1-6): ").strip()
        
        if choice == '6':
            return None
        
        project_map = {
            '1': 'OvalBlockB',
            '2': 'NewMalden',
            '3': 'GreenwichPeninsula',
            '4': 'HollowayPark',
            '5': 'WestCromwellRoad'
        }
        return project_map.get(choice)


def get_report_type_selection():
    """Get report type selection from user.
    
    Returns:
        str: Report type ('summary', 'progression', 'condensed', 'certificates') or None if cancelled
    """
    print("\n" + "="*60)
    print("Select Report Type:")
    print("="*60)
    print("1. Summary Report")
    print("2. Detailed Progression Report")
    print("3. Condensed Progression Report")
    print("4. Certificate Report")
    print("5. Cancel")
    print("="*60)
    
    choice = input("\nEnter your choice (1-5): ").strip()
    
    report_map = {
        '1': 'summary',
        '2': 'progression',
        '3': 'condensed',
        '4': 'certificates'
    }
    
    if choice == '5':
        return None
    
    return report_map.get(choice)


def generate_summary_report(project_name, config, output_dir, db):
    """Generate summary report for a project using dynamic counting.
    
    Args:
        project_name: Name of the project
        config: Project configuration
        output_dir: Output directory path
        db: Database connection
        
    Returns:
        bool: True if successful
    """
    project_slug = slugify(project_name)
    summary_output = output_dir / f"{project_slug}_summary.xlsx"
    
    # Get latest snapshot data
    latest_data_df = db.get_latest_documents(project_name)
    
    if latest_data_df.empty:
        print(f"  ✗ No data for {project_name}")
        return False
    
    # Apply drawing filters for main report
    filtered_data = get_main_report_data(latest_data_df, config)
    
    # Create a single-row summary DataFrame with dynamic counts
    # Get snapshot date/time from the latest data
    cursor = db.conn.cursor()
    cursor.execute("""
        SELECT snapshot_date, snapshot_time 
        FROM documents 
        WHERE project_name = ?
        ORDER BY snapshot_date DESC, snapshot_time DESC
        LIMIT 1
    """, (project_name,))
    
    result = cursor.fetchone()
    if not result:
        print(f"  ✗ No snapshot found for {project_name}")
        return False
    
    snapshot_date, snapshot_time = result
    
    # Create summary row using dynamic counting
    summary_row = create_summary_row(snapshot_date, snapshot_time, filtered_data, config)
    summary_df = pd.DataFrame([summary_row])
    
    # Generate the report
    if save_excel_with_retry(summary_df, None, filtered_data, summary_output, config):
        print(f"  ✓ Summary report: {summary_output}")
        return True
    else:
        print(f"  ✗ Failed to save summary report")
        return False


def generate_progression_report_full(project_name, config, output_dir, db):
    """Generate detailed progression report for a project using dynamic counting.
    
    Args:
        project_name: Name of the project
        config: Project configuration
        output_dir: Output directory path
        db: Database connection
        
    Returns:
        bool: True if successful
    """
    project_slug = slugify(project_name)
    progression_output = output_dir / f"{project_slug}_progression.xlsx"
    
    # Delete existing report to rebuild from scratch
    if progression_output.exists():
        progression_output.unlink()
    
    # Get all snapshots from database (just dates/times)
    query = """
        SELECT DISTINCT snapshot_date, snapshot_time 
        FROM documents 
        WHERE project_name = ?
        ORDER BY snapshot_date, snapshot_time
    """
    cursor = db.conn.execute(query, (project_name,))
    snapshots = cursor.fetchall()
    
    if not snapshots:
        print(f"  ℹ No snapshots found")
        return False
    
    # Process each snapshot with dynamic counting
    for snapshot_date, snapshot_time in snapshots:
        # Get raw documents for this snapshot
        snapshot_docs = db.get_documents_for_snapshot(project_name, snapshot_date, snapshot_time)
        
        if snapshot_docs.empty:
            continue
        
        # Filter to main report documents (drawings/schematics only)
        filtered_docs = get_main_report_data(snapshot_docs, config)
        
        # Convert database date format to display format
        try:
            date_obj = datetime.strptime(snapshot_date, '%Y-%m-%d')
            display_date = date_obj.strftime('%d-%b-%Y')
        except:
            display_date = snapshot_date
        
        # Create dynamic summary row for this snapshot
        summary_row = create_summary_row(display_date, snapshot_time, filtered_docs, config)
        snapshot_summary_df = pd.DataFrame([summary_row])
        
        # Generate progression report (adds one column)
        if not generate_progression_report(snapshot_summary_df, progression_output, config, filtered_docs):
            print(f"  ✗ Failed column: {display_date} {snapshot_time}")
            return False
    
    fill_empty_cells_with_zeros_in_file(str(progression_output))
    print(f"  ✓ Progression report: {progression_output}")
    return True


def generate_condensed_report(project_name, config, output_dir, db, num_weeks=4):
    """Generate condensed progression report for a project using dynamic counting.
    
    Args:
        project_name: Name of the project
        config: Project configuration
        output_dir: Output directory path
        db: Database connection
        num_weeks: Number of recent weeks to include
        
    Returns:
        bool: True if successful
    """
    # Get all snapshots for this project
    cursor = db.conn.cursor()
    cursor.execute("""
        SELECT DISTINCT snapshot_date, snapshot_time
        FROM documents
        WHERE project_name = ?
        ORDER BY snapshot_date, snapshot_time
    """, (project_name,))
    
    all_snapshots = cursor.fetchall()
    
    if not all_snapshots:
        print(f"  ℹ No data for condensed report")
        return False
    
    # Get the last N snapshots (weekly/recent data)
    last_n_snapshots = all_snapshots[-num_weeks:] if len(all_snapshots) >= num_weeks else all_snapshots
    
    # Identify which months are covered by the last N snapshots
    covered_months = set()
    for snapshot_date, snapshot_time in last_n_snapshots:
        date_obj = datetime.strptime(snapshot_date, '%Y-%m-%d')
        month_key = (date_obj.year, date_obj.month)
        covered_months.add(month_key)
    
    # Build condensed snapshot list:
    # 1. Last N snapshots (weekly detail)
    # 2. One snapshot per month for older months NOT covered by last N
    condensed_snapshots = []
    monthly_snapshots = {}  # Group by month for older data
    
    # Add last N snapshots as weekly
    for snapshot_date, snapshot_time in last_n_snapshots:
        condensed_snapshots.append((snapshot_date, snapshot_time, False))  # False = weekly
    
    # Process older snapshots - keep one per month for uncovered months
    older_snapshots = all_snapshots[:-num_weeks] if len(all_snapshots) > num_weeks else []
    for snapshot_date, snapshot_time in older_snapshots:
        date_obj = datetime.strptime(snapshot_date, '%Y-%m-%d')
        month_key = (date_obj.year, date_obj.month)
        
        # Only include if this month is NOT already covered by last N snapshots
        if month_key not in covered_months:
            # Keep the latest snapshot for this month
            if month_key not in monthly_snapshots or date_obj > monthly_snapshots[month_key][0]:
                monthly_snapshots[month_key] = (date_obj, snapshot_date, snapshot_time)
    
    # Add monthly snapshots at the beginning
    for month_key, (date_obj, snapshot_date, snapshot_time) in sorted(monthly_snapshots.items()):
        condensed_snapshots.insert(0, (snapshot_date, snapshot_time, True))  # True = monthly
    
    # Sort all snapshots chronologically
    condensed_snapshots.sort(key=lambda x: (x[0], x[1]))
    
    if not condensed_snapshots:
        print(f"  ℹ No data for condensed report")
        return False
    
    project_slug = slugify(project_name)
    condensed_output = output_dir / f"{project_slug}_progression_condensed.xlsx"
    
    # Delete existing report
    if condensed_output.exists():
        condensed_output.unlink()
    
    # Process each snapshot with dynamic counting
    for snapshot_date, snapshot_time, is_monthly in condensed_snapshots:
        
        if not snapshot_date or not snapshot_time:
            continue
        
        # Get raw documents for this snapshot
        snapshot_docs = db.get_documents_for_snapshot(project_name, snapshot_date, snapshot_time)
        
        if snapshot_docs.empty:
            continue
        
        # Filter to main report documents (drawings/schematics only)
        filtered_docs = get_main_report_data(snapshot_docs, config)
        
        # Convert database date format to display format
        # Monthly: "Jun-2025", Weekly: "07-Oct-2025"
        try:
            date_obj = datetime.strptime(snapshot_date, '%Y-%m-%d')
            if is_monthly:
                display_date = date_obj.strftime('%b-%Y')  # "Jun-2025" for monthly
            else:
                display_date = date_obj.strftime('%d-%b-%Y')  # "07-Oct-2025" for weekly
        except:
            display_date = snapshot_date
        
        # Create dynamic summary row for this snapshot
        summary_row = create_summary_row(display_date, snapshot_time, filtered_docs, config)
        snapshot_summary_df = pd.DataFrame([summary_row])
        
        # Generate progression report (adds one column)
        if generate_progression_report(snapshot_summary_df, condensed_output, config, filtered_docs):
            # Apply blue formatting to monthly columns
            if is_monthly:
                try:
                    book = load_workbook(condensed_output)
                    if 'Progression' in book.sheetnames:
                        sheet = book['Progression']
                        last_col = sheet.max_column
                        date_cell = sheet.cell(row=1, column=last_col)
                        date_cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
                        date_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                        book.save(condensed_output)
                except:
                    pass
    
    fill_empty_cells_with_zeros_in_file(str(condensed_output))
    print(f"  ✓ Condensed report: {condensed_output}")
    return True


def generate_certificate_report_full(project_name, config, output_dir, db):
    """Generate certificate report for a project using dynamic counting.
    
    Args:
        project_name: Name of the project
        config: Project configuration
        output_dir: Output directory path
        db: Database connection
        
    Returns:
        bool: True if successful
    """
    cert_settings = config.get('CERTIFICATE_SETTINGS', {})
    
    if not cert_settings.get('enabled', False):
        print(f"  ℹ Certificates not enabled for this project")
        return False
    
    if not cert_settings.get('generate_report', False):
        print(f"  ℹ Certificate report generation not enabled")
        return False
    
    # Get latest documents
    latest_data_df = db.get_latest_documents(project_name)
    
    if latest_data_df.empty:
        print(f"  ℹ No data found")
        return False
    
    # Filter for certificates only
    cert_data = filter_certificates(latest_data_df, config)
    
    if cert_data.empty:
        print(f"  ℹ No certificate documents found")
        return False
    
    # Get all snapshots and build certificate summary dynamically
    cursor = db.conn.cursor()
    cursor.execute("""
        SELECT DISTINCT snapshot_date, snapshot_time
        FROM documents
        WHERE project_name = ?
        ORDER BY snapshot_date, snapshot_time
    """, (project_name,))
    
    snapshots = cursor.fetchall()
    
    # Build certificate summary using dynamic counting
    cert_summary_rows = []
    for snapshot_date, snapshot_time in snapshots:
        # Get documents for this snapshot
        snapshot_docs = db.get_documents_for_snapshot(project_name, snapshot_date, snapshot_time)
        
        # Filter for certificates only
        snapshot_certs = filter_certificates(snapshot_docs, config)
        
        if not snapshot_certs.empty:
            # Use dynamic counting via create_summary_row
            summary_row = create_summary_row(snapshot_date, snapshot_time, snapshot_certs, config)
            cert_summary_rows.append(summary_row)
    
    if not cert_summary_rows:
        print(f"  ℹ No certificate data in snapshots")
        return False
    
    cert_summary_df = pd.DataFrame(cert_summary_rows)
    cols = ['Date', 'Time'] + [c for c in cert_summary_df.columns if c not in ['Date', 'Time']]
    cert_summary_df = cert_summary_df[cols]
    
    project_slug = slugify(project_name)
    cert_output = output_dir / f"{project_slug}_certificates.xlsx"
    
    if save_certificate_report_with_retry(cert_summary_df, cert_data, cert_output, config):
        print(f"  ✓ Certificate report: {cert_output}")
        return True
    else:
        print(f"  ✗ Failed to save certificate report")
        return False


def process_single_project_all_reports(project_name):
    """Generate all reports for a single project.
    
    Args:
        project_name: Name of the project
        
    Returns:
        bool: True if successful
    """
    print(f"\n{'='*60}")
    print(f"Processing: {project_name}")
    print(f"{'='*60}")
    
    output_dir = Path('output')
    output_dir.mkdir(exist_ok=True)
    
    try:
        with DocumentDatabase() as db:
            config = load_project_config(project_name)
            
            # Check if we have data
            latest_data_df = db.get_latest_documents(project_name)
            if latest_data_df.empty:
                print(f"✗ No data for {project_name}")
                print("  Run 'Update database' option first")
                return False
            
            # Generate all reports
            print("\nGenerating reports...")
            success = True
            
            # 1. Summary Report (with dynamic counting)
            if not generate_summary_report(project_name, config, output_dir, db):
                success = False
            
            # 2. Detailed Progression Report (with dynamic counting)
            if not generate_progression_report_full(project_name, config, output_dir, db):
                success = False
            
            # 3. Condensed Progression Report (with dynamic counting)
            generate_condensed_report(project_name, config, output_dir, db, num_weeks=4)
            
            # 4. Certificate Report (if enabled)
            generate_certificate_report_full(project_name, config, output_dir, db)
            
            if success:
                print(f"\n✓ All reports completed for {project_name}")
            else:
                print(f"\n⚠ Some reports failed for {project_name}")
            
            return success
            
    except Exception as e:
        print(f"✗ Error processing {project_name}: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


def process_all_projects_all_reports():
    """Generate all reports for all projects."""
    print("\n" + "="*60)
    print("PROCESSING ALL PROJECTS - ALL REPORTS")
    print("="*60)
    
    # Update database first
    print("\nStep 1: Checking for new files...")
    try:
        stats = update_database_with_new_files()
        if stats['files_imported'] > 0:
            print(f"✓ Imported {stats['files_imported']} new files")
        else:
            print("✓ Database is up to date")
    except Exception as e:
        print(f"✗ Error updating database: {str(e)}")
        print("Continuing with existing data...")
    
    # Get all projects
    print("\nStep 2: Generating reports...")
    
    with DocumentDatabase() as db:
        projects = db.get_all_projects()
        
        if not projects:
            print("✗ No projects in database")
            return
        
        print(f"Found {len(projects)} projects\n")
        
        success_count = 0
        fail_count = 0
        
        for project_name in projects:
            if process_single_project_all_reports(project_name):
                success_count += 1
            else:
                fail_count += 1
        
        # Summary
        print(f"\n{'='*60}")
        print("PROCESSING COMPLETE")
        print(f"{'='*60}")
        print(f"✓ Successful: {success_count} projects")
        if fail_count > 0:
            print(f"✗ Failed: {fail_count} projects")
        print(f"{'='*60}")


def generate_specific_report_for_projects(report_type, project_names):
    """Generate a specific report type for specified project(s).
    
    Args:
        report_type: Type of report ('summary', 'progression', 'condensed', 'certificates')
        project_names: List of project names or ['ALL']
    """
    output_dir = Path('output')
    output_dir.mkdir(exist_ok=True)
    
    # Update database first
    print("\nChecking for new files...")
    try:
        stats = update_database_with_new_files()
        if stats['files_imported'] > 0:
            print(f"✓ Imported {stats['files_imported']} new files")
    except Exception as e:
        print(f"⚠ Warning: {str(e)}")
    
    # Get list of projects
    with DocumentDatabase() as db:
        if 'ALL' in project_names:
            projects = db.get_all_projects()
            if not projects:
                print("✗ No projects in database")
                return
            print(f"\nGenerating {report_type} reports for ALL {len(projects)} projects")
        else:
            projects = project_names
            print(f"\nGenerating {report_type} report(s)")
    
    print(f"{'='*60}\n")
    
    success_count = 0
    fail_count = 0
    skipped_count = 0
    
    # Process each project
    for project_name in projects:
        print(f"{'='*60}")
        print(f"Project: {project_name}")
        print(f"{'='*60}")
        
        try:
            with DocumentDatabase() as db:
                config = load_project_config(project_name)
                
                # Check if we have data
                latest_data_df = db.get_latest_documents(project_name)
                if latest_data_df.empty:
                    print(f"✗ No data for {project_name}")
                    skipped_count += 1
                    continue
                
                # Generate requested report (all use dynamic counting now)
                if report_type == 'summary':
                    if generate_summary_report(project_name, config, output_dir, db):
                        success_count += 1
                    else:
                        fail_count += 1
                
                elif report_type == 'progression':
                    if generate_progression_report_full(project_name, config, output_dir, db):
                        success_count += 1
                    else:
                        fail_count += 1
                
                elif report_type == 'condensed':
                    if generate_condensed_report(project_name, config, output_dir, db, num_weeks=4):
                        success_count += 1
                    else:
                        skipped_count += 1
                
                elif report_type == 'certificates':
                    if generate_certificate_report_full(project_name, config, output_dir, db):
                        success_count += 1
                    else:
                        skipped_count += 1
        
        except Exception as e:
            print(f"✗ Error: {str(e)}")
            fail_count += 1
        
        print()  # Empty line between projects
    
    # Summary
    print(f"{'='*60}")
    print("GENERATION COMPLETE")
    print(f"{'='*60}")
    print(f"✓ Successful: {success_count}")
    if skipped_count > 0:
        print(f"ℹ Skipped: {skipped_count}")
    if fail_count > 0:
        print(f"✗ Failed: {fail_count}")
    print(f"{'='*60}")


def main():
    """Main function with interactive menu."""
    output_dir = Path('output')
    output_dir.mkdir(exist_ok=True)
    
    print("\n" + "="*60)
    print("Welcome to MBS Reports v2")
    print("Database-Driven Document Register Processing")
    print("="*60)
    
    while True:
        choice = show_menu()
        
        if choice == '1':
            # Generate all reports for all projects
            process_all_projects_all_reports()
            input("\nPress Enter to continue...")
        
        elif choice == '2':
            # Generate all reports for single project
            project_name = get_project_selection(include_all_option=False)
            
            if project_name:
                process_single_project_all_reports(project_name)
            else:
                print("\nCancelled")
            
            input("\nPress Enter to continue...")
        
        elif choice == '3':
            # Generate specific report type
            report_type = get_report_type_selection()
            
            if report_type:
                project_name = get_project_selection(include_all_option=True)
                
                if project_name:
                    if project_name == 'ALL':
                        generate_specific_report_for_projects(report_type, ['ALL'])
                    else:
                        generate_specific_report_for_projects(report_type, [project_name])
                else:
                    print("\nCancelled")
            else:
                print("\nCancelled")
            
            input("\nPress Enter to continue...")
        
        elif choice == '4':
            # Exit
            print("\nGoodbye!")
            break
        
        else:
            print("\n✗ Invalid choice. Please enter 1-4.")
            input("\nPress Enter to continue...")


if __name__ == '__main__':
    main()

