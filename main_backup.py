import os
import pandas as pd
from datetime import datetime
from pathlib import Path
import json
from config import *
import re
import time
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import openpyxl.styles
import argparse
from openpyxl.worksheet.page import PageMargins
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.fill import ColorChoice, PatternFillProperties
from openpyxl.drawing.colors import SchemeColor
import warnings

# Suppress openpyxl default style warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

# Overall Summary Sheet Style Configuration
OVERALL_SUMMARY_STYLES = {
    'title': {
        'font': Font(name='Calibri', size=14, bold=True, color='000000'),
        'alignment': Alignment(horizontal='left', vertical='center'),
        'fill': PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    },
    'timestamp': {
        'font': Font(name='Calibri', size=11, italic=True, color='000000'),
        'alignment': Alignment(horizontal='left', vertical='center'),
        'fill': PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    },
    'section_header': {
        'font': Font(name='Calibri', size=12, bold=True, color='000000'),
        'alignment': Alignment(horizontal='left', vertical='center'),
        'fill': PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
    },
    'column_header': {
        'font': Font(name='Calibri', size=11, bold=True, color='000000'),
        'alignment': Alignment(horizontal='left', vertical='center'),
        'fill': PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
    },
    'data_cell': {
        'font': Font(name='Calibri', size=11, color='000000'),
        'alignment': Alignment(horizontal='left', vertical='center'),
        'fill': PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    },
    'total_cell': {
        'font': Font(name='Calibri', size=11, bold=True, color='000000'),
        'alignment': Alignment(horizontal='left', vertical='center'),
        'fill': PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    },
    'border': Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
}

# Progression Report Status Configuration
PROGRESSION_STATUS_ORDER = {
    'Status A': {
        'display_name': 'Status A',
        'status_terms': [
            'A - Authorized and Accepted',
            'Accepted',
            'A',
            'A - Proceed',
            'A - Proceed (Lead Reviewer)'
        ]
    },
    'Status B': {
        'display_name': 'Status B',
        'status_terms': [
            'B - Partial Sign Off (with comment)',
            'Accepted with Comments',
            'B',
            'B - Proceed with Comments',
            'B - Proceed with Comments (Lead Reviewer)'
        ]
    },
    'Status C': {
        'display_name': 'Status C',
        'status_terms': [
            'Rejected',
            'QA - Rejected',
            'C - Rejected',
            'C',
            'C - Rejected (Lead Reviewer)',
            'C-Rejected'
        ]
    },
    'QC Rejected': {
        'display_name': 'QA/QC Rejected',
        'status_terms': [
            'QC Rejected',
            'QA Rejected'
        ]
    },
    'QC Checked': {
        'display_name': 'QC Checked',
        'status_terms': [
            'QC Checked',
            'QC Accepted'
        ]
    },
    'Under Review': {
        'display_name': 'Under Review/For Commenting',
        'status_terms': [
            'Under Review',
            'For Status Change',
            'For Commenting',
            'Awaiting QC Check',
            'Shared'
        ]
    },
    'Preliminary': {
        'display_name': 'Preliminary',
        'status_terms': [
            'Preliminary'
        ]
    },
    'Other': {
        'display_name': 'Other',
        'status_terms': [
            'Other'
        ]
    }
}

# Status-based conditional formatting
STATUS_STYLES = {
    'STATUS A': {
        'search_terms': ['A - Authorized and Accepted', 'Accepted', 'A', 'A - Proceed', 'A - Proceed (Lead Reviewer)', 'Status A'],
        'style': {
            'font': Font(name='Calibri', size=11, bold=True, color='000000'),
            'fill': PatternFill(start_color='25E82C', end_color='25E82C', fill_type='solid')
        }
    },
    'STATUS B': {
        'search_terms': ['B - Partial Sign Off (with comment)', 'Accepted with Comments', 'B', 'B - Proceed with Comments', 'B - Proceed with Comments (Lead Reviewer)', 'Status B'],
        'style': {
            'font': Font(name='Calibri', size=11, bold=True, color='000000'),
            'fill': PatternFill(start_color='EDDDA1', end_color='EDDDA1', fill_type='solid')
        }
    },
    'STATUS C': {
        'search_terms': ['Rejected', 'QA - Rejected', 'C - Rejected', 'QA Rejected', 'C', 'C - Rejected (Lead Reviewer)', 'C-Rejected', 'Status C'],
        'style': {
            'font': Font(name='Calibri', size=11, bold=True, color='000000'),
            'fill': PatternFill(start_color='ED1111', end_color='ED1111', fill_type='solid')
        }
    },
    'INFORMATION': {
        'search_terms': ['Information', 'Withdrawn-Obsolete', 'QA Passed', 'QA - Passed', 'For Status Change', 'For Commenting', 'Awaiting QC Check', 'QC Accepted', 'QC Checked', 'Under Review'],
        'style': {
            'font': Font(name='Calibri', size=11, bold=True, color='000000'),
            'fill': PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        }
    },
    'REVIEWED': {
        'search_terms': ['Reviewed'],
        'style': {
            'font': Font(name='Calibri', size=11, bold=True, color='000000'),
            'fill': PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        }
    },
    'PUBLISHED': {
        'search_terms': ['Published'],
        'style': {
            'font': Font(name='Calibri', size=11, bold=True, color='000000'),
            'fill': PatternFill(start_color='67DBB5', end_color='67DBB5', fill_type='solid')
        }
    },
    'SHARED': {
        'search_terms': ['Shared'],
        'style': {
            'font': Font(name='Calibri', size=11, bold=True, color='000000'),
            'fill': PatternFill(start_color='E0F090', end_color='E0F090', fill_type='solid')
        }
    },
    'PRELIMINARY': {
        'search_terms': ['Preliminary'],
        'style': {
            'font': Font(name='Calibri', size=11, bold=True, color='000000'),
            'fill': PatternFill(start_color='87CEEB', end_color='87CEEB', fill_type='solid')  # Light blue
        }
    },
    'OTHER': {
        'search_terms': ['Other'],
        'style': {
            'font': Font(name='Calibri', size=11, bold=True, color='000000'),
            'fill': PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')  # Light gray
        }
    },
    # 'SUBMITTED': {
    #     'search_terms': ['H - Submitted', 'H - For Review'],
    #     'style': {
    #         'font': Font(name='Calibri', size=11, bold=True, color='000000'),
    #         'fill': PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    #     }
    # },
    # 'COMPLETED': {
    #     'search_terms': ['I - Completed', 'I - Final'],
    #     'style': {
    #         'font': Font(name='Calibri', size=11, bold=True, color='000000'),
    #         'fill': PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    #     }
    # }
}

def apply_status_style(cell, status_name):
    """Apply conditional formatting based on status name"""
    for style_config in STATUS_STYLES.values():
        if any(term == status_name for term in style_config['search_terms']):
            cell.font = style_config['style']['font']
            cell.fill = style_config['style']['fill']
            return style_config['style']  # Return the style config for reuse
    # If no matching style found, use default data cell style
    cell.font = OVERALL_SUMMARY_STYLES['data_cell']['font']
    cell.fill = OVERALL_SUMMARY_STYLES['data_cell']['fill']
    return OVERALL_SUMMARY_STYLES['data_cell']  # Return default style

# Example of how to use these styles in the code:
# cell = overall_summary['A1']
# cell.font = OVERALL_SUMMARY_STYLES['title']['font']
# cell.alignment = OVERALL_SUMMARY_STYLES['title']['alignment']
# cell.fill = OVERALL_SUMMARY_STYLES['title']['fill']
# cell.border = OVERALL_SUMMARY_STYLES['border']

class DocumentRegisterProcessor:
    def __init__(self):
        self.data_dir = DATA_DIR
        self.reports_dir = REPORTS_DIR
        self.history_file = self.data_dir / "file_history.json"
        self.load_history()
        self.latest_changes = {}  # Store the latest changes

    def load_history(self):
        """Load the file processing history"""
        if self.history_file.exists():
            with open(self.history_file, 'r') as f:
                self.history = json.load(f)
        else:
            self.history = {}

    def save_history(self):
        """Save the file processing history"""
        with open(self.history_file, 'w') as f:
            json.dump(self.history, f, indent=2)

    def process_excel_file(self, file_path):
        """Process a single Excel file or CSV file and track changes"""
        # Skip temporary Excel files
        if file_path.name.startswith('~$'):
            print(f"Skipping temporary file: {file_path.name}")
            return None

        file_name = file_path.name
        current_hash = self._get_file_hash(file_path)
        
        try:
            # Determine file type and read accordingly
            file_path_str = str(file_path).lower()
            
            if file_path_str.endswith('.csv'):
                # Load project config to get CSV settings
                from config import load_project_config
                config = load_project_config(None, file_path)
                
                # Process CSV file
                df = process_csv_file(file_path, config)
                
                # Get timestamp from CSV
                date_str, time_str = get_file_timestamp(file_path)
                if date_str and time_str:
                    try:
                        file_date = datetime.strptime(date_str, '%d-%b-%Y')
                        file_time = datetime.strptime(time_str, '%H:%M').time()
                    except Exception as e:
                        print(f"Warning: Error parsing date from {file_name}: {str(e)}")
                        file_date = datetime.min
                else:
                    print(f"Warning: Could not parse timestamp from {file_name}")
                    file_date = datetime.min
            else:
                # Excel file
                df = pd.read_excel(file_path, **EXCEL_SETTINGS)
                
                # Get the timestamp from cell B4
                timestamp_df = pd.read_excel(file_path, usecols="B", nrows=4, header=None)
                timestamp_str = timestamp_df.iloc[3, 0]  # Get value from B4
                
                # Extract the date and time from the timestamp string
                try:
                    # Split by commas and get the third part (date and time)
                    parts = timestamp_str.split(',')
                    if len(parts) >= 3:
                        date_time_part = parts[2].strip()
                        # Split by space to separate date and time
                        date_time = date_time_part.split()
                        if len(date_time) >= 2:
                            file_date = datetime.strptime(date_time[0], '%d-%b-%Y')
                            file_time = datetime.strptime(date_time[1], '%H:%M').time()
                        else:
                            print(f"Warning: Could not parse date/time from {file_name}")
                            file_date = datetime.min
                    else:
                        print(f"Warning: Could not parse timestamp in {file_name}")
                        file_date = datetime.min
                except Exception as e:
                    print(f"Warning: Error parsing date from {file_name}: {str(e)}")
                    file_date = datetime.min
            
            # Save the processed data with timestamp
            self._save_processed_data(file_name, df, file_date)
            
            # Check if we have previous version
            if file_name in self.history:
                old_hash = self.history[file_name]['hash']
                if old_hash != current_hash:
                    changes = self._detect_changes(df, file_name)
                    self._update_history(file_name, current_hash, changes)
                    self.latest_changes[file_name] = changes  # Store the changes
                    return changes
            else:
                # First time processing this file
                self._update_history(file_name, current_hash, {'new_file': True})
                self.latest_changes[file_name] = {'new_file': True}
                return {'new_file': True}
                
        except Exception as e:
            print(f"Error processing file {file_name}: {str(e)}")
            return None

    def _save_processed_data(self, file_name, df, file_date):
        """Save the processed data for later use"""
        data_file = self.data_dir / f"{file_name}.parquet"
        # Save the dataframe with the file date as metadata
        df.to_parquet(data_file, engine='pyarrow')
        # Save the date separately
        date_file = self.data_dir / f"{file_name}.date"
        with open(date_file, 'w') as f:
            f.write(file_date.isoformat())

    def _get_file_hash(self, file_path):
        """Get a simple hash of the file for change detection"""
        return str(os.path.getmtime(file_path))

    def _detect_changes(self, current_df, file_name):
        """Detect changes between current and previous version"""
        try:
            # Load previous version
            prev_data_file = self.data_dir / f"{file_name}.parquet"
            if not prev_data_file.exists():
                return {'new_file': True}
            
            prev_df = pd.read_parquet(prev_data_file)
            
            # Find documents that exist in both versions
            common_docs = set(current_df['Doc Ref']).intersection(set(prev_df['Doc Ref']))
            
            changes = {
                'status_changes': [],
                'revision_changes': [],
                'date_changes': [],
                'new_documents': [],
                'removed_documents': []
            }
            
            # Check for new and removed documents
            current_docs = set(current_df['Doc Ref'])
            prev_docs = set(prev_df['Doc Ref'])
            
            changes['new_documents'] = list(current_docs - prev_docs)
            changes['removed_documents'] = list(prev_docs - current_docs)
            
            # Check for changes in existing documents
            for doc_ref in common_docs:
                current_doc = current_df[current_df['Doc Ref'] == doc_ref].iloc[0]
                prev_doc = prev_df[prev_df['Doc Ref'] == doc_ref].iloc[0]
                
                # Check status changes
                if current_doc['Status'] != prev_doc['Status']:
                    changes['status_changes'].append({
                        'doc_ref': doc_ref,
                        'doc_title': current_doc['Doc Title'],
                        'old_status': prev_doc['Status'],
                        'new_status': current_doc['Status']
                    })
                
                # Check revision changes
                if current_doc['Rev'] != prev_doc['Rev']:
                    changes['revision_changes'].append({
                        'doc_ref': doc_ref,
                        'doc_title': current_doc['Doc Title'],
                        'old_rev': prev_doc['Rev'],
                        'new_rev': current_doc['Rev']
                    })
                
                # Check date changes
                if current_doc['Date (WET)'] != prev_doc['Date (WET)']:
                    changes['date_changes'].append({
                        'doc_ref': doc_ref,
                        'doc_title': current_doc['Doc Title'],
                        'old_date': prev_doc['Date (WET)'],
                        'new_date': current_doc['Date (WET)']
                    })
            
            return changes
            
        except Exception as e:
            print(f"Error detecting changes: {str(e)}")
            return {'error': str(e)}

    def _update_history(self, file_name, file_hash, changes):
        """Update the processing history"""
        self.history[file_name] = {
            'last_processed': datetime.now().isoformat(),
            'hash': file_hash,
            'changes': changes
        }
        self.save_history()

def get_file_timestamp(file_path):
    """Get the timestamp from cell B4 of the Excel file or from CSV file"""
    try:
        file_path_str = str(file_path).lower()
        
        if file_path_str.endswith('.csv'):
            # For CSV files, get timestamp from 'Report Created' column
            df = pd.read_csv(file_path, nrows=1)
            if 'Report Created' in df.columns and not df['Report Created'].isna().all():
                timestamp_str = df['Report Created'].iloc[0]
                if pd.notna(timestamp_str):
                    # Parse the timestamp (format: "08-07-2025 07:03")
                    try:
                        # Split by space to separate date and time
                        date_part, time_part = timestamp_str.split(' ')
                        # Parse date (DD-MM-YYYY format)
                        from datetime import datetime
                        date_obj = datetime.strptime(date_part, '%d-%m-%Y')
                        time_obj = datetime.strptime(time_part, '%H:%M').time()
                        return date_obj.strftime('%d-%b-%Y'), time_obj.strftime('%H:%M')
                    except Exception as e:
                        print(f"Warning: Could not parse CSV timestamp '{timestamp_str}': {str(e)}")
                        return None, None
            return None, None
        else:
            # Excel file - Read just cell B4 (which is merged from B to I)
            timestamp_df = pd.read_excel(file_path, usecols="B", nrows=4, header=None)
            timestamp_str = timestamp_df.iloc[3, 0]
            
            # Split by commas and get the third part (date and time)
            parts = timestamp_str.split(',')
            if len(parts) >= 3:
                date_time_part = parts[2].strip()
                # Split by space to separate date and time
                date_time = date_time_part.split()
                if len(date_time) >= 2:
                    date_str = date_time[0]  # Keep as text
                    time_str = date_time[1]  # Keep as text
                    return date_str, time_str
            
            print(f"Warning: Could not parse timestamp from {file_path.name}")
            return None, None
    except Exception as e:
        print(f"Error reading timestamp from {file_path.name}: {str(e)}")
        return None, None

def clean_revision(val):
    if pd.isna(val):
        return ''
    s = str(val).replace('\u00A0', ' ').strip().upper()
    # Replace Cyrillic 'С' (U+0421) with Latin 'C'
    s = s.replace('\u0421', 'C')
    return s

def process_csv_file(file_path, config):
    """Process a CSV file and transform it to match expected format"""
    try:
        # Read the CSV file
        csv_settings = config.get('CSV_SETTINGS', {})
        df = pd.read_csv(file_path, **csv_settings)
        
        # Apply MBS filtering if enabled
        mbs_filter = config.get('MBS_FILTER')
        if mbs_filter and mbs_filter.get('enabled', False):
            filter_mask = pd.Series([False] * len(df), index=df.index)
            
            for column in mbs_filter.get('search_columns', []):
                if column in df.columns:
                    case_sensitive = mbs_filter.get('case_sensitive', False)
                    mask = df[column].str.contains('MBS', case=not case_sensitive, na=False)
                    filter_mask = filter_mask | mask
            
            df = df[filter_mask].copy()
            print(f"Filtered to {len(df)} MBS records")
        
        # Apply column mappings if provided
        column_mappings = config.get('COLUMN_MAPPINGS')
        if column_mappings:
            for target_col, source_col in column_mappings.items():
                if source_col in df.columns:
                    df[target_col] = df[source_col]
        
        # Apply custom status mapping for Holloway Park
        if config.get('PROJECT_TITLE') == 'Holloway Park':
            # Import the custom status mapping function
            try:
                from configs.HollowayPark import map_holloway_park_status
                if 'Status' in df.columns or 'Design Status' in df.columns:
                    # Apply the custom status mapping function to each row
                    df['Status'] = df.apply(map_holloway_park_status, axis=1)
                    print("Applied custom Holloway Park status mapping")
            except ImportError:
                print("Warning: Could not import Holloway Park status mapping function")
        
        # Clean revision column
        if 'Rev' in df.columns:
            df['Rev'] = df['Rev'].apply(clean_revision)
        
        return df
        
    except Exception as e:
        print(f"Error processing CSV file {file_path}: {str(e)}")
        raise

def get_counts(df, config=None):
    """Get counts of revisions and statuses from the dataframe"""
    counts = {}
    
    try:
        # Clean the Rev column
        if 'Rev' in df.columns:
            df['Rev'] = df['Rev'].apply(clean_revision)
        
        # Check if certificate separation is enabled
        cert_config = config.get('CERTIFICATE_SETTINGS', {}) if config else {}
        cert_enabled = cert_config.get('enabled', False)
        file_type_col = None
        cert_types = []
        
        if cert_enabled and config:
            file_type_settings = config.get('FILE_TYPE_SETTINGS', {})
            file_type_col = file_type_settings.get('column_name')
            cert_types = cert_config.get('certificate_types', [])
        
        # Separate certificate and non-certificate data if enabled
        if cert_enabled and file_type_col and file_type_col in df.columns:
            cert_data = df[df[file_type_col].isin(cert_types)]
            non_cert_data = df[~df[file_type_col].isin(cert_types)]
        else:
            cert_data = pd.DataFrame()
            non_cert_data = df
        
        # Count revisions
        if cert_enabled and not cert_data.empty:
            # Count all revisions first
            all_rev_counts = df['Rev'].value_counts()
            cert_rev_counts = cert_data['Rev'].value_counts()
            
            # For each revision, separate certificates from non-certificates
            for rev, total_count in all_rev_counts.items():
                cert_count = cert_rev_counts.get(rev, 0)
                non_cert_count = total_count - cert_count
                
                # Add non-certificate count (this will be the regular P01, P02, etc.)
                if non_cert_count > 0:
                    counts[f'Rev_{rev}'] = non_cert_count
            
            # Add total certificate count for P revisions only
            p_cert_total = 0
            for rev, cert_count in cert_rev_counts.items():
                if rev.startswith('P') and cert_count > 0:
                    p_cert_total += cert_count
            
            if p_cert_total > 0:
                counts['Rev_P_Certificates'] = p_cert_total
        else:
            # Regular revision counting
            rev_counts = df['Rev'].value_counts()
            for rev, count in rev_counts.items():
                counts[f'Rev_{rev}'] = count
        
        # Count statuses
        if cert_enabled and not cert_data.empty:
            # Count non-certificate statuses only
            non_cert_status_counts = non_cert_data['Status'].value_counts()
            for status, count in non_cert_status_counts.items():
                counts[f'Status_{status}'] = count
            
            # Count certificate statuses with suffix (separate from regular statuses)
            cert_status_counts = cert_data['Status'].value_counts()
            cert_suffix = cert_config.get('status_suffix', ' (Certificates)')
            for status, count in cert_status_counts.items():
                counts[f'Status_{status}{cert_suffix}'] = count
        else:
            # Regular status counting
            status_counts = df['Status'].value_counts()
            for status, count in status_counts.items():
                counts[f'Status_{status}'] = count
        
        # Count file types if the column exists
        if 'OVL - File Type' in df.columns:
            file_type_counts = df['OVL - File Type'].value_counts()
            for file_type, count in file_type_counts.items():
                counts[f'FileType_{file_type}'] = count
        elif 'Form' in df.columns:
            file_type_counts = df['Form'].value_counts()
            for file_type, count in file_type_counts.items():
                counts[f'FileType_{file_type}'] = count
        elif 'File Type' in df.columns:
            file_type_counts = df['File Type'].value_counts()
            for file_type, count in file_type_counts.items():
                counts[f'FileType_{file_type}'] = count
        
        return counts
    except Exception as e:
        print(f"Error in get_counts: {str(e)}")
        print("DataFrame columns:", df.columns.tolist())
        print("DataFrame head:")
        print(df.head())
        raise

def load_processed_files():
    """Load the record of processed files"""
    try:
        with open('processed_files.json', 'r') as f:
            return json.load(f)
    except FileNotFoundError:
        return {}

def save_processed_files(processed_files):
    """Save the record of processed files"""
    with open('processed_files.json', 'w') as f:
        json.dump(processed_files, f, indent=2, default=str)

def compare_values(current_val, prev_val, col_name):
    """Compare values and return True if they are actually different"""
    # Convert both values to strings, handling NaN/None
    current_str = str(current_val).strip() if pd.notna(current_val) else ''
    prev_str = str(prev_val).strip() if pd.notna(prev_val) else ''
    
    # If both are empty, no change
    if not current_str and not prev_str:
        return False
    
    # If one is empty and the other isn't, that's a change
    if not current_str or not prev_str:
        return True
    
    # For date fields, normalize the format
    if 'Date' in col_name or 'WET' in col_name:
        try:
            # Try to parse and normalize the dates
            current_date = pd.to_datetime(current_str)
            prev_date = pd.to_datetime(prev_str)
            # Compare the normalized date strings
            current_str = current_date.strftime('%d-%b-%Y')
            prev_str = prev_date.strftime('%d-%b-%Y')
        except:
            # If date parsing fails, use the original strings
            pass
    
    # Compare the normalized strings
    return current_str != prev_str

def save_excel_with_retry(summary_df, changes_df, latest_data_df, output_file, config, max_retries=3):
    """Try to save the Excel file with retries"""
    for attempt in range(max_retries):
        try:
            # Try to load existing file
            try:
                book = load_workbook(output_file)
                # Remove existing Overall Summary sheet if it exists (pandas won't replace it)
                if 'Overall Summary' in book.sheetnames:
                    book.remove(book['Overall Summary'])
                    book.save(output_file)  # Save the removal
                # Remove existing Summary Data sheet if it exists
                if 'Summary Data' in book.sheetnames:
                    book.remove(book['Summary Data'])
                # Remove existing Latest Data sheet if it exists
                if 'Latest Data' in book.sheetnames:
                    book.remove(book['Latest Data'])
                with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    summary_df.to_excel(writer, sheet_name='Summary Data', index=False)
                    latest_data_df.to_excel(writer, sheet_name='Latest Data', index=False)
            except FileNotFoundError:
                # If file doesn't exist, create new one
                with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                    summary_df.to_excel(writer, sheet_name='Summary Data', index=False)
                    latest_data_df.to_excel(writer, sheet_name='Latest Data', index=False)
            # Create or update Overall Summary sheet first
            book = load_workbook(output_file)
            if 'Overall Summary' in book.sheetnames:
                book.remove(book['Overall Summary'])
            overall_summary = book.create_sheet('Overall Summary', 0)  # Create at index 0 (first position)
            
            # Set print layout: fit to 1 page, center, narrow margins
            overall_summary.page_setup.fitToWidth = 1
            overall_summary.page_setup.fitToHeight = 0
            overall_summary.page_setup.horizontalCentered = True
            overall_summary.page_setup.verticalCentered = True
            overall_summary.page_margins = PageMargins(left=0.25, right=0.25, top=0.75, bottom=0.75, header=0.3, footer=0.3)
            # Extra reliability for centering
            overall_summary.sheet_properties.pageSetUpPr.horizontalCentered = True
            overall_summary.sheet_properties.pageSetUpPr.verticalCentered = True
            
            # Merge and center the title across A1:O1
            overall_summary.merge_cells('A1:O1')
            
            # Set row height for the title row
            overall_summary.row_dimensions[1].height = 70

            # Add title with project name on first line and rest on second line
            project_title = config.get('PROJECT_TITLE', '')
            if project_title:
                title_text = f"{project_title}\nDocument Register Overall Summary"
            else:
                title_text = "Document Register Overall Summary"
            overall_summary['A1'] = title_text
            
            # Style exactly like the working test script
            overall_summary['A1'].font = Font(name='Calibri', size=14, bold=True, color='000000')
            overall_summary['A1'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            overall_summary['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Get the latest data from Summary Data
            summary_data = pd.read_excel(output_file, sheet_name='Summary Data')
            latest_row = summary_data.iloc[-1]  # Get the last row
            
            # Add data export timestamp from Summary Data (now in row 2)
            export_date = latest_row['Date'].strftime("%d-%m-%Y") if isinstance(latest_row['Date'], datetime) else latest_row['Date']
            export_time = latest_row['Time'].strftime("%H-%M-%S") if isinstance(latest_row['Time'], datetime) else latest_row['Time']
            overall_summary['A2'] = f'Data Export: {export_date} {export_time}'
            overall_summary['A2'].font = OVERALL_SUMMARY_STYLES['timestamp']['font']
            overall_summary['A2'].alignment = OVERALL_SUMMARY_STYLES['timestamp']['alignment']
            overall_summary['A2'].fill = OVERALL_SUMMARY_STYLES['timestamp']['fill']
            
            # Get all column names from Summary Data
            all_columns = summary_data.columns.tolist()
            
            # Add total documents
            overall_summary['A5'] = 'Total Documents:'
            # Get total from Latest Data sheet
            latest_data = pd.read_excel(output_file, sheet_name='Latest Data')
            total_docs = len(latest_data)  # len() already gives us the correct count
            overall_summary['B5'] = total_docs
            overall_summary['B5'].font = OVERALL_SUMMARY_STYLES['total_cell']['font']
            overall_summary['B5'].alignment = OVERALL_SUMMARY_STYLES['total_cell']['alignment']
            overall_summary['B5'].fill = OVERALL_SUMMARY_STYLES['total_cell']['fill']
            
            # Filter and sort revision columns
            rev_columns = [col for col in all_columns if col.startswith('Rev_')]
            
            # Check if certificate separation is enabled
            cert_config = config.get('CERTIFICATE_SETTINGS', {})
            cert_enabled = cert_config.get('enabled', False)
            
            # Group revisions by type
            p_revs = sorted([col for col in rev_columns if col.startswith('Rev_P')], 
                          key=lambda x: int(x.split('_')[1][1:]) if x.split('_')[1][1:].isdigit() else float('inf'))
            
            # Add certificates to P revisions if enabled and present
            if cert_enabled and 'Rev_P_Certificates' in rev_columns:
                p_revs.append('Rev_P_Certificates')
            
            c_revs = sorted([col for col in rev_columns if col.startswith('Rev_C')],
                          key=lambda x: int(x.split('_')[1][1:]) if x.split('_')[1][1:].isdigit() else float('inf'))
            
            # Filter out certificates from other_revs if it's being handled with P revisions
            other_revs = sorted([col for col in rev_columns if not (
                col.startswith('Rev_P') or 
                col.startswith('Rev_C') or 
                (cert_enabled and col == 'Rev_P_Certificates')
            )])
            
            # Function to add revision and status summary
            def add_revision_summary(start_row, rev_columns, title):
                """Add a revision summary section with a combined status summary"""
                # Add section header
                overall_summary[f'A{start_row}'] = title
                overall_summary[f'A{start_row}'].font = OVERALL_SUMMARY_STYLES['section_header']['font']
                overall_summary[f'A{start_row}'].alignment = OVERALL_SUMMARY_STYLES['section_header']['alignment']
                overall_summary[f'A{start_row}'].fill = OVERALL_SUMMARY_STYLES['section_header']['fill']
                
                # Add revision headers
                overall_summary[f'A{start_row + 1}'] = 'Revision'
                overall_summary[f'B{start_row + 1}'] = 'Count'
                overall_summary[f'C{start_row + 1}'] = 'Status'
                overall_summary[f'D{start_row + 1}'] = 'Count'
                
                # Style headers
                for col in ['A', 'B', 'C', 'D']:
                    overall_summary[f'{col}{start_row + 1}'].font = OVERALL_SUMMARY_STYLES['column_header']['font']
                    overall_summary[f'{col}{start_row + 1}'].alignment = OVERALL_SUMMARY_STYLES['column_header']['alignment']
                    overall_summary[f'{col}{start_row + 1}'].fill = OVERALL_SUMMARY_STYLES['column_header']['fill']
                    overall_summary[f'{col}{start_row + 1}'].border = OVERALL_SUMMARY_STYLES['border']
                
                # Check if this is P revision summary and certificates are enabled
                is_p_revision = title == 'P Revision Summary'
                cert_config = config.get('CERTIFICATE_SETTINGS', {})
                cert_enabled = cert_config.get('enabled', False) and is_p_revision
                
                # Get file type column name
                file_type_col = None
                if cert_enabled:
                    file_type_settings = config.get('FILE_TYPE_SETTINGS', {})
                    file_type_col = file_type_settings.get('column_name')
                
                # Separate certificates from regular documents
                if cert_enabled and file_type_col and file_type_col in latest_data.columns:
                    cert_types = cert_config.get('certificate_types', [])
                    # Filter data for certificates and non-certificates
                    cert_data = latest_data[latest_data[file_type_col].isin(cert_types)]
                    non_cert_data = latest_data[~latest_data[file_type_col].isin(cert_types)]
                else:
                    cert_data = pd.DataFrame()
                    non_cert_data = latest_data
                
                # Get status counts for all revisions in this group (excluding certificates for P revisions)
                status_counts = {}
                total_count = 0
                cert_total_count = 0
                cert_status_counts = {}
                
                for rev_col in rev_columns:
                    rev_name = rev_col.replace('Rev_', '')
                    
                    # Skip certificate column here - we'll handle it separately
                    if rev_col == 'Rev_P_Certificates':
                        cert_total_count += latest_row.get(rev_col, 0)
                        # Certificate statuses are already counted in the Summary Data table with suffix
                        # We'll get them from the latest_row data instead of recounting
                        continue
                    
                    # For regular revisions, use the separated count from summary data
                    count = latest_row.get(rev_col, 0)
                    total_count += count
                    
                    # Count statuses for this revision (excluding certificates if separation is enabled)
                    if cert_enabled and file_type_col:
                        # Count statuses for non-certificate documents only
                        rev_data = non_cert_data[non_cert_data['Rev'] == rev_name]
                    else:
                        # Count statuses for all documents of this revision
                        rev_data = latest_data[latest_data['Rev'] == rev_name]
                    
                    for status, status_count in rev_data['Status'].value_counts().items():
                        status_counts[status] = status_counts.get(status, 0) + status_count
                
                # Add revision data
                row = start_row + 2
                for rev_col in rev_columns:
                    # Handle certificates specially
                    if rev_col == 'Rev_P_Certificates':
                        # Skip the regular processing for certificates - we'll handle it separately
                        continue
                    
                    rev_name = rev_col.replace('Rev_', '')
                    
                    # Always use the count from the summary data (which already has certificates separated)
                    count = latest_row.get(rev_col, 0)
                    
                    # Add revision name and count
                    overall_summary[f'A{row}'] = rev_name
                    overall_summary[f'B{row}'] = count
                    overall_summary[f'A{row}'].font = OVERALL_SUMMARY_STYLES['data_cell']['font']
                    overall_summary[f'A{row}'].alignment = OVERALL_SUMMARY_STYLES['data_cell']['alignment']
                    overall_summary[f'A{row}'].fill = OVERALL_SUMMARY_STYLES['data_cell']['fill']
                    overall_summary[f'A{row}'].border = OVERALL_SUMMARY_STYLES['border']
                    overall_summary[f'B{row}'].font = OVERALL_SUMMARY_STYLES['data_cell']['font']
                    overall_summary[f'B{row}'].alignment = OVERALL_SUMMARY_STYLES['data_cell']['alignment']
                    overall_summary[f'B{row}'].fill = OVERALL_SUMMARY_STYLES['data_cell']['fill']
                    overall_summary[f'B{row}'].border = OVERALL_SUMMARY_STYLES['border']
                    row += 1
                
                # Add certificate summary row if enabled and Rev_Certificates column exists
                if cert_enabled and 'Rev_P_Certificates' in rev_columns:
                    cert_count = latest_row.get('Rev_P_Certificates', 0)
                    cert_label = cert_config.get('summary_label', 'P01-PXX (Certificates)')
                    overall_summary[f'A{row}'] = cert_label
                    overall_summary[f'B{row}'] = cert_count
                    overall_summary[f'A{row}'].font = OVERALL_SUMMARY_STYLES['data_cell']['font']
                    overall_summary[f'A{row}'].alignment = OVERALL_SUMMARY_STYLES['data_cell']['alignment']
                    overall_summary[f'A{row}'].fill = OVERALL_SUMMARY_STYLES['data_cell']['fill']
                    overall_summary[f'A{row}'].border = OVERALL_SUMMARY_STYLES['border']
                    overall_summary[f'B{row}'].font = OVERALL_SUMMARY_STYLES['data_cell']['font']
                    overall_summary[f'B{row}'].alignment = OVERALL_SUMMARY_STYLES['data_cell']['alignment']
                    overall_summary[f'B{row}'].fill = OVERALL_SUMMARY_STYLES['data_cell']['fill']
                    overall_summary[f'B{row}'].border = OVERALL_SUMMARY_STYLES['border']
                    row += 1
                    # Update cert_total_count for the total calculation
                    cert_total_count = cert_count
                
                # Add total row for revisions
                overall_summary[f'A{row}'] = 'Total'
                overall_summary[f'B{row}'] = total_count + (cert_total_count if cert_enabled else 0)
                overall_summary[f'A{row}'].font = OVERALL_SUMMARY_STYLES['total_cell']['font']
                overall_summary[f'A{row}'].alignment = OVERALL_SUMMARY_STYLES['total_cell']['alignment']
                overall_summary[f'A{row}'].fill = OVERALL_SUMMARY_STYLES['total_cell']['fill']
                overall_summary[f'A{row}'].border = OVERALL_SUMMARY_STYLES['border']
                overall_summary[f'B{row}'].font = OVERALL_SUMMARY_STYLES['total_cell']['font']
                overall_summary[f'B{row}'].alignment = OVERALL_SUMMARY_STYLES['total_cell']['alignment']
                overall_summary[f'B{row}'].fill = OVERALL_SUMMARY_STYLES['total_cell']['fill']
                overall_summary[f'B{row}'].border = OVERALL_SUMMARY_STYLES['border']
                
                # Add status summary
                status_row = start_row + 2
                total_status_count = 0
                
                # Create ordered list with Published first, then others
                ordered_statuses = []
                published_statuses = []
                status_a_statuses = []
                status_b_statuses = []
                status_c_statuses = []
                other_statuses = []
                
                # Also create certificate status lists if enabled
                cert_published_statuses = []
                cert_status_a_statuses = []
                cert_status_b_statuses = []
                cert_status_c_statuses = []
                cert_other_statuses = []
                
                # Categorize statuses
                for status, count in status_counts.items():
                    categorized = False
                    for style_name, style_config in STATUS_STYLES.items():
                        if any(term == status for term in style_config['search_terms']):
                            if style_name == 'PUBLISHED':
                                published_statuses.append((status, count))
                            elif style_name == 'STATUS A':
                                status_a_statuses.append((status, count))
                            elif style_name == 'STATUS B':
                                status_b_statuses.append((status, count))
                            elif style_name == 'STATUS C':
                                status_c_statuses.append((status, count))
                            else:
                                other_statuses.append((status, count))
                            categorized = True
                            break
                    if not categorized:
                        other_statuses.append((status, count))
                
                # Categorize certificate statuses if enabled
                if cert_enabled:
                    # Get certificate status counts from Summary Data table (they have suffix)
                    cert_suffix = cert_config.get('status_suffix', ' (Certificates)')
                    for col_name in summary_data.columns:
                        if col_name.startswith('Status_') and col_name.endswith(cert_suffix):
                            # Extract the base status name and get the count
                            status = col_name.replace('Status_', '').replace(cert_suffix, '')
                            count = latest_row.get(col_name, 0)
                            if count > 0:
                                categorized = False
                                for style_name, style_config in STATUS_STYLES.items():
                                    if any(term == status for term in style_config['search_terms']):
                                        if style_name == 'PUBLISHED':
                                            cert_published_statuses.append((status, count))
                                        elif style_name == 'STATUS A':
                                            cert_status_a_statuses.append((status, count))
                                        elif style_name == 'STATUS B':
                                            cert_status_b_statuses.append((status, count))
                                        elif style_name == 'STATUS C':
                                            cert_status_c_statuses.append((status, count))
                                        else:
                                            cert_other_statuses.append((status, count))
                                        categorized = True
                                        break
                                if not categorized:
                                    cert_other_statuses.append((status, count))
                
                # Build ordered list: Published first, then A, B, C, Other
                ordered_statuses = (sorted(published_statuses) + 
                                  sorted(status_a_statuses) + 
                                  sorted(status_b_statuses) + 
                                  sorted(status_c_statuses) + 
                                  sorted(other_statuses))
                
                # Add certificate statuses with suffix if enabled
                if cert_enabled and (cert_published_statuses or cert_status_a_statuses or cert_status_b_statuses or cert_status_c_statuses or cert_other_statuses):
                    cert_suffix = cert_config.get('status_suffix', ' (Certificates)')
                    cert_ordered_statuses = (sorted(cert_published_statuses) + 
                                           sorted(cert_status_a_statuses) + 
                                           sorted(cert_status_b_statuses) + 
                                           sorted(cert_status_c_statuses) + 
                                           sorted(cert_other_statuses))
                    
                    # Add certificate statuses with suffix to the main list
                    for status, count in cert_ordered_statuses:
                        ordered_statuses.append((f"{status}{cert_suffix}", count))
                
                for status, count in ordered_statuses:
                    total_status_count += count
                    
                    # Add status name with conditional formatting
                    status_cell = overall_summary[f'C{status_row}']
                    status_cell.value = status
                    # Remove certificate suffix for style matching
                    status_for_style = status.replace(cert_config.get('status_suffix', ' (Certificates)'), '') if cert_enabled else status
                    style = apply_status_style(status_cell, status_for_style)
                    status_cell.alignment = OVERALL_SUMMARY_STYLES['data_cell']['alignment']
                    status_cell.border = OVERALL_SUMMARY_STYLES['border']
                    
                    # Add count with matching style
                    count_cell = overall_summary[f'D{status_row}']
                    count_cell.value = count
                    count_cell.font = Font(
                        name=style['font'].name,
                        size=style['font'].size,
                        bold=style['font'].bold,
                        italic=style['font'].italic,
                        color=style['font'].color
                    )
                    count_cell.fill = style['fill']
                    count_cell.alignment = OVERALL_SUMMARY_STYLES['data_cell']['alignment']
                    count_cell.border = OVERALL_SUMMARY_STYLES['border']
                    
                    status_row += 1
                
                # Add total row for status counts
                overall_summary[f'C{status_row}'] = 'Total'
                overall_summary[f'D{status_row}'] = total_status_count
                overall_summary[f'C{status_row}'].font = OVERALL_SUMMARY_STYLES['total_cell']['font']
                overall_summary[f'C{status_row}'].alignment = OVERALL_SUMMARY_STYLES['total_cell']['alignment']
                overall_summary[f'C{status_row}'].fill = OVERALL_SUMMARY_STYLES['total_cell']['fill']
                overall_summary[f'C{status_row}'].border = OVERALL_SUMMARY_STYLES['border']
                overall_summary[f'D{status_row}'].font = OVERALL_SUMMARY_STYLES['total_cell']['font']
                overall_summary[f'D{status_row}'].alignment = OVERALL_SUMMARY_STYLES['total_cell']['alignment']
                overall_summary[f'D{status_row}'].fill = OVERALL_SUMMARY_STYLES['total_cell']['fill']
                overall_summary[f'D{status_row}'].border = OVERALL_SUMMARY_STYLES['border']
                
                return max(row, status_row) + 2  # Return the next row to start from
            
            # Add P revision summary
            current_row = 7
            current_row = add_revision_summary(current_row, p_revs, 'P Revision Summary')
            
            # Add C revision summary
            current_row = add_revision_summary(current_row, c_revs, 'C Revision Summary')
            
            # Add other revision summary
            current_row = add_revision_summary(current_row, other_revs, 'Other Revision Summary')
            
            # Initialize row variable for border formatting
            row = current_row
            
            # Add file type summary section if enabled in config
            if config.get('FILE_TYPE_SETTINGS', {}).get('include_in_summary', False):
                file_type_col = config['FILE_TYPE_SETTINGS']['column_name']
                if file_type_col in latest_data.columns:
                    # Add file type summary section
                    file_type_start_row = current_row + 2
                    overall_summary[f'A{file_type_start_row}'] = config['FILE_TYPE_SETTINGS']['summary_title']
                    overall_summary[f'A{file_type_start_row}'].font = OVERALL_SUMMARY_STYLES['section_header']['font']
                    overall_summary[f'A{file_type_start_row}'].alignment = OVERALL_SUMMARY_STYLES['section_header']['alignment']
                    overall_summary[f'A{file_type_start_row}'].fill = OVERALL_SUMMARY_STYLES['section_header']['fill']
                    
                    # Add file type headers
                    overall_summary[f'A{file_type_start_row + 1}'] = 'File Type'
                    overall_summary[f'B{file_type_start_row + 1}'] = 'Count'
                    overall_summary[f'A{file_type_start_row + 1}'].font = OVERALL_SUMMARY_STYLES['column_header']['font']
                    overall_summary[f'A{file_type_start_row + 1}'].alignment = OVERALL_SUMMARY_STYLES['column_header']['alignment']
                    overall_summary[f'A{file_type_start_row + 1}'].fill = OVERALL_SUMMARY_STYLES['column_header']['fill']
                    overall_summary[f'B{file_type_start_row + 1}'].font = OVERALL_SUMMARY_STYLES['column_header']['font']
                    overall_summary[f'B{file_type_start_row + 1}'].alignment = OVERALL_SUMMARY_STYLES['column_header']['alignment']
                    overall_summary[f'B{file_type_start_row + 1}'].fill = OVERALL_SUMMARY_STYLES['column_header']['fill']
                    
                    # Add file type data
                    row = file_type_start_row + 2
                    file_type_counts = latest_data[file_type_col].value_counts()
                    total_file_types = 0
                    for file_type, count in file_type_counts.items():
                        total_file_types += count
                        overall_summary[f'A{row}'] = file_type
                        overall_summary[f'B{row}'] = count
                        overall_summary[f'A{row}'].font = OVERALL_SUMMARY_STYLES['data_cell']['font']
                        overall_summary[f'A{row}'].alignment = OVERALL_SUMMARY_STYLES['data_cell']['alignment']
                        overall_summary[f'A{row}'].fill = OVERALL_SUMMARY_STYLES['data_cell']['fill']
                        overall_summary[f'A{row}'].border = OVERALL_SUMMARY_STYLES['border']
                        overall_summary[f'B{row}'].font = OVERALL_SUMMARY_STYLES['data_cell']['font']
                        overall_summary[f'B{row}'].alignment = OVERALL_SUMMARY_STYLES['data_cell']['alignment']
                        overall_summary[f'B{row}'].fill = OVERALL_SUMMARY_STYLES['data_cell']['fill']
                        overall_summary[f'B{row}'].border = OVERALL_SUMMARY_STYLES['border']
                        row += 1
                    
                    # Add total row for file types
                    overall_summary[f'A{row}'] = 'Total'
                    overall_summary[f'B{row}'] = total_file_types
                    overall_summary[f'A{row}'].font = OVERALL_SUMMARY_STYLES['total_cell']['font']
                    overall_summary[f'A{row}'].alignment = OVERALL_SUMMARY_STYLES['total_cell']['alignment']
                    overall_summary[f'A{row}'].fill = OVERALL_SUMMARY_STYLES['total_cell']['fill']
                    overall_summary[f'A{row}'].border = OVERALL_SUMMARY_STYLES['border']
                    overall_summary[f'B{row}'].font = OVERALL_SUMMARY_STYLES['total_cell']['font']
                    overall_summary[f'B{row}'].alignment = OVERALL_SUMMARY_STYLES['total_cell']['alignment']
                    overall_summary[f'B{row}'].fill = OVERALL_SUMMARY_STYLES['total_cell']['fill']
                    overall_summary[f'B{row}'].border = OVERALL_SUMMARY_STYLES['border']
            
            # Add borders and formatting
            for row_num in range(2, row + 1):  # Start from row 2 to skip the title row
                for col in ['A', 'B', 'C', 'D']:
                    cell = overall_summary[f'{col}{row_num}']
                    cell.border = OVERALL_SUMMARY_STYLES['border']
                    cell.alignment = OVERALL_SUMMARY_STYLES['data_cell']['alignment']
            
            # Create pie charts for P and C revision statuses
            chart_start_col = 'G'  # Start charts in column G to give more space
            
            # Function to create a pie chart for revision status data
            def create_status_pie_chart(revision_type, chart_title, chart_position):
                """Create a pie chart for status distribution of a specific revision type"""
                # Get status data for this revision type
                if revision_type == 'P':
                    # Check if certificate separation is enabled for P revisions
                    cert_config = config.get('CERTIFICATE_SETTINGS', {})
                    cert_enabled = cert_config.get('enabled', False)
                    
                    if cert_enabled:
                        # Get file type column and exclude certificates
                        file_type_settings = config.get('FILE_TYPE_SETTINGS', {})
                        file_type_col = file_type_settings.get('column_name')
                        
                        if file_type_col and file_type_col in latest_data.columns:
                            cert_types = cert_config.get('certificate_types', [])
                            # Exclude certificates from P revision chart
                            rev_data = latest_data[
                                (latest_data['Rev'].str.startswith('P')) & 
                                (~latest_data[file_type_col].isin(cert_types))
                            ]
                        else:
                            rev_data = latest_data[latest_data['Rev'].str.startswith('P')]
                    else:
                        rev_data = latest_data[latest_data['Rev'].str.startswith('P')]
                elif revision_type == 'C':
                    rev_data = latest_data[latest_data['Rev'].str.startswith('C')]
                else:
                    return None
                
                if rev_data.empty:
                    return None
                
                # Count statuses
                status_counts = rev_data['Status'].value_counts()
                
                if len(status_counts) == 0:
                    return None
                
                # Group statuses according to STATUS_STYLES categories
                grouped_counts = {
                    'Published': 0,     # Published gets its own category but will be grouped with Status A for charts
                    'Status A': 0,
                    'Status B': 0, 
                    'Status C': 0,
                    'Other': 0
                }
                
                for status, count in status_counts.items():
                    # Check which category this status belongs to
                    categorized = False
                    for style_name, style_config in STATUS_STYLES.items():
                        if any(term == status for term in style_config['search_terms']):
                            if style_name == 'PUBLISHED':
                                grouped_counts['Published'] += count
                            if style_name == 'STATUS A':
                                grouped_counts['Status A'] += count
                            elif style_name == 'STATUS B':
                                grouped_counts['Status B'] += count
                            elif style_name == 'STATUS C':
                                grouped_counts['Status C'] += count
                            else:
                                if style_name != 'PUBLISHED':  # Don't double-count Published
                                    grouped_counts['Other'] += count
                            categorized = True
                            break
                    
                    # If not categorized, add to Other
                    if not categorized:
                        grouped_counts['Other'] += count
                
                # For chart purposes, combine Published with Status A
                chart_grouped_counts = grouped_counts.copy()
                if chart_grouped_counts['Published'] > 0:
                    chart_grouped_counts['Status A'] += chart_grouped_counts['Published']
                    del chart_grouped_counts['Published']
                
                # Remove categories with zero counts
                chart_grouped_counts = {k: v for k, v in chart_grouped_counts.items() if v > 0}
                
                if not chart_grouped_counts:
                    return None
                
                # Find a good place to put chart data (in a hidden area)
                data_start_row = 100  # Use row 100+ to keep chart data out of the way
                chart_col_offset = 0 if revision_type == 'P' else 4  # Offset C chart data more
                
                # Add chart data to hidden area
                data_row = data_start_row
                chart_labels_with_counts = []
                for category, count in chart_grouped_counts.items():
                    label_with_count = f"{category} ({count})"
                    chart_labels_with_counts.append(label_with_count)
                    overall_summary[f'{chr(ord(chart_start_col) + chart_col_offset)}{data_row}'] = label_with_count
                    overall_summary[f'{chr(ord(chart_start_col) + chart_col_offset + 1)}{data_row}'] = count
                    data_row += 1
                
                # Create pie chart
                chart = PieChart()
                chart.title = chart_title
                chart.width = 16  # Make charts wider
                chart.height = 12  # Make charts taller
                
                # Define data range from hidden area
                labels = Reference(overall_summary, 
                                 min_col=ord(chart_start_col) - ord('A') + 1 + chart_col_offset, 
                                 min_row=data_start_row, 
                                 max_row=data_row - 1)
                data = Reference(overall_summary, 
                               min_col=ord(chart_start_col) - ord('A') + 2 + chart_col_offset, 
                               min_row=data_start_row, 
                               max_row=data_row - 1)
                
                chart.add_data(data, titles_from_data=False)
                chart.set_categories(labels)
                
                # Configure data labels to show only percentages (not "Series 1")
                chart.dataLabels = DataLabelList()
                chart.dataLabels.showCatName = False  # Don't show category name on slices
                chart.dataLabels.showVal = False      # Don't show values on slices
                chart.dataLabels.showPercent = True
                chart.dataLabels.showSerName = False  # This should remove "Series 1"
                
                # Set legend to show our custom labels with counts
                chart.legend.position = 'r'  # Position legend to the right
                chart.legend.layout = None  # Let Excel auto-position the legend
                
                # Style the chart colors to match STATUS_STYLES
                # Define colors for each category
                colors = {
                    'Status A': '00B050',  # Green
                    'Status B': 'EDDDA1',  # Beige (matching STATUS_STYLES)
                    'Status C': 'ED1111',  # Red
                    'Other': '808080'      # Grey
                }
                
                # Apply colors to chart slices
                try:
                    if chart.series:
                        series = chart.series[0]
                        for i, (category, count) in enumerate(chart_grouped_counts.items()):
                            # Extract the category name without the count
                            category_name = category.split(' (')[0]
                            if category_name in colors:
                                # Create a data point with the specified color
                                pt = DataPoint(idx=i)
                                # Set the fill color
                                pt.graphicalProperties.solidFill = colors[category_name]
                                series.dPt.append(pt)
                except Exception as e:
                    # If coloring fails, continue without custom colors
                    print(f"Warning: Could not apply chart colors: {e}")
                
                # Alternative approach using series formatting
                try:
                    if chart.series and len(chart.series) > 0:
                        series = chart.series[0]
                        # Clear any existing data points
                        series.dPt = []
                        
                        for i, (category, count) in enumerate(chart_grouped_counts.items()):
                            category_name = category.split(' (')[0]
                            if category_name in colors:
                                # Create data point
                                pt = DataPoint(idx=i)
                                # Create fill properties
                                fill = PatternFillProperties()
                                fill.solidFill = ColorChoice(srgbClr=colors[category_name])
                                pt.graphicalProperties.solidFill = fill.solidFill
                                series.dPt.append(pt)
                except Exception as e:
                    # If this approach fails too, try a simpler method
                    try:
                        for i, (category, count) in enumerate(chart_grouped_counts.items()):
                            category_name = category.split(' (')[0]
                            if category_name in colors and chart.series:
                                pt = DataPoint(idx=i)
                                # Simple color assignment
                                pt.graphicalProperties.solidFill = colors[category_name]
                                if not hasattr(chart.series[0], 'dPt') or chart.series[0].dPt is None:
                                    chart.series[0].dPt = []
                                chart.series[0].dPt.append(pt)
                    except Exception as e2:
                        print(f"Warning: All chart coloring methods failed: {e2}")
                        pass
                
                # Position the chart
                overall_summary.add_chart(chart, chart_position)
                
                return data_row
            
            # Create P revision status chart - position it lower to give more space from title
            create_status_pie_chart('P', 'P Revision Status Distribution', 'G2')
            
            # Create C revision status chart - position it below P chart with spacing
            create_status_pie_chart('C', 'C Revision Status Distribution', 'G27')
            
            # Auto-adjust column widths for Overall Summary
            for column in overall_summary.columns:
                max_length = 0
                column_letter = None
                for cell in column:
                    if hasattr(cell, 'column_letter'):
                        column_letter = cell.column_letter
                        break
                
                if not column_letter:
                    continue  # Skip if we can't get column letter
                
                # Skip the first row for merged columns in Overall Summary
                start_row = 2 if column_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O'] else 1
                for cell in column[start_row-1:]:
                    try:
                        # Skip merged cells
                        if hasattr(cell, 'value') and cell.value is not None:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                    except:
                        pass
                # Set minimum width of 8 for empty/narrow columns, otherwise use calculated width
                adjusted_width = max(8, max_length + 2)
                overall_summary.column_dimensions[column_letter].width = adjusted_width
            
            # Auto-adjust column widths for all other sheets
            for sheet_name in ['Summary Data', 'Changes', 'Latest Data']:
                if sheet_name in book.sheetnames:
                    sheet = book[sheet_name]
                    for column in sheet.columns:
                        max_length = 0
                        # Get column letter from first non-merged cell
                        column_letter = None
                        for cell in column:
                            if hasattr(cell, 'column_letter'):
                                column_letter = cell.column_letter
                                break
                        
                        if not column_letter:
                            continue  # Skip if we can't get column letter
                        
                        for cell in column:
                            try:
                                # Skip merged cells
                                if hasattr(cell, 'value') and cell.value is not None:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = (max_length + 2)
                        sheet.column_dimensions[column_letter].width = adjusted_width
            
            # FINAL ATTEMPT: Set alignment as the very last operation
            overall_summary['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Final save with the Overall Summary sheet
            book.save(output_file)
            return True
            
        except PermissionError:
            if attempt < max_retries - 1:
                print(f"File {output_file} is in use. Waiting before retry...")
                time.sleep(2)
            else:
                print(f"Could not save to {output_file} - file is in use.")
                return False
    return False

def generate_standalone_report(input_file, output_file, config):
    """Generate a standalone report from a single input file"""
    try:
        # Read the input file
        current_df = pd.read_excel(input_file, **config['EXCEL_SETTINGS'])
        # Convert all columns to string
        for col in current_df.columns:
            current_df[col] = current_df[col].astype(str)
        
        # Get counts for summary
        counts = get_counts(current_df, config)
        
        # Create summary DataFrame with single row
        summary_data = [{
            'Date': datetime.now().strftime('%d-%b-%Y'),
            'Time': datetime.now().strftime('%H:%M')
        }]
        for key in sorted(counts.keys()):
            summary_data[0][key] = counts.get(key, 0)
        
        summary_df = pd.DataFrame(summary_data)
        
        # Create empty changes DataFrame since this is a standalone report
        changes_df = pd.DataFrame(columns=list(current_df.columns) + ['Change Type', 'Change Details'])
        
        # Save to Excel
        if save_excel_with_retry(summary_df, changes_df, current_df, output_file, config):
            print(f"\nStandalone report generated in {output_file}")
            return True
        else:
            print("\nPlease close any open Excel files and try again.")
            
    except Exception as e:
        print(f"Error generating standalone report: {str(e)}")
        return False

def slugify(text):
    return re.sub(r'[^A-Za-z0-9]+', '_', text).strip('_')

def detect_new_revision_types(sheet, new_revisions, revision_type):
    """Detect if there are new revision types that don't exist in the current progression report"""
    if not sheet or not new_revisions:
        return []
    
    # Get existing revision labels from column A
    existing_revisions = set()
    section_start_row = None
    section_end_row = None
    
    # Find the section for this revision type
    section_title = f'{revision_type} Revision Progression'
    section_found = False
    
    for row in range(1, sheet.max_row + 1):
        cell_value = sheet[f'A{row}'].value
        if cell_value == section_title:
            section_found = True
            section_start_row = row + 2  # Skip header and column header rows
            continue
        
        # If we've found the section, collect revision labels until we hit the next section or total
        if section_found and cell_value:
            if 'Total' in str(cell_value) or 'Progression' in str(cell_value):
                section_end_row = row - 1
                break
            # Extract revision name from the label (remove any prefix)
            revision_name = str(cell_value).replace('Rev_', '').replace('Status_', '')
            if revision_name and not revision_name.startswith('P Revisions') and not revision_name.startswith('C Revisions'):
                existing_revisions.add(revision_name)
    
    # Compare with new revisions
    new_revision_names = {rev.replace('Rev_', '') for rev in new_revisions}
    missing_revisions = new_revision_names - existing_revisions
    
    return sorted(list(missing_revisions))

def fill_empty_cells_with_zeros_in_file(progression_report_path):
    """Open the progression report, fill empty cells in tables with zeros, and save."""
    from openpyxl import load_workbook
    print(f"Filling empty cells in {progression_report_path}...")
    wb = load_workbook(progression_report_path)
    if 'Progression Report' not in wb.sheetnames:
        print("No 'Progression Report' sheet found.")
        return
    sheet = wb['Progression Report']
    
    # Find all date headers and their column numbers
    date_columns = set()
    for row_num in range(1, sheet.max_row + 1):
        for col_num in range(2, sheet.max_column + 1):
            col_letter = chr(ord('A') + col_num - 1)
            date_cell = sheet[f'{col_letter}{row_num}']
            if date_cell.value and '-' in str(date_cell.value) and len(str(date_cell.value)) > 5:
                date_columns.add(col_num)
    
    print(f"  Found date columns: {sorted(date_columns)}")
    
    # Find all table sections
    date_headers = []
    for row_num in range(1, sheet.max_row + 1):
        for col_num in range(2, sheet.max_column + 1):
            col_letter = chr(ord('A') + col_num - 1)
            date_cell = sheet[f'{col_letter}{row_num}']
            if date_cell.value and '-' in str(date_cell.value) and len(str(date_cell.value)) > 5:
                # This looks like a date header, find the next total row
                for next_row in range(row_num + 1, sheet.max_row + 1):
                    next_cell = sheet[f'A{next_row}']
                    if next_cell.value and 'Total' in str(next_cell.value):
                        date_headers.append((row_num, next_row))
                        break
                break
    
    # For each table section, fill empty cells with zeros only in date columns
    total_cells_filled = 0
    for start_row, end_row in date_headers:
        for row in range(start_row + 1, end_row):  # Skip the date header row
            for col_num in date_columns:  # Only process columns with date headers
                col_letter = chr(ord('A') + col_num - 1)
                cell = sheet[f'{col_letter}{row}']
                if cell.value is None or cell.value == '':
                    sheet[f'{col_letter}{row}'] = 0
                    # Apply consistent formatting with other data cells
                    sheet[f'{col_letter}{row}'].font = Font(name='Calibri', size=11)
                    sheet[f'{col_letter}{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    total_cells_filled += 1
    
    wb.save(progression_report_path)
    print(f"Filled {total_cells_filled} empty cells in {progression_report_path}.")

def generate_progression_report(summary_df, output_file, config, latest_data_df=None):
    """Generate a report showing the progression of revisions and statuses over time"""
    try:
        print(f"Debug: Starting progression report generation for {output_file}")
        print(f"Debug: summary_df has {len(summary_df)} rows")
        
        # Create a new workbook or load existing
        if os.path.exists(output_file):
            wb = load_workbook(output_file)
            print(f"Debug: Loaded existing progression report")
        else:
            wb = Workbook()
            # Remove the default 'Sheet' worksheet
            wb.remove(wb['Sheet'])
            print(f"Debug: Created new progression report")
        
        # Get or create the Progression Report sheet
        if 'Progression Report' in wb.sheetnames:
            sheet = wb['Progression Report']
            # Get the last used column
            last_col = 1  # Start with column A
            for col in sheet.columns:
                if any(cell.value for cell in col):
                    last_col = max(last_col, col[0].column)
            next_col = last_col + 1
            print(f"Debug: Existing sheet found, next column will be {next_col}")
        else:
            sheet = wb.create_sheet('Progression Report')
            next_col = 2  # Start with column B (A is for labels)
            print(f"Debug: Created new sheet, starting with column {next_col}")
        
        # Set up the sheet if it's new
        if next_col == 2:
            sheet.page_setup.fitToWidth = 1
            sheet.page_setup.fitToHeight = 0
            sheet.page_setup.horizontalCentered = True
            sheet.page_setup.verticalCentered = True
            sheet.page_margins = PageMargins(left=0.25, right=0.25, top=0.75, bottom=0.75, header=0.3, footer=0.3)
            
            # Add title
            sheet.merge_cells('A1:Z1')
            sheet['A1'] = f"{config.get('PROJECT_TITLE', '')} Document Register Progression Report"
            sheet['A1'].font = Font(name='Calibri', size=14, bold=True)
            sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Get all revision and status columns
        all_columns = summary_df.columns.tolist()
        rev_columns = [col for col in all_columns if col.startswith('Rev_')]
        status_columns = [col for col in all_columns if col.startswith('Status_')]
        
        # Sort revisions by type (P, C, other)
        p_revs = sorted([col for col in rev_columns if col.startswith('Rev_P')], 
                       key=lambda x: int(x.split('_')[1][1:]) if x.split('_')[1][1:].isdigit() else float('inf'))
        c_revs = sorted([col for col in rev_columns if col.startswith('Rev_C')],
                       key=lambda x: int(x.split('_')[1][1:]) if x.split('_')[1][1:].isdigit() else float('inf'))
        other_revs = sorted([col for col in rev_columns if not (col.startswith('Rev_P') or col.startswith('Rev_C'))])
        
        # Detect new revision types if this is an existing report
        if next_col > 2:  # Existing report
            new_p_revs = detect_new_revision_types(sheet, p_revs, 'P')
            new_c_revs = detect_new_revision_types(sheet, c_revs, 'C')
            
            if new_p_revs:
                print(f"WARNING: New P revision types detected: {', '.join(new_p_revs)}")
                print("These revisions will not be properly aligned in the progression report.")
                print("Consider regenerating the progression report to include proper row structure.")
            
            if new_c_revs:
                print(f"WARNING: New C revision types detected: {', '.join(new_c_revs)}")
                print("These revisions will not be properly aligned in the progression report.")
                print("Consider regenerating the progression report to include proper row structure.")
        
        # Function to get status count for a specific status group (all documents)
        def get_status_count(status_group):
            total = 0
            status_terms = PROGRESSION_STATUS_ORDER[status_group]['status_terms']
            
            for col in status_columns:
                # Extract the actual status name from the column name
                status_name = col.replace('Status_', '')
                if status_name in status_terms:
                    total += summary_df.iloc[-1].get(col, 0)
            return total
        
        # Function to get status count for a specific status group filtered by revision type
        def get_filtered_status_count(status_group, revision_type):
            if latest_data_df is None:
                # Fallback to unfiltered count if no latest data available
                return get_status_count(status_group)
            
            total = 0
            status_terms = PROGRESSION_STATUS_ORDER[status_group]['status_terms']
            
            # Filter data by revision type
            if revision_type == 'P':
                filtered_data = latest_data_df[latest_data_df['Rev'].str.startswith('P', na=False)]
            elif revision_type == 'C':
                filtered_data = latest_data_df[latest_data_df['Rev'].str.startswith('C', na=False)]
            else:
                # For other revision types, return 0 or handle as needed
                return 0
            
            # Check if this is Holloway Park project (custom status mapping)
            is_holloway_park = config.get('PROJECT_TITLE') == 'Holloway Park'
            
            if is_holloway_park:
                # For Holloway Park, map status group names to the custom mapped statuses
                status_mapping = {
                    'Status A': 'Status A',
                    'Status B': 'Status B', 
                    'Status C': 'Status C',
                    'Preliminary': 'Preliminary',
                    'Other': 'Other'
                }
                
                # Get the mapped status name for this status group
                mapped_status = status_mapping.get(status_group)
                if mapped_status:
                    count = len(filtered_data[filtered_data['Status'] == mapped_status])
                    total += count
            else:
                # For other projects, use the original logic
                for status_term in status_terms:
                    count = len(filtered_data[filtered_data['Status'] == status_term])
                    total += count
            
            return total
        
        # Function to get count of uncategorized statuses for a revision type
        def get_other_status_count(revision_type):
            if latest_data_df is None:
                return 0
            
            # Filter data by revision type
            if revision_type == 'P':
                filtered_data = latest_data_df[latest_data_df['Rev'].str.startswith('P', na=False)]
            elif revision_type == 'C':
                filtered_data = latest_data_df[latest_data_df['Rev'].str.startswith('C', na=False)]
            else:
                return 0
            
            # Check if this is Holloway Park project (custom status mapping)
            is_holloway_park = config.get('PROJECT_TITLE') == 'Holloway Park'
            
            if is_holloway_park:
                # For Holloway Park, 'Other' status is already counted in the status groups
                # So we return 0 to avoid double counting
                return 0
            else:
                # For other projects, use the original logic
                # Get all defined status terms
                all_defined_statuses = set()
                for status_group in PROGRESSION_STATUS_ORDER.values():
                    all_defined_statuses.update(status_group['status_terms'])
                
                # Count documents with statuses not in the defined list
                other_count = 0
                for status in filtered_data['Status'].unique():
                    if status not in all_defined_statuses:
                        count = len(filtered_data[filtered_data['Status'] == status])
                        other_count += count
                
                return other_count
        
        # Start row for data
        current_row = 3
        
        # Track total row positions for each section
        total_row_positions = {
            'P Revisions Total': None,
            'P Status Total': None,
            'C Revisions Total': None,
            'C Status Total': None
        }
        
        # Function to fill empty cells with zeros in table sections
        def fill_empty_cells_with_zeros():
            """Fill any empty cells in table sections with zeros"""
            print("  Debug: Starting fill_empty_cells_with_zeros function")
            
            # Find all date headers (column headers) to identify table sections
            date_headers = []
            for row_num in range(1, sheet.max_row + 1):
                cell = sheet[f'A{row_num}']
                # Look for date headers - they should be in columns B onwards and contain date-like patterns
                for col_num in range(2, sheet.max_column + 1):
                    col_letter = chr(ord('A') + col_num - 1)
                    date_cell = sheet[f'{col_letter}{row_num}']
                    if date_cell.value and '-' in str(date_cell.value) and len(str(date_cell.value)) > 5:
                        # This looks like a date header, find the next total row
                        for next_row in range(row_num + 1, sheet.max_row + 1):
                            next_cell = sheet[f'A{next_row}']
                            if next_cell.value and 'Total' in str(next_cell.value):
                                date_headers.append((row_num, next_row))
                                print(f"  Debug: Found table section from row {row_num} to {next_row}")
                                break
                        break
            
            print(f"  Debug: Found {len(date_headers)} table sections")
            
            # For each table section, fill empty cells with zeros
            for start_row, end_row in date_headers:
                # Get the column range (from B to the last column with data)
                # Always include the current column being added (next_col)
                max_col = next_col - 1  # The current column being added
                # Also check existing columns for any that might have data
                for col in sheet.columns:
                    if any(cell.value for cell in col):
                        max_col = max(max_col, col[0].column)
                
                print(f"  Debug: Processing section rows {start_row + 1} to {end_row}, columns B to {chr(ord('A') + max_col - 1)}")
                
                # Fill empty cells in this section
                cells_filled = 0
                for row in range(start_row + 1, end_row):  # Skip the date header row
                    for col_num in range(2, max_col + 1):  # Start from column B
                        col_letter = chr(ord('A') + col_num - 1)
                        cell = sheet[f'{col_letter}{row}']
                        if cell.value is None or cell.value == '':
                            sheet[f'{col_letter}{row}'] = 0
                            sheet[f'{col_letter}{row}'].font = Font(name='Calibri', size=11)
                            sheet[f'{col_letter}{row}'].alignment = Alignment(horizontal='center', vertical='center')
                            cells_filled += 1
                
                print(f"  Debug: Filled {cells_filled} empty cells in this section")
            
            print("  Debug: Completed fill_empty_cells_with_zeros function")
        
        # Function to add a section header
        def add_section_header(title, row):
            if next_col == 2:  # Only add headers for new sheets
                sheet.merge_cells(f'A{row}:Z{row}')
                sheet[f'A{row}'] = title
                sheet[f'A{row}'].font = Font(name='Calibri', size=12, bold=True)
                sheet[f'A{row}'].fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
            else:
                # For existing sheets, find the section header by name
                section_row_found = None
                for row_num in range(1, sheet.max_row + 1):
                    # Check if this cell is part of a merged range
                    cell = sheet[f'A{row_num}']
                    if hasattr(cell, 'coordinate') and cell.coordinate in sheet.merged_cells:
                        # For merged cells, get the value from the top-left cell of the merged range
                        for merged_range in sheet.merged_cells.ranges:
                            if cell.coordinate in merged_range:
                                # Get the top-left cell of the merged range
                                top_left_cell = sheet[f'A{merged_range.min_row}']
                                cell_value = top_left_cell.value
                                break
                    else:
                        cell_value = cell.value
                    
                    if cell_value and title in str(cell_value):
                        section_row_found = row_num
                        break
                
                if section_row_found:
                    # Use the found row position
                    row = section_row_found
            
            return row + 1
        
        # Function to add column headers
        def add_column_headers(row):
            # Add date header for the new column
            date_col = chr(ord('A') + next_col - 1)
            latest_data = summary_df.iloc[-1]  # Get the latest data
            sheet[f'{date_col}{row}'] = latest_data['Date']  # Only use the date
            sheet[f'{date_col}{row}'].font = Font(name='Calibri', size=11, bold=True)
            sheet[f'{date_col}{row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            return row + 1
        
        # Function to add data rows with proper row matching
        def add_data_rows(columns, start_row, title_prefix=''):
            row = start_row
            latest_data = summary_df.iloc[-1]  # Get the latest data
            rows_inserted = 0  # Track how many rows we've inserted
            
            if next_col == 2:
                # New sheet - create row headers and add data
                for col in columns:
                    # Add row header
                    sheet[f'A{row}'] = f"{title_prefix}{col.replace('Rev_', '').replace('Status_', '')}"
                    sheet[f'A{row}'].font = Font(name='Calibri', size=11)
                    sheet[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
                    
                    # Add data for the new column
                    date_col = chr(ord('A') + next_col - 1)
                    value = latest_data.get(col, 0)
                    sheet[f'{date_col}{row}'] = value
                    sheet[f'{date_col}{row}'].font = Font(name='Calibri', size=11)
                    sheet[f'{date_col}{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    row += 1
            else:
                # Existing sheet - match data to existing row headers
                # First, read existing row headers from column A
                existing_headers = []
                current_row_num = start_row
                
                # Read headers until we hit a total row or next section
                while current_row_num <= sheet.max_row:
                    cell_value = sheet[f'A{current_row_num}'].value
                    if not cell_value or 'Total' in str(cell_value) or 'Progression' in str(cell_value):
                        break
                    existing_headers.append((current_row_num, str(cell_value)))
                    current_row_num += 1
                
                # Create a mapping of revision names to their data values
                revision_data = {}
                for col in columns:
                    revision_name = col.replace('Rev_', '').replace('Status_', '')
                    revision_data[revision_name] = latest_data.get(col, 0)
                
                # Add data to existing rows and track which revisions we've handled
                handled_revisions = set()
                for header_row, header_text in existing_headers:
                    # Extract revision name from header (remove any prefix)
                    header_revision = header_text.replace('Rev_', '').replace('Status_', '')
                    
                    # Find matching revision data
                    if header_revision in revision_data:
                        # Add data to this row
                        date_col = chr(ord('A') + next_col - 1)
                        value = revision_data[header_revision]
                        
                        # Check if the cell is a merged cell before trying to edit it
                        cell = sheet[f'{date_col}{header_row}']
                        if not hasattr(cell, 'coordinate') or not cell.coordinate in sheet.merged_cells:
                            sheet[f'{date_col}{header_row}'] = value
                            sheet[f'{date_col}{header_row}'].font = Font(name='Calibri', size=11)
                            sheet[f'{date_col}{header_row}'].alignment = Alignment(horizontal='center', vertical='center')
                        else:
                            print(f"  Warning: Skipping merged cell at {date_col}{header_row}")
                        
                        handled_revisions.add(header_revision)
                    else:
                        # This row doesn't have data for this revision - add 0
                        date_col = chr(ord('A') + next_col - 1)
                        
                        # Check if the cell is a merged cell before trying to edit it
                        cell = sheet[f'{date_col}{header_row}']
                        if not hasattr(cell, 'coordinate') or not cell.coordinate in sheet.merged_cells:
                            sheet[f'{date_col}{header_row}'] = 0
                            sheet[f'{date_col}{header_row}'].font = Font(name='Calibri', size=11)
                            sheet[f'{date_col}{header_row}'].alignment = Alignment(horizontal='center', vertical='center')
                        else:
                            print(f"  Warning: Skipping merged cell at {date_col}{header_row}")
                
                # Find revisions that weren't handled (new revisions)
                new_revisions = set(revision_data.keys()) - handled_revisions
                
                if new_revisions:
                    print(f"INFO: Adding {len(new_revisions)} new revision(s) to existing progression report")
                    
                    # Sort new revisions properly (P01, P02, P10, P11, etc.)
                    def sort_revision_key(rev):
                        if rev.startswith('P') or rev.startswith('C'):
                            # Extract the number part
                            prefix = rev[0]  # P or C
                            number_part = rev[1:]
                            
                            # Handle special cases like P_Certificates
                            if '_' in number_part:
                                # For special cases like P_Certificates, put them at the end
                                return (prefix, float('inf'))
                            
                            try:
                                # Convert to int for proper numeric sorting
                                return (prefix, int(number_part))
                            except ValueError:
                                # If it's not a number, sort as string but after numeric ones
                                return (prefix, float('inf'))
                        else:
                            # For other revisions, sort as string
                            return ('Z', rev)  # Put at end
                    
                    sorted_new_revisions = sorted(new_revisions, key=sort_revision_key)
                    
                    # For each new revision, insert it at the appropriate position
                    for new_rev in sorted_new_revisions:
                        # Find the correct insertion position based on sorting
                        insert_position = None
                        
                        # Get all existing revisions in this section for comparison
                        section_revisions = []
                        for header_row, header_text in existing_headers:
                            if 'Total' not in header_text:
                                rev_name = header_text.replace('Rev_', '').replace('Status_', '')
                                section_revisions.append((header_row, rev_name))
                        
                        # Sort existing revisions
                        section_revisions.sort(key=lambda x: sort_revision_key(x[1]))
                        
                        # Find where to insert the new revision
                        for i, (header_row, rev_name) in enumerate(section_revisions):
                            if sort_revision_key(new_rev) < sort_revision_key(rev_name):
                                insert_position = header_row
                                break
                        
                        # If no position found, insert at the end (before total)
                        if insert_position is None:
                            if section_revisions:
                                insert_position = max(row for row, _ in section_revisions) + 1
                            else:
                                insert_position = start_row
                        
                        # Insert new row at the correct position
                        sheet.insert_rows(insert_position)
                        rows_inserted += 1  # Track the insertion
                        
                        # Update merged cells that are affected by the row insertion
                        # Get all merged ranges and update those that start at or after the insertion point
                        merged_ranges_to_update = []
                        for merged_range in list(sheet.merged_cells.ranges):
                            start_row = merged_range.min_row
                            end_row = merged_range.max_row
                            
                            # If the merged range starts at or after the insertion point, it needs to be updated
                            if start_row >= insert_position:
                                # Unmerge the current range
                                sheet.unmerge_cells(str(merged_range))
                                # Store the new range coordinates
                                new_start_row = start_row + 1
                                new_end_row = end_row + 1
                                merged_ranges_to_update.append((merged_range, new_start_row, new_end_row))
                        
                        # Re-merge the updated ranges
                        for old_range, new_start_row, new_end_row in merged_ranges_to_update:
                            # Create new range string
                            start_col = old_range.min_col
                            end_col = old_range.max_col
                            new_range = f"{chr(ord('A') + start_col - 1)}{new_start_row}:{chr(ord('A') + end_col - 1)}{new_end_row}"
                            sheet.merge_cells(new_range)
                        
                        # Update all row numbers in existing_headers that are >= insert_position
                        updated_headers = []
                        for header_row, header_text in existing_headers:
                            if header_row >= insert_position:
                                updated_headers.append((header_row + 1, header_text))
                            else:
                                updated_headers.append((header_row, header_text))
                        existing_headers = updated_headers
                        
                        # Update total row positions that are >= insert_position
                        for total_label, total_row in total_row_positions.items():
                            if total_row is not None and total_row >= insert_position:
                                total_row_positions[total_label] = total_row + 1
                        
                        # Add the new revision header
                        sheet[f'A{insert_position}'] = f"{title_prefix}{new_rev}"
                        sheet[f'A{insert_position}'].font = Font(name='Calibri', size=11)
                        sheet[f'A{insert_position}'].alignment = Alignment(horizontal='left', vertical='center')
                        
                        # Add the data
                        date_col = chr(ord('A') + next_col - 1)
                        value = revision_data[new_rev]
                        sheet[f'{date_col}{insert_position}'] = value
                        sheet[f'{date_col}{insert_position}'].font = Font(name='Calibri', size=11)
                        sheet[f'{date_col}{insert_position}'].alignment = Alignment(horizontal='center', vertical='center')
                        
                        # Update existing_headers to include the new row
                        existing_headers.append((insert_position, f"{title_prefix}{new_rev}"))
                        
                        print(f"  Added new revision: {new_rev} at row {insert_position}")
                
                # Return the row after the last handled row, adjusted for inserted rows
                if existing_headers:
                    row = max(header_row for header_row, _ in existing_headers) + 1
                else:
                    row = start_row
            
            return row
        
        # Function to add status data rows with revision filtering
        def add_status_data_rows(start_row, revision_type):
            row = start_row
            latest_data = summary_df.iloc[-1]  # Get the latest data
            
            # Add each status in the defined order
            for status_group in PROGRESSION_STATUS_ORDER:
                # Add row header if it's a new sheet
                if next_col == 2:
                    sheet[f'A{row}'] = PROGRESSION_STATUS_ORDER[status_group]['display_name']
                    sheet[f'A{row}'].font = Font(name='Calibri', size=11)
                    sheet[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
                
                # Add filtered data for the new column
                date_col = chr(ord('A') + next_col - 1)
                value = get_filtered_status_count(status_group, revision_type)
                sheet[f'{date_col}{row}'] = value
                sheet[f'{date_col}{row}'].font = Font(name='Calibri', size=11)
                sheet[f'{date_col}{row}'].alignment = Alignment(horizontal='center', vertical='center')
                
                row += 1
            
            # Check if this is Holloway Park project
            is_holloway_park = config.get('PROJECT_TITLE') == 'Holloway Park'
            
            # For Holloway Park, skip the "Other Status" row since 'Other' is already included in status groups
            # For other projects, add "Other Status" row to maintain consistent structure
            if not is_holloway_park:
                # Add row header if it's a new sheet
                if next_col == 2:
                    sheet[f'A{row}'] = 'Other Status'
                    sheet[f'A{row}'].font = Font(name='Calibri', size=11)
                    sheet[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
                
                # Add filtered data for the new column (always add, even if 0)
                date_col = chr(ord('A') + next_col - 1)
                other_count = get_other_status_count(revision_type)
                sheet[f'{date_col}{row}'] = other_count
                sheet[f'{date_col}{row}'].font = Font(name='Calibri', size=11)
                sheet[f'{date_col}{row}'].alignment = Alignment(horizontal='center', vertical='center')
                
                row += 1
            
            return row
        
        # Function to calculate total for a set of columns
        def calculate_total(columns):
            total = 0
            latest_data = summary_df.iloc[-1]
            for col in columns:
                value = latest_data.get(col, 0)
                # Handle nan values - treat them as 0
                if pd.isna(value):
                    value = 0
                total += value
            return total
        
        # Function to add a total row
        def add_total_row(start_row, total_value, label='Total'):
            # Use tracked position if available, otherwise find by name
            if next_col > 2:
                # Check if we have a tracked position for this total row
                if label in total_row_positions and total_row_positions[label] is not None:
                    actual_row = total_row_positions[label]
                else:
                    # Fallback: Search for the total row by name in column A
                    total_row_found = None
                    for row_num in range(1, sheet.max_row + 1):
                        # Check if this cell is part of a merged range
                        cell = sheet[f'A{row_num}']
                        if hasattr(cell, 'coordinate') and cell.coordinate in sheet.merged_cells:
                            # For merged cells, get the value from the top-left cell of the merged range
                            for merged_range in sheet.merged_cells.ranges:
                                if cell.coordinate in merged_range:
                                    # Get the top-left cell of the merged range
                                    top_left_cell = sheet[f'A{merged_range.min_row}']
                                    cell_value = top_left_cell.value
                                    break
                        else:
                            cell_value = cell.value
                        
                        if cell_value and label in str(cell_value):
                            total_row_found = row_num
                            break
                    
                    if total_row_found:
                        # Use the found row position and store it for future use
                        actual_row = total_row_found
                        total_row_positions[label] = actual_row
                    else:
                        # If not found, use the provided start_row
                        actual_row = start_row
                        total_row_positions[label] = actual_row
            else:
                # New sheet - add row header
                actual_row = start_row
                sheet[f'A{actual_row}'] = label
                sheet[f'A{actual_row}'].font = Font(name='Calibri', size=11, bold=True)
                sheet[f'A{actual_row}'].alignment = Alignment(horizontal='left', vertical='center')
                sheet[f'A{actual_row}'].fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
                # Store the position for future use
                total_row_positions[label] = actual_row
            
            # Add total data for the new column
            date_col = chr(ord('A') + next_col - 1)
            
            # Check if the cell is a merged cell before trying to edit it
            cell = sheet[f'{date_col}{actual_row}']
            if not hasattr(cell, 'coordinate') or not cell.coordinate in sheet.merged_cells:
                sheet[f'{date_col}{actual_row}'] = total_value
                sheet[f'{date_col}{actual_row}'].font = Font(name='Calibri', size=11, bold=True)
                sheet[f'{date_col}{actual_row}'].alignment = Alignment(horizontal='center', vertical='center')
                sheet[f'{date_col}{actual_row}'].fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
                print(f"  Added total {total_value} to {date_col}{actual_row} for {label}")
            else:
                print(f"  Warning: Skipping merged cell at {date_col}{actual_row} for total row {label}")
                # Try to find the correct cell in the merged range
                for merged_range in sheet.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        # Find the correct column within the merged range
                        if merged_range.min_col <= ord(date_col) - ord('A') + 1 <= merged_range.max_col:
                            # This column is within the merged range, try to write to the top-left cell
                            top_left_cell = sheet[f'{date_col}{merged_range.min_row}']
                            if not hasattr(top_left_cell, 'coordinate') or not top_left_cell.coordinate in sheet.merged_cells:
                                sheet[f'{date_col}{merged_range.min_row}'] = total_value
                                sheet[f'{date_col}{merged_range.min_row}'].font = Font(name='Calibri', size=11, bold=True)
                                sheet[f'{date_col}{merged_range.min_row}'].alignment = Alignment(horizontal='center', vertical='center')
                                sheet[f'{date_col}{merged_range.min_row}'].fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
                                print(f"  Added total {total_value} to {date_col}{merged_range.min_row} (merged range) for {label}")
                                break
                        else:
                            print(f"  Column {date_col} is not within merged range {merged_range}")
                        break
            
            return actual_row + 1
        
        # Add P Revision section
        current_row = add_section_header('P Revision Progression', current_row)
        current_row = add_column_headers(current_row)
        current_row = add_data_rows(p_revs, current_row)
        # Add total row for P revisions
        p_total = calculate_total(p_revs)
        current_row = add_total_row(current_row, p_total, 'P Revisions Total')
        current_row += 1  # Add spacing
        
        # Add P Revision Status section
        current_row = add_section_header('P Revision Status Progression', current_row)
        current_row = add_column_headers(current_row)
        current_row = add_status_data_rows(current_row, 'P')
        # Add total row for P revision statuses
        p_status_total = sum(get_filtered_status_count(status_group, 'P') for status_group in PROGRESSION_STATUS_ORDER) + get_other_status_count('P')
        current_row = add_total_row(current_row, p_status_total, 'P Status Total')
        current_row += 1  # Add spacing
        
        # Add C Revision section
        current_row = add_section_header('C Revision Progression', current_row)
        current_row = add_column_headers(current_row)
        current_row = add_data_rows(c_revs, current_row)
        # Add total row for C revisions
        c_total = calculate_total(c_revs)
        current_row = add_total_row(current_row, c_total, 'C Revisions Total')
        current_row += 1  # Add spacing
        
        # Add C Revision Status section
        current_row = add_section_header('C Revision Status Progression', current_row)
        current_row = add_column_headers(current_row)
        current_row = add_status_data_rows(current_row, 'C')
        # Add total row for C revision statuses
        c_status_total = sum(get_filtered_status_count(status_group, 'C') for status_group in PROGRESSION_STATUS_ORDER) + get_other_status_count('C')
        current_row = add_total_row(current_row, c_status_total, 'C Status Total')
        
        # Note: fill_empty_cells_with_zeros is now called once at the end of all processing
        # to avoid running it multiple times per file

        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                if hasattr(cell, 'column_letter'):
                    column_letter = cell.column_letter
                    break
            
            if not column_letter:
                continue
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = max(8, max_length + 2)
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook
        wb.save(output_file)
        return True
        
    except Exception as e:
        print(f"Error generating progression report: {str(e)}")
        return False

def load_processed_files_per_project():
    """Load the record of processed files per project"""
    try:
        with open('processed_files_per_project.json', 'r') as f:
            return json.load(f)
    except FileNotFoundError:
        return {}

def save_processed_files_per_project(processed_files):
    """Save the record of processed files per project"""
    with open('processed_files_per_project.json', 'w') as f:
        json.dump(processed_files, f, indent=2, default=str)

def get_project_files_with_timestamps(project_input_dir):
    """Get all Excel files in a project directory with their timestamps, sorted by date"""
    files_with_timestamps = []
    
    # Check if this is Holloway Park project (by folder name)
    is_holloway_park = 'HP' in str(project_input_dir) or 'holloway' in str(project_input_dir).lower()
    
    # Get file patterns to search for
    if is_holloway_park:
        file_patterns = ["*.xlsx", "*.csv"]
    else:
        file_patterns = ["*.xlsx"]
    
    for pattern in file_patterns:
        for file_path in project_input_dir.glob(pattern):
            if file_path.name.startswith('~$'):  # Skip temporary files
                continue
                
            # Get timestamp from B4 (Excel) or Report Created column (CSV)
            date_str, time_str = get_file_timestamp(file_path)
            if not date_str or not time_str:
                print(f"Skipping {file_path.name} - could not read timestamp")
                continue
                
            # Convert to datetime for comparison
            try:
                date = datetime.strptime(date_str, '%d-%b-%Y')
                time = datetime.strptime(time_str, '%H:%M').time()
                files_with_timestamps.append((file_path, date, time, date_str, time_str))
            except ValueError as e:
                print(f"Warning: Could not parse date/time from {file_path.name}: {str(e)}")
                continue
    
    # Sort by date and time (oldest first)
    files_with_timestamps.sort(key=lambda x: (x[1], x[2]))
    return files_with_timestamps

def detect_project_files():
    """Detect all files in project folders and update the JSON tracking file"""
    print("Detecting files in project folders...")
    
    # Setup directories
    input_dir = Path('input')
    
    # Define project folders with consistent naming
    project_folders = {
        'OVB': input_dir / 'OVB',
        'NM': input_dir / 'NM', 
        'GP': input_dir / 'GP',
        'HP': input_dir / 'HP'
    }
    
    # Load existing processed files record
    processed_files = load_processed_files_per_project()
    
    # Handle legacy "NewMalden" entry by merging with "NM"
    if "NewMalden" in processed_files and "NM" not in processed_files:
        processed_files["NM"] = processed_files.pop("NewMalden")
    elif "NewMalden" in processed_files and "NM" in processed_files:
        # Merge the two entries
        processed_files["NM"].update(processed_files.pop("NewMalden"))
    
    # Process each project folder
    for project_code, project_input_dir in project_folders.items():
        if not project_input_dir.exists():
            print(f"Project folder {project_input_dir} does not exist, skipping...")
            continue
        
        print(f"\nProcessing project: {project_code}")
        print(f"Folder: {project_input_dir}")
        
        # Get all files with timestamps for this project
        files_with_timestamps = get_project_files_with_timestamps(project_input_dir)
        
        if not files_with_timestamps:
            print(f"No valid files found in {project_input_dir}")
            continue
        
        print(f"Found {len(files_with_timestamps)} files:")
        
        # Initialize project in processed files if not exists
        if project_code not in processed_files:
            processed_files[project_code] = {}
        
        # Check each file and add to processed files (simulating processing)
        for file_path, date, time, date_str, time_str in files_with_timestamps:
            file_key = f"{date_str}_{time_str}"
            
            if file_path.name in processed_files[project_code]:
                if processed_files[project_code][file_path.name] == file_key:
                    print(f"  ✓ {file_path.name} ({date_str} {time_str}) - Already processed")
                else:
                    print(f"  ↻ {file_path.name} ({date_str} {time_str}) - Updated, will reprocess")
                    processed_files[project_code][file_path.name] = file_key
            else:
                print(f"  + {file_path.name} ({date_str} {time_str}) - New file, will process")
                processed_files[project_code][file_path.name] = file_key
    
    # Save updated processed files record
    save_processed_files_per_project(processed_files)
    print(f"\nUpdated processed_files_per_project.json with all detected files")
    
    return processed_files

def show_menu():
    """Display the main menu and get user choice"""
    print("\n" + "="*60)
    print("           DOCUMENT REGISTER REPORT GENERATOR")
    print("="*60)
    print("1. Process latest file (original behavior)")
    print("2. Process all projects (all unprocessed files)")
    print("3. Process single project (all unprocessed files)")
    print("4. Detect files only (scan and update tracking)")
    print("5. Generate standalone report")
    print("6. Exit")
    print("-"*60)
    
    while True:
        try:
            choice = input("Enter your choice (1-6): ").strip()
            if choice in ['1', '2', '3', '4', '5', '6']:
                return choice
            else:
                print("Invalid choice. Please enter a number between 1 and 6.")
        except KeyboardInterrupt:
            print("\n\nExiting...")
            return '6'
        except EOFError:
            print("\n\nExiting...")
            return '6'

def get_project_selection():
    """Get project selection from user"""
    print("\n" + "-"*40)
    print("PROJECT SELECTION")
    print("-"*40)
    print("Available projects:")
    print("1. GreenwichPeninsula (GP)")
    print("2. NewMalden (NM)")
    print("3. OvalBlockB (OVB)")
    print("4. HollowayPark (HP)")
    
    while True:
        project_choice = input("Select project (1-4): ").strip()
        if project_choice == '1':
            return 'GreenwichPeninsula', 'GP'
        elif project_choice == '2':
            return 'NewMalden', 'NM'
        elif project_choice == '3':
            return 'OvalBlockB', 'OVB'
        elif project_choice == '4':
            return 'HollowayPark', 'HP'
        else:
            print("Invalid choice. Please enter 1, 2, 3, or 4.")

def get_standalone_input():
    """Get input file and project for standalone report"""
    print("\n" + "-"*40)
    print("STANDALONE REPORT GENERATION")
    print("-"*40)
    
    # Get input file
    while True:
        input_file = input("Enter input file path: ").strip().strip('"')
        if not input_file:
            print("Input file path cannot be empty.")
            continue
        
        input_path = Path(input_file)
        if not input_path.exists():
            print(f"File '{input_file}' does not exist.")
            continue
        
        if not input_path.suffix.lower() in ['.xlsx', '.xls']:
            print("Please select an Excel file (.xlsx or .xls)")
            continue
        
        break
    
    # Get project name
    print("\nAvailable projects:")
    print("1. GreenwichPeninsula")
    print("2. NewMalden") 
    print("3. OvalBlockB")
    print("4. HollowayPark")
    print("5. Auto-detect (recommended)")
    
    while True:
        project_choice = input("Select project (1-5): ").strip()
        if project_choice == '1':
            project = 'GreenwichPeninsula'
            break
        elif project_choice == '2':
            project = 'NewMalden'
            break
        elif project_choice == '3':
            project = 'OvalBlockB'
            break
        elif project_choice == '4':
            project = 'HollowayPark'
            break
        elif project_choice == '5':
            project = None  # Auto-detect
            break
        else:
            print("Invalid choice. Please enter 1, 2, 3, 4, or 5.")
    
    return input_path, project

def main():
    """Main function with interactive menu"""
    # Setup directories
    input_dir = Path('input')
    output_dir = Path('output')
    output_dir.mkdir(exist_ok=True)
    
    while True:
        choice = show_menu()
        
        if choice == '1':
            # Process latest file (original behavior)
            print("\nProcessing latest file...")
            
            # Load processed files record
            processed_files = load_processed_files()
            
            # Dictionary to store all counts
            all_counts = {}
            all_changes = []
            
            # Get the previous latest data
            previous_latest_data = None
            output_file = output_dir / 'summary.xlsx'
            if output_file.exists():
                try:
                    previous_latest_data = pd.read_excel(output_file, sheet_name='Latest Data')
                    # Convert all columns to string
                    for col in previous_latest_data.columns:
                        previous_latest_data[col] = previous_latest_data[col].astype(str)
                except Exception as e:
                    print(f"Warning: Could not load previous latest data: {str(e)}")
            
            # Find the most recent file based on timestamps
            latest_file = None
            latest_timestamp = None
            
            for file_path in input_dir.glob("*.xlsx"):
                if file_path.name.startswith('~$'):  # Skip temporary files
                    continue
                    
                # Get timestamp from B4
                date_str, time_str = get_file_timestamp(file_path)
                if not date_str or not time_str:
                    print(f"Skipping {file_path.name} - could not read timestamp")
                    continue
                    
                # Convert to datetime for comparison
                try:
                    date = datetime.strptime(date_str, '%d-%b-%Y')
                    time = datetime.strptime(time_str, '%H:%M').time()
                    if latest_timestamp is None or (date, time) > latest_timestamp:
                        latest_timestamp = (date, time)
                        latest_file = file_path
                except ValueError as e:
                    print(f"Warning: Could not parse date/time from {file_path.name}: {str(e)}")
                    continue
            
            if latest_file is None:
                print("No valid files found to process")
                input("\nPress Enter to continue...")
                continue
            
            print(f"\nProcessing latest file: {latest_file.name}")
            
            # Load project configuration based on the latest file
            config = load_project_config(None, latest_file)
            
            # Read the latest file
            try:
                current_df = pd.read_excel(latest_file, **config['EXCEL_SETTINGS'])
                
                # Convert all columns to string
                for col in current_df.columns:
                    try:
                        current_df[col] = current_df[col].astype(str)
                    except Exception as e:
                        print(f"Warning: Error converting column '{col}' to string: {str(e)}")
                
                # Compare with previous latest data if it exists
                if previous_latest_data is not None:
                    # Create a dictionary of previous data using composite key (Doc Ref + Doc Path)
                    prev_data_dict = {}
                    for _, row in previous_latest_data.iterrows():
                        key = (row['Doc Ref'], row['Doc Path'])
                        prev_data_dict[key] = row
                    
                    # Note: We're only generating summary reports, so we don't need detailed change tracking
                    # The comparison logic has been removed to simplify the script
                
                # Get counts for summary
                try:
                    counts = get_counts(current_df, config)
                    all_counts[latest_timestamp] = counts
                except Exception as e:
                    print(f"Error getting counts: {str(e)}")
                    raise
                
                # Update processed files record
                date_str, time_str = get_file_timestamp(latest_file)
                file_key = f"{date_str}_{time_str}"  # Use the string versions
                processed_files[latest_file.name] = file_key
                
            except Exception as e:
                print(f"Error processing {latest_file.name}: {str(e)}")
                input("\nPress Enter to continue...")
                continue
            
            # Create summary DataFrame
            summary_data = []
            for (date, time) in sorted(all_counts.keys()):
                row = {
                    'Date': date.strftime('%d-%b-%Y'),  # Convert to string
                    'Time': time.strftime('%H:%M')      # Convert to string
                }
                counts = all_counts[(date, time)]
                for key in sorted(counts.keys()):
                    row[key] = counts.get(key, 0)
                summary_data.append(row)
            
            summary_df = pd.DataFrame(summary_data)
            
            # Create changes DataFrame
            changes_df = pd.DataFrame(all_changes) if all_changes else pd.DataFrame(columns=list(current_df.columns) + ['Change Type', 'Change Details'])
            
            # Save to Excel
            project_slug = slugify(config.get('PROJECT_TITLE', 'summary'))
            output_file = output_dir / f"{project_slug}_summary.xlsx"
            
            if save_excel_with_retry(summary_df, changes_df, current_df, output_file, config):
                # Save processed files record
                save_processed_files(processed_files)
                print(f"\nSummary updated in {output_file}")
                
                # Generate progression report
                progression_output = output_dir / f"{project_slug}_progression.xlsx"
                if generate_progression_report(summary_df, progression_output, config, current_df):
                    print(f"Progression report generated in {progression_output}")
                else:
                    print("Failed to generate progression report")
            else:
                print("\nPlease close any open Excel files and try again.")
            
            input("\nPress Enter to continue...")
            
        elif choice == '2':
            # Process all projects
            print("\nProcessing all projects...")
            process_all_projects()
            input("\nPress Enter to continue...")
            
        elif choice == '3':
            # Process single project
            project_name, project_code = get_project_selection()
            print(f"\nProcessing single project: {project_name}")
            process_single_project(project_name, project_code)
            input("\nPress Enter to continue...")
            
        elif choice == '4':
            # Detect files only
            print("\nDetecting files in project folders...")
            detect_project_files()
            input("\nPress Enter to continue...")
            
        elif choice == '5':
            # Generate standalone report
            try:
                input_file, project = get_standalone_input()
                
                # Load project configuration based on input file
                config = load_project_config(project, input_file)
                
                # Generate output filename
                project_slug = slugify(config.get('PROJECT_TITLE', 'standalone_report'))
                output_file = output_dir / f"{project_slug}_standalone_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                
                if generate_standalone_report(input_file, output_file, config):
                    print(f"\nStandalone report generated successfully!")
                else:
                    print("\nFailed to generate standalone report.")
                    
            except KeyboardInterrupt:
                print("\nCancelled.")
            except Exception as e:
                print(f"\nError: {str(e)}")
            
            input("\nPress Enter to continue...")
            
        elif choice == '6':
            # Exit
            print("\nGoodbye!")
            break

def process_all_projects():
    """Process all projects in their respective input folders"""
    # Setup directories
    input_dir = Path('input')
    output_dir = Path('output')
    output_dir.mkdir(exist_ok=True)
    
    # Define project folders
    project_folders = {
        'OVB': input_dir / 'OVB',
        'NM': input_dir / 'NM', 
        'GP': input_dir / 'GP',
        'HP': input_dir / 'HP'
    }
    
    # Load processed files record
    processed_files = load_processed_files_per_project()
    
    # Process each project
    for project_code, project_input_dir in project_folders.items():
        if not project_input_dir.exists():
            print(f"\nProject folder {project_input_dir} does not exist, skipping...")
            continue
        
        # Map project codes to full names
        project_names = {
            'OVB': 'OvalBlockB',
            'NM': 'NewMalden',
            'GP': 'GreenwichPeninsula',
            'HP': 'HollowayPark'
        }
        
        project_name = project_names.get(project_code, project_code)
        
        # Process this project
        success = process_project_files(project_name, project_input_dir, output_dir, processed_files)
        
        if success:
            print(f"Successfully processed project: {project_name}")
        else:
            print(f"Failed to process project: {project_name}")
    
    # Save processed files record
    save_processed_files_per_project(processed_files)
    print(f"\n{'='*60}")
    print("All projects processed!")
    print(f"{'='*60}")

def process_project_files(project_name, project_input_dir, output_dir, processed_files):
    """Process all files for a specific project in chronological order"""
    print(f"\n{'='*60}")
    print(f"Processing project: {project_name}")
    print(f"{'='*60}")
    
    # Get all files with timestamps for this project
    files_with_timestamps = get_project_files_with_timestamps(project_input_dir)
    
    if not files_with_timestamps:
        print(f"No valid files found in {project_input_dir}")
        return False
    
    print(f"Found {len(files_with_timestamps)} files to process:")
    for file_path, date, time, date_str, time_str in files_with_timestamps:
        print(f"  - {file_path.name} ({date_str} {time_str})")
    
    # Initialize project-specific data
    all_counts = {}
    latest_data_df = None
    
    # Get the previous latest data for this project
    project_slug = slugify(project_name)
    project_output_file = output_dir / f"{project_slug}_summary.xlsx"
    previous_latest_data = None
    
    if project_output_file.exists():
        try:
            previous_latest_data = pd.read_excel(project_output_file, sheet_name='Latest Data')
            for col in previous_latest_data.columns:
                previous_latest_data[col] = previous_latest_data[col].astype(str)
        except Exception as e:
            print(f"Warning: Could not load previous latest data for {project_name}: {str(e)}")
    
    if project_name not in processed_files:
        processed_files[project_name] = {}
    
    # Track if any files were processed
    files_processed = False
    
    for file_path, date, time, date_str, time_str in files_with_timestamps:
        file_key = f"{date_str}_{time_str}"
        if file_path.name in processed_files[project_name]:
            if processed_files[project_name][file_path.name] == file_key:
                print(f"Skipping {file_path.name} - already processed")
                continue
            else:
                print(f"File {file_path.name} has been updated, reprocessing...")
        print(f"\nProcessing: {file_path.name} ({date_str} {time_str})")
        config = load_project_config(project_name, file_path)
        try:
            # Determine file type and read accordingly
            file_path_str = str(file_path).lower()
            
            if file_path_str.endswith('.csv'):
                # Process CSV file
                current_df = process_csv_file(file_path, config)
            else:
                # Process Excel file
                current_df = pd.read_excel(file_path, **config['EXCEL_SETTINGS'])
            
            # Convert all columns to string
            for col in current_df.columns:
                try:
                    current_df[col] = current_df[col].astype(str)
                except Exception as e:
                    print(f"Warning: Error converting column '{col}' to string: {str(e)}")
            
            try:
                counts = get_counts(current_df, config)
                all_counts[(date, time)] = counts
            except Exception as e:
                print(f"Error getting counts: {str(e)}")
                raise
            processed_files[project_name][file_path.name] = file_key
            latest_data_df = current_df.copy()
            previous_latest_data = current_df.copy()
            files_processed = True
            
            # Generate progression report for this file (add new column)
            progression_output = output_dir / f"{project_slug}_progression.xlsx"
            
            # Create a single-row summary DataFrame for this file
            file_summary_data = [{
                'Date': date.strftime('%d-%b-%Y'),
                'Time': time.strftime('%H:%M')
            }]
            for key in sorted(counts.keys()):
                file_summary_data[0][key] = counts.get(key, 0)
            file_summary_df = pd.DataFrame(file_summary_data)
            
            print(f"Adding progression data for {date_str} {time_str}...")
            if generate_progression_report(file_summary_df, progression_output, config, current_df):
                print(f"Progression report updated with new column")
            else:
                print("Failed to update progression report")
                
        except Exception as e:
            print(f"Error processing {file_path.name}: {str(e)}")
            continue

    # Guard: If nothing was processed, don't continue
    if not files_processed or latest_data_df is None or 'config' not in locals():
        print("No new files processed for this project.")
        return False

    summary_data = []
    for (date, time) in sorted(all_counts.keys()):
        row = {
            'Date': date.strftime('%d-%b-%Y'),
            'Time': time.strftime('%H:%M')
        }
        counts = all_counts[(date, time)]
        for key in sorted(counts.keys()):
            row[key] = counts.get(key, 0)
        summary_data.append(row)
    summary_df = pd.DataFrame(summary_data)
    
    # Save to Excel (summary only)
    if save_excel_with_retry(summary_df, None, latest_data_df, project_output_file, config):
        print(f"\nSummary updated in {project_output_file}")
    else:
        print("\nPlease close any open Excel files and try again.")
        return False
        
    # After all files are processed, fill empty cells in the progression report
    progression_output = output_dir / f"{project_slug}_progression.xlsx"
    fill_empty_cells_with_zeros_in_file(str(progression_output))
    return True

def process_single_project(project_name, project_code):
    """Process all files for a specific project in chronological order"""
    print(f"\n{'='*60}")
    print(f"Processing project: {project_name}")
    print(f"{'='*60}")
    
    # Setup directories
    input_dir = Path('input')
    output_dir = Path('output')
    output_dir.mkdir(exist_ok=True)
    
    # Define project folders
    project_folders = {
        'OVB': input_dir / 'OVB',
        'NM': input_dir / 'NM',
        'GP': input_dir / 'GP',
        'HP': input_dir / 'HP'
    }
    
    # Load processed files record
    processed_files = load_processed_files_per_project()
    
    # Process this project
    success = process_project_files(project_name, project_folders[project_code], output_dir, processed_files)
    
    if success:
        print(f"Successfully processed project: {project_name}")
    else:
        print(f"Failed to process project: {project_name}")
    
    # Save processed files record
    save_processed_files_per_project(processed_files)
    print(f"\n{'='*60}")
    print(f"Project {project_name} processed!")
    print(f"{'='*60}")

if __name__ == "__main__":
    main() 