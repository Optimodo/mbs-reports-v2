"""Layout tracking report generation module.

This module generates reports for tracking apartment layout drawings,
showing coverage by apartment type with detailed progress tracking.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule
from datetime import datetime
from pathlib import Path

from analyzers.document_tracker import get_layout_tracking_summary


# Styling constants
HEADER_FONT = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

SECTION_FONT = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
SECTION_FILL = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')

SUBHEADER_FONT = Font(name='Calibri', size=10, bold=True)
SUBHEADER_FILL = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

BORDER_THIN = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def get_progress_color(percentage):
    """Get color code based on progress percentage."""
    if percentage >= 90:
        return '25E82C'  # Green
    elif percentage >= 75:
        return '92D050'  # Light green
    elif percentage >= 50:
        return 'FFC000'  # Orange/yellow
    elif percentage >= 25:
        return 'FF9900'  # Orange
    else:
        return 'ED1111'  # Red


def add_progress_bar(ws, cell_ref, percentage):
    """Add a data bar progress indicator to a cell."""
    cell = ws[cell_ref]
    cell.value = percentage / 100
    
    # Get color based on percentage
    progress_color_hex = get_progress_color(percentage)
    bar_color = Color(rgb=progress_color_hex)
    
    # Create data bar rule
    data_bar = DataBarRule(
        start_type='num', start_value=0,
        end_type='num', end_value=1,
        color=bar_color,
        showValue=True,
        minLength=None, maxLength=None
    )
    
    ws.conditional_formatting.add(cell_ref, data_bar)
    
    # Format the cell
    cell.number_format = '0%'
    cell.font = Font(name='Calibri', size=11, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )


def add_overview_section(ws, layout_summary, accommodation_data):
    """Add overview statistics section."""
    row = 5
    
    # Section header
    ws[f'A{row}'] = 'LAYOUT TRACKING OVERVIEW'
    ws[f'A{row}'].font = SECTION_FONT
    ws[f'A{row}'].fill = SECTION_FILL
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(f'A{row}:G{row}')
    row += 2
    
    # Statistics
    total_types = len(accommodation_data.get('apartment_types', {})) if accommodation_data else 0
    total_layouts = layout_summary.get('total_layouts', 0)
    withdrawn = layout_summary.get('withdrawn_count', 0)
    categorized = layout_summary.get('categorized', 0)
    uncategorized = layout_summary.get('uncategorized', 0)
    
    stats = [
        ('Total Apartment Types in Project:', total_types),
        ('', ''),
        ('Total Layout Drawings Found:', total_layouts),
        ('Withdrawn (Excluded):', withdrawn),
        ('Categorized Layouts:', categorized),
        ('Uncategorized Layouts:', uncategorized),
    ]
    
    for label, value in stats:
        if label:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(name='Calibri', size=10, bold=True)
            ws[f'B{row}'].font = Font(name='Calibri', size=10)
            ws[f'B{row}'].alignment = Alignment(horizontal='left')
        row += 1
    
    return row + 2


def add_apartment_layouts_section(ws, apartment_progress, start_row):
    """Add detailed apartment layout tracking table."""
    
    # Section header
    ws[f'A{start_row}'] = 'APARTMENT LAYOUT CATEGORIES'
    ws[f'A{start_row}'].font = SECTION_FONT
    ws[f'A{start_row}'].fill = SECTION_FILL
    ws[f'A{start_row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(f'A{start_row}:G{start_row}')
    start_row += 2
    
    # Column headers - simplified
    headers = ['Layout Category', 'Types Found', 'Documents', 'Progress Bar']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx)
        cell.value = header
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER_THIN
    
    # Merge the progress bar header from D to G
    ws.merge_cells(f'D{start_row}:G{start_row}')
    progress_bar_cell = ws.cell(row=start_row, column=4)
    progress_bar_cell.value = 'Progress Bar'
    progress_bar_cell.font = SUBHEADER_FONT
    progress_bar_cell.fill = SUBHEADER_FILL
    progress_bar_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add borders to progress bar header cells (D through G = columns 4-7)
    for col_num in range(4, 8):
        cell = ws.cell(row=start_row, column=col_num)
        cell.border = BORDER_THIN
    
    start_row += 1
    
    # Add data for each layout category
    for layout_key, progress in apartment_progress.items():
        display_name = progress['display_name']
        # Use required types (excluding greylisted) for denominator
        types_found = f"{progress['types_with_layout']} / {progress.get('total_required_types', progress['total_expected_types'])}"
        coverage_pct = progress['coverage_percentage']
        unique_count = progress['unique_document_count']
        duplicate_count = progress['duplicate_count']
        total_docs = progress['document_count']
        
        # Layout category name
        ws.cell(row=start_row, column=1, value=display_name).font = Font(name='Calibri', size=10, bold=True)
        
        # Types found (fraction) - shows required types only
        ws.cell(row=start_row, column=2, value=types_found)
        ws.cell(row=start_row, column=2).alignment = Alignment(horizontal='center')
        
        # Total documents
        coverage_pct = progress['coverage_percentage']
        ws.cell(row=start_row, column=3, value=total_docs)
        ws.cell(row=start_row, column=3).alignment = Alignment(horizontal='center')
        
        # Merge cells for progress bar (D through G)
        ws.merge_cells(f'D{start_row}:G{start_row}')
        
        # Create visual progress bar using data bar
        add_progress_bar(ws, f'D{start_row}', coverage_pct)
        
        # Add border to entire merged progress bar area (D through G = columns 4-7)
        for col_num in range(4, 8):
            cell = ws.cell(row=start_row, column=col_num)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        start_row += 1
    
    start_row += 2
    
    return start_row + 2


def add_missing_types_summary(ws, apartment_progress, start_row):
    """Add condensed summary of missing apartment types (for main report page)."""
    
    # Section header
    ws[f'A{start_row}'] = 'MISSING APARTMENT TYPES'
    ws[f'A{start_row}'].font = SECTION_FONT
    ws[f'A{start_row}'].fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
    ws[f'A{start_row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(f'A{start_row}:G{start_row}')
    start_row += 2
    
    # Column headers
    # Category header (column A)
    cat_header = ws.cell(row=start_row, column=1)
    cat_header.value = 'Category'
    cat_header.font = SUBHEADER_FONT
    cat_header.fill = SUBHEADER_FILL
    cat_header.alignment = Alignment(horizontal='center', vertical='center')
    cat_header.border = BORDER_THIN
    
    # Missing Types header (columns B:G, merged) with borders on all cells
    ws.merge_cells(f'B{start_row}:G{start_row}')
    for col in range(2, 8):  # Columns B through G
        cell = ws.cell(row=start_row, column=col)
        cell.border = BORDER_THIN
        cell.fill = SUBHEADER_FILL
    # Set content on the first merged cell
    types_header = ws.cell(row=start_row, column=2)
    types_header.value = 'Missing Types'
    types_header.font = SUBHEADER_FONT
    types_header.alignment = Alignment(horizontal='center', vertical='center')
    
    start_row += 1
    
    # Add condensed data (one row per category)
    for layout_key, progress in apartment_progress.items():
        missing_types = progress.get('missing_types', [])
        greylisted_missing = progress.get('greylisted_missing_types', [])
        display_name = progress['display_name']
        
        # Category name in column A with border
        cat_cell = ws.cell(row=start_row, column=1, value=display_name)
        cat_cell.font = Font(name='Calibri', size=10, bold=True)
        cat_cell.alignment = Alignment(horizontal='left', vertical='center')
        cat_cell.border = BORDER_THIN
        
        # Missing types - merged cells B:G, TRUNCATED to first 10 types
        if missing_types:
            # Show first 10 types, indicate if more exist
            types_str = ', '.join(missing_types[:10])
            if len(missing_types) > 10:
                types_str += f' ... (+{len(missing_types) - 10} more)'
            # Add note about greylisted if any
            if greylisted_missing:
                types_str += f' (+ {len(greylisted_missing)} optional)'
            
            # Merge cells B:G and apply borders to ALL merged cells
            ws.merge_cells(f'B{start_row}:G{start_row}')
            for col in range(2, 8):  # Columns B through G
                cell = ws.cell(row=start_row, column=col)
                cell.border = BORDER_THIN
            # Set the content on the first merged cell (B)
            cell = ws.cell(row=start_row, column=2)
            cell.value = types_str
            cell.font = Font(name='Calibri', size=9, color='C00000')
            cell.alignment = Alignment(horizontal='left', vertical='center')
            
        elif greylisted_missing:
            # Only greylisted missing
            types_str = f'{len(greylisted_missing)} optional type(s) not provided'
            ws.merge_cells(f'B{start_row}:G{start_row}')
            for col in range(2, 8):  # Columns B through G
                cell = ws.cell(row=start_row, column=col)
                cell.border = BORDER_THIN
            cell = ws.cell(row=start_row, column=2)
            cell.value = types_str
            cell.font = Font(name='Calibri', size=9, color='999999', italic=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        else:
            ws.merge_cells(f'B{start_row}:G{start_row}')
            for col in range(2, 8):  # Columns B through G
                cell = ws.cell(row=start_row, column=col)
                cell.border = BORDER_THIN
            cell = ws.cell(row=start_row, column=2)
            cell.value = 'All types covered ✓'
            cell.font = Font(name='Calibri', size=9, color='70AD47', bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        start_row += 1
    
    return start_row + 2


def add_duplicates_section(ws, apartment_progress, start_row):
    """Add section showing types with alternative/duplicate layouts."""
    
    # Check if there are any duplicates
    has_duplicates = any(progress['duplicate_count'] > 0 for progress in apartment_progress.values())
    
    if not has_duplicates:
        return start_row
    
    # Section header
    ws[f'A{start_row}'] = 'APARTMENT TYPES WITH ALTERNATIVE LAYOUTS'
    ws[f'A{start_row}'].font = SECTION_FONT
    ws[f'A{start_row}'].fill = PatternFill(start_color='FF6600', end_color='FF6600', fill_type='solid')
    ws[f'A{start_row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(f'A{start_row}:H{start_row}')
    start_row += 2
    
    ws[f'A{start_row}'] = 'Some apartment types have multiple layout variations (e.g., for different plot numbers or design alternatives).'
    ws[f'A{start_row}'].font = Font(name='Calibri', size=9, italic=True, color='666666')
    ws.merge_cells(f'A{start_row}:H{start_row}')
    start_row += 2
    
    for layout_key, progress in apartment_progress.items():
        duplicate_types = progress.get('duplicate_types', [])
        duplicate_count = progress['duplicate_count']
        
        if duplicate_types:
            display_name = progress['display_name']
            ws[f'A{start_row}'] = f'{display_name}:'
            ws[f'A{start_row}'].font = Font(name='Calibri', size=10, bold=True, color='FF6600')
            ws[f'B{start_row}'] = f"{duplicate_count} alternative layout(s) for {len(duplicate_types)} type(s)"
            ws[f'B{start_row}'].font = Font(name='Calibri', size=10)
            start_row += 1
            
            # List types with alternatives
            types_str = ', '.join(duplicate_types[:15])
            if len(duplicate_types) > 15:
                types_str += f' ... and {len(duplicate_types) - 15} more'
            ws[f'B{start_row}'] = types_str
            ws[f'B{start_row}'].font = Font(name='Calibri', size=9, color='666666')
            ws.merge_cells(f'B{start_row}:H{start_row}')
            start_row += 2
    
    return start_row + 2


def add_communal_layouts_section(ws, communal_progress, start_row, project_structure=None):
    """Add detailed communal layout tracking table."""
    
    # Section header
    ws[f'A{start_row}'] = 'COMMUNAL LAYOUT CATEGORIES'
    ws[f'A{start_row}'].font = SECTION_FONT
    ws[f'A{start_row}'].fill = SECTION_FILL
    ws[f'A{start_row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(f'A{start_row}:G{start_row}')
    start_row += 1
    
    # Column headers - simplified
    headers = ['Layout Category', 'Block', 'Documents', 'Progress Bar']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx)
        cell.value = header
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER_THIN
    
    # Merge the progress bar header from D to G
    ws.merge_cells(f'D{start_row}:G{start_row}')
    progress_bar_cell = ws.cell(row=start_row, column=4)
    progress_bar_cell.value = 'Progress Bar'
    progress_bar_cell.font = SUBHEADER_FONT
    progress_bar_cell.fill = SUBHEADER_FILL
    progress_bar_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add borders to progress bar header cells (D through G = columns 4-7)
    for col_num in range(4, 8):
        cell = ws.cell(row=start_row, column=col_num)
        cell.border = BORDER_THIN
    
    start_row += 1
    
    # Add summary data for each layout category
    for layout_key, progress in communal_progress.items():
        display_name = progress['display_name']
        coverage_pct = progress['coverage_percentage']
        total_floors = progress['total_expected_floors']
        doc_count = progress['document_count']
        
        # Layout category name
        ws.cell(row=start_row, column=1, value=display_name).font = Font(name='Calibri', size=10, bold=True)
        
        # Block name - centered (combining all blocks for summary view)
        ws.cell(row=start_row, column=2, value='All Blocks').alignment = Alignment(horizontal='center', vertical='center')
        
        # Documents count - centered
        coverage_pct = progress['coverage_percentage']
        docs_cell = ws.cell(row=start_row, column=3, value=doc_count)
        docs_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Merge cells for progress bar (D through G)
        ws.merge_cells(f'D{start_row}:G{start_row}')
        
        # Create visual progress bar using data bar
        add_progress_bar(ws, f'D{start_row}', coverage_pct)
        
        # Add border to entire merged progress bar area (D through G = columns 4-7)
        for col_num in range(4, 8):
            cell = ws.cell(row=start_row, column=col_num)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        start_row += 1
    
    start_row += 2
    
    # Add detailed block coverage section
    ws[f'A{start_row}'] = 'BLOCK COVERAGE BREAKDOWN'
    ws[f'A{start_row}'].font = SECTION_FONT
    ws[f'A{start_row}'].fill = SECTION_FILL
    ws[f'A{start_row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(f'A{start_row}:G{start_row}')
    start_row += 1
    
    # Block coverage headers (simplified layout)
    block_headers = ['Layout Category', 'Block', 'Covered Floors', 'Missing Floors', 'Status']
    
    for col_idx, header in enumerate(block_headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER_THIN
    start_row += 1
    
    # Add block-by-block data
    for layout_key, progress in communal_progress.items():
        display_name = progress['display_name']
        expected_by_block = progress.get('expected_floors_by_block', {})
        coverage_by_block = progress.get('coverage_by_block', {})
        greylisted_blocks = set(progress.get('greylisted_blocks', []))
        
        # Get ALL blocks from PROJECT_STRUCTURE (not just those with expected floors)
        if project_structure and 'blocks' in project_structure:
            all_blocks = sorted(project_structure['blocks'].keys())
        else:
            # Fallback to blocks in expected_by_block
            all_blocks = sorted(expected_by_block.keys())
        
        # Layout category name (only show once per category, vertically centered)
        if all_blocks:
            category_cell = ws.cell(row=start_row, column=1, value=display_name)
            category_cell.font = Font(name='Calibri', size=10, bold=True)
            category_cell.alignment = Alignment(horizontal='center', vertical='center')
            # Merge cells vertically for this category
            if len(all_blocks) > 1:
                ws.merge_cells(f'A{start_row}:A{start_row + len(all_blocks) - 1}')
        
        for block in all_blocks:
            expected_floors = expected_by_block.get(block, [])
            covered_floors = coverage_by_block.get(block, [])
            is_greylisted = block in greylisted_blocks
            
            # For greylisted blocks: show ALL covered floors, no missing floors (it's optional!)
            # For required blocks: only count floors that are both expected AND covered
            if is_greylisted:
                covered_count = len(covered_floors)  # Show all covered floors for optional blocks
                missing_floors = set()  # No missing floors for optional blocks
            else:
                covered_expected_floors = set(expected_floors) & set(covered_floors)
                covered_count = len(covered_expected_floors)  # Use intersection for required blocks
                missing_floors = set(expected_floors) - set(covered_floors)
            
            expected_count = len(expected_floors)
            missing_count = len(missing_floors)
            
            # Block name (centered)
            # Only show as grey/italic if greylisted AND no layouts
            block_cell = ws.cell(row=start_row, column=2, value=f"Block {block}")
            block_cell.alignment = Alignment(horizontal='center', vertical='center')
            if is_greylisted and covered_count == 0:
                block_cell.font = Font(name='Calibri', size=10, italic=True, color='999999')
            
            # Covered floors (centered)
            covered_cell = ws.cell(row=start_row, column=3, value=covered_count)
            covered_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Missing floors - show actual floor numbers instead of count (centered)
            # Only show as grey/italic if greylisted AND no layouts
            if missing_count == 0:
                missing_text = "None"
                missing_cell = ws.cell(row=start_row, column=4, value=missing_text)
            else:
                missing_list = sorted(list(missing_floors))
                missing_text = ', '.join(map(str, missing_list[:10]))  # Limit to 10 floors for readability
                if len(missing_list) > 10:
                    missing_text += f' ... and {len(missing_list) - 10} more'
                missing_cell = ws.cell(row=start_row, column=4, value=missing_text)
                # Show in red if not greylisted OR if greylisted but has layouts
                if not is_greylisted or covered_count > 0:
                    missing_cell.font = Font(name='Calibri', size=10, color='C00000')
                else:
                    missing_cell.font = Font(name='Calibri', size=10, color='999999', italic=True)
            missing_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Status - improved text with greylist support
            if is_greylisted:
                # For optional blocks, show coverage status (never "missing")
                if covered_count == 0:
                    status = "N/A (Optional)"
                    status_cell = ws.cell(row=start_row, column=5, value=status)
                    status_cell.font = Font(name='Calibri', size=10, color='999999', italic=True)
                else:
                    status = f"✓ {covered_count} floor(s) (Optional)"
                    status_cell = ws.cell(row=start_row, column=5, value=status)
                    status_cell.font = Font(name='Calibri', size=10, color='00B050')
            elif missing_count == 0 and expected_count > 0:
                status = "✓ Complete"
                status_cell = ws.cell(row=start_row, column=5, value=status)
                status_cell.font = Font(name='Calibri', size=10, bold=True, color='00B050')
            elif expected_count == 0:
                status = "N/A"
                status_cell = ws.cell(row=start_row, column=5, value=status)
                status_cell.font = Font(name='Calibri', size=10, color='666666')
            else:
                status = f"Missing {missing_count} floors"
                status_cell = ws.cell(row=start_row, column=5, value=status)
                status_cell.font = Font(name='Calibri', size=10, color='C00000')
            
            start_row += 1
        
        # Add spacing between layout categories
        start_row += 1
    
    return start_row


def add_missing_types_section(ws, apartment_progress, start_row):
    """Add section showing missing apartment types by category."""
    
    # Section header
    ws[f'A{start_row}'] = 'MISSING APARTMENT TYPES BY CATEGORY'
    ws[f'A{start_row}'].font = SECTION_FONT
    ws[f'A{start_row}'].fill = SECTION_FILL
    ws[f'A{start_row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(f'A{start_row}:H{start_row}')
    start_row += 2
    
    for layout_key, progress in apartment_progress.items():
        missing_types = progress.get('missing_types', [])
        greylisted_missing = progress.get('greylisted_missing_types', [])
        display_name = progress['display_name']
        
        if missing_types:
            ws[f'A{start_row}'] = f'{display_name}:'
            ws[f'A{start_row}'].font = Font(name='Calibri', size=10, bold=True, color='4472C4')
            ws[f'B{start_row}'] = f"{len(missing_types)} type(s) missing"
            ws[f'B{start_row}'].font = Font(name='Calibri', size=10)
            start_row += 1
            
            # List ALL missing types (no truncation) with wrapping
            types_str = ', '.join(missing_types)
            # Add greylisted info if any
            if greylisted_missing:
                types_str += f' (+ {len(greylisted_missing)} optional)'
            
            # Merge cells and enable wrapping
            ws.merge_cells(f'B{start_row}:H{start_row}')
            cell = ws[f'B{start_row}']
            cell.value = types_str
            cell.font = Font(name='Calibri', size=9, color='666666')
            cell.alignment = Alignment(horizontal='left', wrap_text=True, vertical='top')
            
            # Auto-adjust row height based on content length
            estimated_chars = len(types_str)
            column_width = 100  # Approximate character width for merged B:H
            lines_needed = max(1, (estimated_chars // column_width) + 1)
            ws.row_dimensions[start_row].height = max(20, lines_needed * 15)
            
            start_row += 2
        elif greylisted_missing:
            # Only greylisted missing
            ws[f'A{start_row}'] = f'{display_name}:'
            ws[f'A{start_row}'].font = Font(name='Calibri', size=10, bold=True, color='4472C4')
            ws[f'B{start_row}'] = f"{len(greylisted_missing)} optional type(s) not provided"
            ws[f'B{start_row}'].font = Font(name='Calibri', size=10, color='999999', italic=True)
            start_row += 1
            types_str = ', '.join(greylisted_missing[:20])
            ws[f'B{start_row}'] = types_str
            ws[f'B{start_row}'].font = Font(name='Calibri', size=9, color='999999', italic=True)
            ws.merge_cells(f'B{start_row}:H{start_row}')
            start_row += 2
        else:
            ws[f'A{start_row}'] = f'{display_name}: All types covered ✓'
            ws[f'A{start_row}'].font = Font(name='Calibri', size=10, bold=True, color='70AD47')
            start_row += 2
    
    return start_row


def add_alternatives_section(ws, apartment_progress, start_row):
    """Add section showing types with alternative/duplicate layouts."""
    
    # Check if there are any duplicates
    has_duplicates = any(progress['duplicate_count'] > 0 for progress in apartment_progress.values())
    
    if not has_duplicates:
        return start_row
    
    # Section header
    ws[f'A{start_row}'] = 'APARTMENT TYPES WITH ALTERNATIVE LAYOUTS'
    ws[f'A{start_row}'].font = SECTION_FONT
    ws[f'A{start_row}'].fill = PatternFill(start_color='FF6600', end_color='FF6600', fill_type='solid')
    ws[f'A{start_row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(f'A{start_row}:H{start_row}')
    start_row += 2
    
    ws[f'A{start_row}'] = 'Some apartment types have multiple layout variations (e.g., for different plot numbers or design alternatives).'
    ws[f'A{start_row}'].font = Font(name='Calibri', size=9, italic=True, color='666666')
    ws.merge_cells(f'A{start_row}:H{start_row}')
    start_row += 2
    
    for layout_key, progress in apartment_progress.items():
        duplicate_types = progress.get('duplicate_types', [])
        duplicate_count = progress['duplicate_count']
        
        if duplicate_types:
            display_name = progress['display_name']
            ws[f'A{start_row}'] = f'{display_name}:'
            ws[f'A{start_row}'].font = Font(name='Calibri', size=10, bold=True, color='FF6600')
            ws[f'B{start_row}'] = f"{duplicate_count} alternative layout(s) for {len(duplicate_types)} type(s)"
            ws[f'B{start_row}'].font = Font(name='Calibri', size=10)
            start_row += 1
            
            # List types with alternatives
            types_str = ', '.join(duplicate_types[:15])
            if len(duplicate_types) > 15:
                types_str += f' ... and {len(duplicate_types) - 15} more'
            ws[f'B{start_row}'] = types_str
            ws[f'B{start_row}'].font = Font(name='Calibri', size=9, color='666666')
            ws.merge_cells(f'B{start_row}:H{start_row}')
            start_row += 2
    
    return start_row


def add_block_coverage_details_section(ws, communal_progress, start_row):
    """Add block coverage details without example documents."""
    
    # Section header
    ws[f'A{start_row}'] = 'BLOCK COVERAGE DETAILS'
    ws[f'A{start_row}'].font = SECTION_FONT
    ws[f'A{start_row}'].fill = SECTION_FILL
    ws[f'A{start_row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(f'A{start_row}:H{start_row}')
    start_row += 2
    
    for layout_key, progress in communal_progress.items():
        display_name = progress['display_name']
        coverage_by_block = progress.get('coverage_by_block', {})
        expected_floors_by_block = progress.get('expected_floors_by_block', {})
        
        if coverage_by_block:
            # Layout category header
            ws[f'A{start_row}'] = f'{display_name} Block Coverage:'
            ws[f'A{start_row}'].font = Font(name='Calibri', size=11, bold=True, color='4472C4')
            start_row += 1
            
            # Block coverage details
            for block_name in sorted(coverage_by_block.keys()):
                covered_floors = set(coverage_by_block.get(block_name, []))
                expected_floors = set(expected_floors_by_block.get(block_name, []))
                missing_floors = expected_floors - covered_floors
                
                # Block row
                ws[f'A{start_row}'] = f'  Block {block_name}:'
                ws[f'A{start_row}'].font = Font(name='Calibri', size=10, bold=True)
                
                if missing_floors:
                    ws[f'B{start_row}'] = f'Covered: {len(covered_floors)} floors, Missing: {len(missing_floors)} floor(s)'
                    ws[f'B{start_row}'].font = Font(name='Calibri', size=10, color='D32F2F')
                    
                    # Show specific missing floors
                    missing_list = sorted(list(missing_floors))
                    missing_str = ', '.join(map(str, missing_list[:10]))  # Show first 10
                    if len(missing_list) > 10:
                        missing_str += f' ... and {len(missing_list) - 10} more'
                    
                    ws[f'C{start_row}'] = f'Missing floors: {missing_str}'
                    ws[f'C{start_row}'].font = Font(name='Calibri', size=9, color='D32F2F')
                    ws.merge_cells(f'C{start_row}:H{start_row}')
                else:
                    ws[f'B{start_row}'] = f'Covered: {len(covered_floors)} floors ✓'
                    ws[f'B{start_row}'].font = Font(name='Calibri', size=10, color='70AD47')
                
                start_row += 1
            
            start_row += 1
    
    return start_row


def add_details_tab(wb, layout_summary, apartment_progress, communal_progress):
    """Add detailed breakdown tab with missing types, alternatives, and block coverage."""
    
    ws = wb.create_sheet("Details")
    
    # Report header
    ws['A1'] = "Layout Tracking Details"
    ws['A1'].font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='203864', end_color='203864', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:H1')
    ws.row_dimensions[1].height = 25
    
    current_row = 5
    
    # Missing apartment types section
    if apartment_progress:
        current_row = add_missing_types_section(ws, apartment_progress, current_row)
        current_row += 2
        
        # Alternative layouts section
        current_row = add_alternatives_section(ws, apartment_progress, current_row)
        current_row += 2
    
    # Block coverage details removed - now shown on main page only
    # This keeps the details page focused on apartment types (missing types, alternatives)
    
    # Set column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 60
    ws.column_dimensions['H'].width = 10
    
    # Page setup
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = 9  # A4
    ws.print_options.horizontalCentered = True


def add_categorized_tab(wb, layout_summary):
    """Add detailed categorized layouts tab for debugging."""
    
    categorized_count = layout_summary.get('categorized', 0)
    
    if categorized_count == 0:
        return
    
    # Create new sheet
    ws = wb.create_sheet("Categorized Layouts")
    
    # Title
    ws['A1'] = 'Categorized Layout Drawings (Debug View)'
    ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='203864', end_color='203864', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:J1')
    ws.row_dimensions[1].height = 25
    
    # Summary
    ws['A3'] = f'Total Categorized Layouts: {categorized_count}'
    ws['A3'].font = Font(name='Calibri', size=12, bold=True, color='203864')
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A3:J3')
    
    ws['A4'] = 'This view shows all successfully categorized layouts with detection data for debugging filter logic.'
    ws['A4'].font = Font(name='Calibri', size=9, italic=True, color='666666')
    ws['A4'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A4:J4')
    
    # Headers
    headers = ['Doc Title', 'Doc Ref', 'Category', 'Layout Type', 'Apt Type / Block', 'Floor Coverage', 'File Type', 'Status', 'Rev', 'Date']
    row = 6
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER_THIN
    row += 1
    
    # Get categorized data
    categorized_data = layout_summary.get('categorized_data')
    if categorized_data is not None and not categorized_data.empty:
        categorized = categorized_data[categorized_data['category'].notna()]
        
        # Sort by category, then layout type, then apartment type/block
        categorized = categorized.sort_values(by=['category', 'layout_type', 'apartment_type', 'block'], na_position='last')
        
        for idx, layout in categorized.iterrows():
            category = layout.get('category', '')
            
            # Column 1: Doc Title
            cell1 = ws.cell(row=row, column=1, value=layout.get('Doc Title', ''))
            cell1.alignment = Alignment(horizontal='left', vertical='center')
            
            # Column 2: Doc Ref
            cell2 = ws.cell(row=row, column=2, value=layout.get('Doc Ref', ''))
            cell2.alignment = Alignment(horizontal='center', vertical='center')
            
            # Column 3: Category (apartment/communal)
            cell3 = ws.cell(row=row, column=3, value=str(category).title())
            cell3.alignment = Alignment(horizontal='center', vertical='center')
            if category == 'apartment':
                cell3.fill = PatternFill(start_color='E7F3FF', end_color='E7F3FF', fill_type='solid')
            elif category == 'communal':
                cell3.fill = PatternFill(start_color='FFF3E7', end_color='FFF3E7', fill_type='solid')
            
            # Column 4: Layout Type
            cell4 = ws.cell(row=row, column=4, value=layout.get('layout_type', ''))
            cell4.alignment = Alignment(horizontal='left', vertical='center')
            
            # Column 5: Apartment Type OR Block (depending on category)
            # Just display pre-calculated data from analyzer - no detection here
            if category == 'apartment':
                apt_type = layout.get('apartment_type', '')
                cell5 = ws.cell(row=row, column=5, value=f"Type: {apt_type}" if apt_type else '')
            elif category == 'communal':
                # Block already extracted by analyzer - just display it
                block = layout.get('block', '')
                cell5 = ws.cell(row=row, column=5, value=f"Block: {block}" if block else '')
            else:
                cell5 = ws.cell(row=row, column=5, value='')
            cell5.alignment = Alignment(horizontal='center', vertical='center')
            
            # Column 6: Floor Coverage (for communal only)
            # Floor coverage already extracted by analyzer - just format for display
            if category == 'communal':
                floor_coverage = layout.get('floor_coverage', '')
                if floor_coverage:
                    try:
                        # Parse stored floor list and display
                        floors = eval(floor_coverage) if isinstance(floor_coverage, str) else floor_coverage
                        if floors:
                            floors_str = ', '.join(map(str, sorted(floors)))
                            cell = ws.cell(row=row, column=6, value=floors_str)
                            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        else:
                            ws.cell(row=row, column=6, value='')
                    except:
                        # If parsing fails, just display as-is
                        cell = ws.cell(row=row, column=6, value=str(floor_coverage))
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                else:
                    ws.cell(row=row, column=6, value='')
            else:
                ws.cell(row=row, column=6, value='')
            
            # Column 7: File Type
            cell7 = ws.cell(row=row, column=7, value=layout.get('File Type', ''))
            cell7.alignment = Alignment(horizontal='center', vertical='center')
            
            # Column 8: Status
            cell8 = ws.cell(row=row, column=8, value=layout.get('Status', ''))
            cell8.alignment = Alignment(horizontal='center', vertical='center')
            
            # Column 9: Rev
            cell9 = ws.cell(row=row, column=9, value=layout.get('Rev', ''))
            cell9.alignment = Alignment(horizontal='center', vertical='center')
            
            # Column 10: Date
            cell10 = ws.cell(row=row, column=10, value=layout.get('Date (WET)', ''))
            cell10.alignment = Alignment(horizontal='center', vertical='center')
            
            # Alternate row colors
            if row % 2 == 0:
                for col in range(1, 11):
                    current_cell = ws.cell(row=row, column=col)
                    # Only apply if no category color already applied
                    if current_cell.fill.start_color.index == '00000000':
                        current_cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            
            row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 60  # Doc Title
    ws.column_dimensions['B'].width = 20  # Doc Ref
    ws.column_dimensions['C'].width = 12  # Category
    ws.column_dimensions['D'].width = 30  # Layout Type
    ws.column_dimensions['E'].width = 20  # Apt Type / Block
    ws.column_dimensions['F'].width = 25  # Floor Coverage (narrower to allow wrapping for long floor lists)
    ws.column_dimensions['G'].width = 15  # File Type
    ws.column_dimensions['H'].width = 15  # Status
    ws.column_dimensions['I'].width = 8   # Rev
    ws.column_dimensions['J'].width = 12  # Date
    
    # Freeze panes
    ws.freeze_panes = 'A7'


def add_uncategorized_tab(wb, layout_summary):
    """Add detailed uncategorized layouts tab (like certificate report)."""
    
    uncategorized_count = layout_summary.get('uncategorized', 0)
    
    if uncategorized_count == 0:
        return
    
    # Create new sheet
    ws = wb.create_sheet("Uncategorized Layouts")
    
    # Title
    ws['A1'] = 'Uncategorized Layout Drawings'
    ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:F1')
    ws.row_dimensions[1].height = 25
    
    # Summary
    ws['A3'] = f'Total Uncategorized Layouts: {uncategorized_count}'
    ws['A3'].font = Font(name='Calibri', size=12, bold=True, color='C00000')
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A3:F3')
    
    ws['A4'] = 'These drawings were not matched to any configured layout category. Review titles to identify new patterns.'
    ws['A4'].font = Font(name='Calibri', size=9, italic=True, color='666666')
    ws['A4'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A4:F4')
    
    # Headers
    headers = ['Doc Title', 'Doc Ref', 'File Type', 'Status', 'Rev', 'Date']
    row = 6
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER_THIN
    row += 1
    
    # Get uncategorized data
    categorized_data = layout_summary.get('categorized_data')
    if categorized_data is not None and not categorized_data.empty:
        uncategorized = categorized_data[categorized_data['category'].isna()]
        
        for idx, layout in uncategorized.iterrows():
            ws.cell(row=row, column=1, value=layout.get('Doc Title', ''))
            ws.cell(row=row, column=2, value=layout.get('Doc Ref', ''))
            ws.cell(row=row, column=3, value=layout.get('File Type', ''))
            ws.cell(row=row, column=4, value=layout.get('Status', ''))
            ws.cell(row=row, column=5, value=layout.get('Rev', ''))
            ws.cell(row=row, column=6, value=layout.get('Date (WET)', ''))
            
            # Alternate row colors
            if row % 2 == 0:
                for col in range(1, 7):
                    ws.cell(row=row, column=col).fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            
            row += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = None
        for cell in column:
            try:
                if hasattr(cell, 'column_letter'):
                    if column_letter is None:
                        column_letter = cell.column_letter
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except:
                pass
        
        if column_letter is not None:
            adjusted_width = min(max_length + 2, 80)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze panes
    ws.freeze_panes = 'A7'


def save_layout_report(latest_data, output_file, config):
    """
    Generate and save layout tracking report to Excel.
    
    Args:
        latest_data: DataFrame of latest documents
        output_file: Path to output Excel file
        config: Project configuration dictionary
        
    Returns:
        True if successful, False otherwise
    """
    try:
        # Get layout tracking config
        layout_tracking = config.get('APARTMENT_LAYOUT_TRACKING', {})
        
        if not layout_tracking.get('enabled', False):
            print("⚠️  Layout tracking not enabled for this project")
            return False
        
        # Get accommodation data
        accommodation_data = config.get('ACCOMMODATION_DATA', {})
        
        # Get project structure
        project_structure = config.get('PROJECT_STRUCTURE', {})
        
        # Get layout tracking summary
        print("📊 Analyzing layout drawings...")
        layout_summary = get_layout_tracking_summary(latest_data, layout_tracking, accommodation_data, project_structure)
        
        if not layout_summary or layout_summary.get('total_layouts', 0) == 0:
            print("ℹ️  No layout drawings found - generating empty report")
            # Create empty report structure
            layout_summary = {
                'total_layouts': 0,
                'categorized': 0,
                'uncategorized': 0,
                'withdrawn_count': 0,
                'apartment_progress': {},
                'communal_progress': {},
                'uncategorized_layouts': []
            }
        
        print(f"✓ Found {layout_summary['total_layouts']} layout drawings")
        print(f"  - Withdrawn (excluded): {layout_summary.get('withdrawn_count', 0)}")
        print(f"  - Categorized: {layout_summary['categorized']}")
        print(f"  - Uncategorized: {layout_summary['uncategorized']}")
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Layout Summary"
        
        # Report header
        project_name = config.get('PROJECT_TITLE', 'Unknown Project')
        ws['A1'] = f"{project_name} - Apartment/Communal Layouts Tracking Report"
        ws['A1'].font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color='203864', end_color='203864', fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:G1')
        ws.row_dimensions[1].height = 25
        
        # Report date - formatted like certificate report
        now = datetime.now()
        day = now.day
        suffix = 'th' if 11 <= day <= 13 else {1: 'st', 2: 'nd', 3: 'rd'}.get(day % 10, 'th')
        date_str = f"{day}{suffix} {now.strftime('%B %Y')}"
        
        ws['A2'] = date_str
        ws['A2'].font = Font(name='Calibri', size=10, italic=True)
        ws['A2'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A2:G2')
        
        # Check if report is empty
        if layout_summary.get('total_layouts', 0) == 0:
            # Add a message explaining the empty report
            ws['A4'] = 'No Layout Drawings Found'
            ws['A4'].font = Font(name='Calibri', size=14, bold=True, color='FF0000')
            ws['A4'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A4:G4')
            
            ws['A6'] = 'This report is empty because:'
            ws['A6'].font = Font(name='Calibri', size=11, bold=True)
            
            ws['A7'] = '1. No layout drawings were detected in the document register, OR'
            ws['A8'] = '2. Layout tracking detection patterns are not configured yet'
            ws['A10'] = 'To configure layout tracking, edit your project config file:'
            ws['A11'] = '  • Set file_type_patterns (e.g., ["DR"] for drawings)'
            ws['A12'] = '  • Define apartment_layouts and communal_layouts patterns'
            ws['A13'] = '  • Configure PROJECT_STRUCTURE with block floor expectations'
            
            ws['A15'] = 'See GreenwichPeninsula.py for a complete example configuration.'
            ws['A15'].font = Font(name='Calibri', size=11, italic=True)
            
            # Set column width
            ws.column_dimensions['A'].width = 80
        else:
            current_row = add_overview_section(ws, layout_summary, accommodation_data)
        
            # Add apartment layout section (summary only)
            apartment_progress = layout_summary.get('apartment_progress', {})
            if apartment_progress:
                current_row = add_apartment_layouts_section(ws, apartment_progress, current_row)
                
                # Add condensed missing types summary right after apartment layouts
                current_row = add_missing_types_summary(ws, apartment_progress, current_row)
            
            # Add communal layout section (summary only)
            communal_progress = layout_summary.get('communal_progress', {})
            if communal_progress:
                current_row = add_communal_layouts_section(ws, communal_progress, current_row, project_structure)
        
            # Set column widths - adjusted for new layout
            ws.column_dimensions['A'].width = 35
            ws.column_dimensions['B'].width = 20  # Fixed width for types found/block
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 12
            
            # Add details tab with missing types, alternatives, and block coverage (only if we have data)
            apartment_progress = layout_summary.get('apartment_progress', {})
            communal_progress = layout_summary.get('communal_progress', {})
            add_details_tab(wb, layout_summary, apartment_progress, communal_progress)
            
            # Add categorized tab (for debugging)
            add_categorized_tab(wb, layout_summary)
            
            # Add uncategorized tab
            add_uncategorized_tab(wb, layout_summary)
        
        # Page setup for printing
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.paperSize = 9  # A4
        ws.print_options.horizontalCentered = True
        
        # Save workbook
        wb.save(output_file)
        
        # Post-process to fix data bar issues (remove gaps, allow gradient)
        from utils.openpyxl_databars_fixer import process_report_databars
        process_report_databars(output_file)
        
        print(f"\n✅ Layout tracking report saved: {output_file}")
        return True
        
    except Exception as e:
        print(f"\n❌ Error generating layout report: {str(e)}")
        import traceback
        traceback.print_exc()
        return False
