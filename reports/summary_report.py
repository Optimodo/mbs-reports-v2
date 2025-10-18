"""Summary report generation module."""

import warnings
import pandas as pd
from datetime import datetime
from pathlib import Path
import time
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.page import PageMargins
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.fill import ColorChoice, PatternFillProperties

from styles.formatting import (
    OVERALL_SUMMARY_STYLES,
    STATUS_STYLES,
    apply_status_style
)
from utils.status_mapping import (
    get_status_category,
    get_status_color,
    get_grouped_status_counts,
    get_status_display_order
)

# Suppress openpyxl warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')


def get_chart_safe_color(config_color, category):
    """Convert config colors to chart-safe colors.
    
    Ensures that white/very light colors are converted to visible colors for charts.
    
    Args:
        config_color: Color from config (hex string)
        category: Status category name for fallback mapping
        
    Returns:
        str: Chart-safe hex color
    """
    # Define chart-safe colors for common white/light statuses
    chart_safe_mapping = {
        'Other': 'C0C0C0',        # Light gray
        'Information': 'D3D3D3',   # Light gray
        'Review': 'E6E6FA',       # Lavender
        'IFC-pending': 'F0E68C',  # Khaki
        'Under Review': 'DDA0DD', # Plum
        'Shared': '98FB98',       # Pale green
        'Published': '87CEEB',    # Sky blue
    }
    
    # Check if the color is too light (white or very light colors)
    light_colors = ['FFFFFF', 'FFFFF0', 'FFFACD', 'FFF8DC', 'F5F5DC', 'FDF5E6']
    
    if config_color.upper() in light_colors:
        # Use category-specific fallback or generic light gray
        return chart_safe_mapping.get(category, 'C0C0C0')
    else:
        # Use the original color if it's dark enough
        return config_color


def save_excel_with_retry(summary_df, changes_df, latest_data_df, output_file, config, max_retries=3):
    """Try to save the Excel file with retries.
    
    This function generates a comprehensive summary Excel report with:
    - Overall Summary sheet with charts
    - Summary Data sheet with historical data
    - Latest Data sheet with current document listing
    
    Args:
        summary_df: DataFrame with summary data over time
        changes_df: DataFrame with change tracking data (unused currently)
        latest_data_df: DataFrame with latest document data
        output_file: Path to output Excel file
        config: Project configuration dictionary
        max_retries: Maximum number of retry attempts
        
    Returns:
        bool: True if successful, False otherwise
    """
    for attempt in range(max_retries):
        try:
            # Try to load existing file
            try:
                book = load_workbook(output_file)
                # Remove existing Overall Summary sheet if it exists (pandas won't replace it)
                if 'Overall Summary' in book.sheetnames:
                    book.remove(book['Overall Summary'])
                    book.save(output_file)  # Save the removal
                # Remove existing Summary Data sheet if it exists
                if 'Summary Data' in book.sheetnames:
                    book.remove(book['Summary Data'])
                # Remove existing Latest Data sheet if it exists
                if 'Latest Data' in book.sheetnames:
                    book.remove(book['Latest Data'])
                with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    summary_df.to_excel(writer, sheet_name='Summary Data', index=False)
                    latest_data_df.to_excel(writer, sheet_name='Latest Data', index=False)
            except FileNotFoundError:
                # If file doesn't exist, create new one
                with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                    summary_df.to_excel(writer, sheet_name='Summary Data', index=False)
                    latest_data_df.to_excel(writer, sheet_name='Latest Data', index=False)
            
            # Create or update Overall Summary sheet first
            book = load_workbook(output_file)
            if 'Overall Summary' in book.sheetnames:
                book.remove(book['Overall Summary'])
            overall_summary = book.create_sheet('Overall Summary', 0)  # Create at index 0 (first position)
            
            # Set print layout: fit to 1 page, center, narrow margins
            overall_summary.page_setup.fitToWidth = 1
            overall_summary.page_setup.fitToHeight = 0
            overall_summary.page_setup.horizontalCentered = True
            overall_summary.page_setup.verticalCentered = True
            overall_summary.page_margins = PageMargins(left=0.25, right=0.25, top=0.75, bottom=0.75, header=0.3, footer=0.3)
            # Extra reliability for centering
            overall_summary.sheet_properties.pageSetUpPr.horizontalCentered = True
            overall_summary.sheet_properties.pageSetUpPr.verticalCentered = True
            
            # Merge and center the title across A1:O1
            overall_summary.merge_cells('A1:O1')
            
            # Set row height for the title row
            overall_summary.row_dimensions[1].height = 70

            # Add title with project name on first line and rest on second line
            project_title = config.get('PROJECT_TITLE', '')
            if project_title:
                title_text = f"{project_title}\nDocument Register Overall Summary"
            else:
                title_text = "Document Register Overall Summary"
            overall_summary['A1'] = title_text
            
            # Style exactly like the working test script
            overall_summary['A1'].font = Font(name='Calibri', size=14, bold=True, color='000000')
            overall_summary['A1'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            overall_summary['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Get the latest data from Summary Data
            summary_data = pd.read_excel(output_file, sheet_name='Summary Data')
            latest_row = summary_data.iloc[-1]  # Get the last row
            
            # Add data export timestamp from Summary Data (now in row 2)
            export_date = latest_row['Date'].strftime("%d-%m-%Y") if isinstance(latest_row['Date'], datetime) else latest_row['Date']
            export_time = latest_row['Time'].strftime("%H-%M-%S") if isinstance(latest_row['Time'], datetime) else latest_row['Time']
            overall_summary['A2'] = f'Data Export: {export_date} {export_time}'
            overall_summary['A2'].font = OVERALL_SUMMARY_STYLES['timestamp']['font']
            overall_summary['A2'].alignment = OVERALL_SUMMARY_STYLES['timestamp']['alignment']
            overall_summary['A2'].fill = OVERALL_SUMMARY_STYLES['timestamp']['fill']
            
            # Get all column names from Summary Data
            all_columns = summary_data.columns.tolist()
            
            # Add total documents
            overall_summary['A5'] = 'Total Documents:'
            # Get total from Latest Data sheet
            latest_data = pd.read_excel(output_file, sheet_name='Latest Data')
            total_docs = len(latest_data)  # len() already gives us the correct count
            overall_summary['B5'] = total_docs
            overall_summary['B5'].font = OVERALL_SUMMARY_STYLES['total_cell']['font']
            overall_summary['B5'].alignment = OVERALL_SUMMARY_STYLES['total_cell']['alignment']
            overall_summary['B5'].fill = OVERALL_SUMMARY_STYLES['total_cell']['fill']
            
            # Filter and sort revision columns
            rev_columns = [col for col in all_columns if col.startswith('Rev_')]
            
            # Group revisions by type (P, C, and other)
            # Note: Certificates are now excluded at the data filtering stage
            p_revs = sorted([col for col in rev_columns if col.startswith('Rev_P')], 
                          key=lambda x: int(x.split('_')[1][1:]) if x.split('_')[1][1:].isdigit() else float('inf'))
            
            c_revs = sorted([col for col in rev_columns if col.startswith('Rev_C')],
                          key=lambda x: int(x.split('_')[1][1:]) if x.split('_')[1][1:].isdigit() else float('inf'))
            
            other_revs = sorted([col for col in rev_columns if not (
                col.startswith('Rev_P') or 
                col.startswith('Rev_C')
            )])
            
            # Function to add revision and status summary
            def add_revision_summary(start_row, rev_columns, title):
                """Add a revision summary section with a combined status summary"""
                # Add section header
                overall_summary[f'A{start_row}'] = title
                overall_summary[f'A{start_row}'].font = OVERALL_SUMMARY_STYLES['section_header']['font']
                overall_summary[f'A{start_row}'].alignment = OVERALL_SUMMARY_STYLES['section_header']['alignment']
                overall_summary[f'A{start_row}'].fill = OVERALL_SUMMARY_STYLES['section_header']['fill']
                
                # Add revision headers
                overall_summary[f'A{start_row + 1}'] = 'Revision'
                overall_summary[f'B{start_row + 1}'] = 'Count'
                overall_summary[f'C{start_row + 1}'] = 'Status'
                overall_summary[f'D{start_row + 1}'] = 'Count'
                
                # Style headers
                for col in ['A', 'B', 'C', 'D']:
                    overall_summary[f'{col}{start_row + 1}'].font = OVERALL_SUMMARY_STYLES['column_header']['font']
                    overall_summary[f'{col}{start_row + 1}'].alignment = OVERALL_SUMMARY_STYLES['column_header']['alignment']
                    overall_summary[f'{col}{start_row + 1}'].fill = OVERALL_SUMMARY_STYLES['column_header']['fill']
                    overall_summary[f'{col}{start_row + 1}'].border = OVERALL_SUMMARY_STYLES['border']
                
                # Get status counts for all revisions in this group
                # Note: Certificates are already filtered out at the data loading stage
                status_counts = {}
                total_count = 0
                
                for rev_col in rev_columns:
                    rev_name = rev_col.replace('Rev_', '')
                    
                    # Get count from summary data
                    count = latest_row.get(rev_col, 0)
                    # Handle NaN values from pandas
                    total_count += 0 if pd.isna(count) else count
                    
                    # Count statuses for this revision
                    rev_data = latest_data[latest_data['Rev'] == rev_name]
                    
                    # Use get_grouped_status_counts to properly group raw status values
                    if not rev_data.empty:
                        grouped_counts = get_grouped_status_counts(rev_data['Status'], config)
                        for status, status_count in grouped_counts.items():
                            status_counts[status] = status_counts.get(status, 0) + status_count
                
                # Add revision data
                row = start_row + 2
                for rev_col in rev_columns:
                    rev_name = rev_col.replace('Rev_', '')
                    
                    # Get the count from the summary data
                    count = latest_row.get(rev_col, 0)
                    # Handle NaN values from pandas
                    count = 0 if pd.isna(count) else count
                    
                    # Only add revision row if there are actually documents with this revision
                    if count > 0:
                        # Add revision name and count
                        overall_summary[f'A{row}'] = rev_name
                        overall_summary[f'B{row}'] = count
                        overall_summary[f'A{row}'].font = OVERALL_SUMMARY_STYLES['data_cell']['font']
                        overall_summary[f'A{row}'].alignment = OVERALL_SUMMARY_STYLES['data_cell']['alignment']
                        overall_summary[f'A{row}'].fill = OVERALL_SUMMARY_STYLES['data_cell']['fill']
                        overall_summary[f'A{row}'].border = OVERALL_SUMMARY_STYLES['border']
                        overall_summary[f'B{row}'].font = OVERALL_SUMMARY_STYLES['data_cell']['font']
                        overall_summary[f'B{row}'].alignment = OVERALL_SUMMARY_STYLES['data_cell']['alignment']
                        overall_summary[f'B{row}'].fill = OVERALL_SUMMARY_STYLES['data_cell']['fill']
                        overall_summary[f'B{row}'].border = OVERALL_SUMMARY_STYLES['border']
                        row += 1
                
                # Add total row for revisions
                overall_summary[f'A{row}'] = 'Total'
                overall_summary[f'B{row}'] = total_count
                overall_summary[f'A{row}'].font = OVERALL_SUMMARY_STYLES['total_cell']['font']
                overall_summary[f'A{row}'].alignment = OVERALL_SUMMARY_STYLES['total_cell']['alignment']
                overall_summary[f'A{row}'].fill = OVERALL_SUMMARY_STYLES['total_cell']['fill']
                overall_summary[f'A{row}'].border = OVERALL_SUMMARY_STYLES['border']
                overall_summary[f'B{row}'].font = OVERALL_SUMMARY_STYLES['total_cell']['font']
                overall_summary[f'B{row}'].alignment = OVERALL_SUMMARY_STYLES['total_cell']['alignment']
                overall_summary[f'B{row}'].fill = OVERALL_SUMMARY_STYLES['total_cell']['fill']
                overall_summary[f'B{row}'].border = OVERALL_SUMMARY_STYLES['border']
                
                # Add status summary
                status_row = start_row + 2
                total_status_count = 0
                
                # Create ordered list using project-specific categories
                ordered_statuses = []
                
                # Database now contains grouped categories, so use them directly
                # Certificates are already filtered out at the data loading stage
                filtered_status_counts = status_counts
                
                # Build ordered list using project-specific display order
                if config and 'STATUS_DISPLAY_ORDER' in config:
                    # Use project-specific display order
                    display_order = config['STATUS_DISPLAY_ORDER']
                    
                    # Add statuses in the order defined by STATUS_DISPLAY_ORDER
                    for category in display_order:
                        if category in filtered_status_counts:
                            ordered_statuses.append((category, filtered_status_counts[category]))
                    
                    # Add any remaining statuses that weren't in display order
                    for status, count in filtered_status_counts.items():
                        if status not in display_order:
                            ordered_statuses.append((status, count))
                else:
                    # Fallback to alphabetical order if no display order defined
                    ordered_statuses = sorted(filtered_status_counts.items())
                
                for status, count in ordered_statuses:
                    total_status_count += count
                    
                    # Add status name with conditional formatting
                    status_cell = overall_summary[f'C{status_row}']
                    status_cell.value = status
                    style = apply_status_style(status_cell, status, config)
                    status_cell.alignment = OVERALL_SUMMARY_STYLES['data_cell']['alignment']
                    status_cell.border = OVERALL_SUMMARY_STYLES['border']
                    
                    # Add count with matching style
                    count_cell = overall_summary[f'D{status_row}']
                    count_cell.value = count
                    count_cell.font = Font(
                        name=style['font'].name,
                        size=style['font'].size,
                        bold=style['font'].bold,
                        italic=style['font'].italic,
                        color=style['font'].color
                    )
                    count_cell.fill = style['fill']
                    count_cell.alignment = OVERALL_SUMMARY_STYLES['data_cell']['alignment']
                    count_cell.border = OVERALL_SUMMARY_STYLES['border']
                    
                    status_row += 1
                
                # Add total row for status counts
                overall_summary[f'C{status_row}'] = 'Total'
                overall_summary[f'D{status_row}'] = total_status_count
                overall_summary[f'C{status_row}'].font = OVERALL_SUMMARY_STYLES['total_cell']['font']
                overall_summary[f'C{status_row}'].alignment = OVERALL_SUMMARY_STYLES['total_cell']['alignment']
                overall_summary[f'C{status_row}'].fill = OVERALL_SUMMARY_STYLES['total_cell']['fill']
                overall_summary[f'C{status_row}'].border = OVERALL_SUMMARY_STYLES['border']
                overall_summary[f'D{status_row}'].font = OVERALL_SUMMARY_STYLES['total_cell']['font']
                overall_summary[f'D{status_row}'].alignment = OVERALL_SUMMARY_STYLES['total_cell']['alignment']
                overall_summary[f'D{status_row}'].fill = OVERALL_SUMMARY_STYLES['total_cell']['fill']
                overall_summary[f'D{status_row}'].border = OVERALL_SUMMARY_STYLES['border']
                
                return max(row, status_row) + 2  # Return the next row to start from
            
            # Add P revision summary
            current_row = 7
            current_row = add_revision_summary(current_row, p_revs, 'P Revision Summary')
            
            # Add C revision summary
            current_row = add_revision_summary(current_row, c_revs, 'C Revision Summary')
            
            # Add other revision summary
            current_row = add_revision_summary(current_row, other_revs, 'Other Revision Summary')
            
            # Initialize row variable for border formatting
            row = current_row
            
            # Add file type summary section if enabled in config
            if config.get('FILE_TYPE_SETTINGS', {}).get('include_in_summary', False):
                file_type_col = config['FILE_TYPE_SETTINGS']['column_name']
                if file_type_col in latest_data.columns:
                    # Add file type summary section
                    file_type_start_row = current_row + 2
                    overall_summary[f'A{file_type_start_row}'] = config['FILE_TYPE_SETTINGS']['summary_title']
                    overall_summary[f'A{file_type_start_row}'].font = OVERALL_SUMMARY_STYLES['section_header']['font']
                    overall_summary[f'A{file_type_start_row}'].alignment = OVERALL_SUMMARY_STYLES['section_header']['alignment']
                    overall_summary[f'A{file_type_start_row}'].fill = OVERALL_SUMMARY_STYLES['section_header']['fill']
                    
                    # Add file type headers
                    overall_summary[f'A{file_type_start_row + 1}'] = 'File Type'
                    overall_summary[f'B{file_type_start_row + 1}'] = 'Count'
                    overall_summary[f'A{file_type_start_row + 1}'].font = OVERALL_SUMMARY_STYLES['column_header']['font']
                    overall_summary[f'A{file_type_start_row + 1}'].alignment = OVERALL_SUMMARY_STYLES['column_header']['alignment']
                    overall_summary[f'A{file_type_start_row + 1}'].fill = OVERALL_SUMMARY_STYLES['column_header']['fill']
                    overall_summary[f'B{file_type_start_row + 1}'].font = OVERALL_SUMMARY_STYLES['column_header']['font']
                    overall_summary[f'B{file_type_start_row + 1}'].alignment = OVERALL_SUMMARY_STYLES['column_header']['alignment']
                    overall_summary[f'B{file_type_start_row + 1}'].fill = OVERALL_SUMMARY_STYLES['column_header']['fill']
                    
                    # Add file type data
                    row = file_type_start_row + 2
                    file_type_counts = latest_data[file_type_col].value_counts()
                    total_file_types = 0
                    for file_type, count in file_type_counts.items():
                        total_file_types += count
                        overall_summary[f'A{row}'] = file_type
                        overall_summary[f'B{row}'] = count
                        overall_summary[f'A{row}'].font = OVERALL_SUMMARY_STYLES['data_cell']['font']
                        overall_summary[f'A{row}'].alignment = OVERALL_SUMMARY_STYLES['data_cell']['alignment']
                        overall_summary[f'A{row}'].fill = OVERALL_SUMMARY_STYLES['data_cell']['fill']
                        overall_summary[f'A{row}'].border = OVERALL_SUMMARY_STYLES['border']
                        overall_summary[f'B{row}'].font = OVERALL_SUMMARY_STYLES['data_cell']['font']
                        overall_summary[f'B{row}'].alignment = OVERALL_SUMMARY_STYLES['data_cell']['alignment']
                        overall_summary[f'B{row}'].fill = OVERALL_SUMMARY_STYLES['data_cell']['fill']
                        overall_summary[f'B{row}'].border = OVERALL_SUMMARY_STYLES['border']
                        row += 1
                    
                    # Add total row for file types
                    overall_summary[f'A{row}'] = 'Total'
                    overall_summary[f'B{row}'] = total_file_types
                    overall_summary[f'A{row}'].font = OVERALL_SUMMARY_STYLES['total_cell']['font']
                    overall_summary[f'A{row}'].alignment = OVERALL_SUMMARY_STYLES['total_cell']['alignment']
                    overall_summary[f'A{row}'].fill = OVERALL_SUMMARY_STYLES['total_cell']['fill']
                    overall_summary[f'A{row}'].border = OVERALL_SUMMARY_STYLES['border']
                    overall_summary[f'B{row}'].font = OVERALL_SUMMARY_STYLES['total_cell']['font']
                    overall_summary[f'B{row}'].alignment = OVERALL_SUMMARY_STYLES['total_cell']['alignment']
                    overall_summary[f'B{row}'].fill = OVERALL_SUMMARY_STYLES['total_cell']['fill']
                    overall_summary[f'B{row}'].border = OVERALL_SUMMARY_STYLES['border']
            
            # Add borders and formatting
            for row_num in range(2, row + 1):  # Start from row 2 to skip the title row
                for col in ['A', 'B', 'C', 'D']:
                    cell = overall_summary[f'{col}{row_num}']
                    cell.border = OVERALL_SUMMARY_STYLES['border']
                    cell.alignment = OVERALL_SUMMARY_STYLES['data_cell']['alignment']
            
            # Create pie charts for P and C revision statuses
            chart_start_col = 'G'  # Start charts in column G to give more space
            
            # Function to create a pie chart for revision status data
            def create_status_pie_chart(revision_type, chart_title, chart_position):
                """Create a pie chart for status distribution of a specific revision type"""
                # Get status data for this revision type
                # Note: Certificates are already filtered out at the data loading stage
                if revision_type == 'P':
                    rev_data = latest_data[latest_data['Rev'].str.startswith('P')]
                elif revision_type == 'C':
                    rev_data = latest_data[latest_data['Rev'].str.startswith('C')]
                else:
                    return None
                
                if rev_data.empty:
                    return None
                
                # Use get_grouped_status_counts to properly group raw status values
                chart_grouped_counts = get_grouped_status_counts(rev_data['Status'], config)
                
                if len(chart_grouped_counts) == 0:
                    return None
                
                if not chart_grouped_counts:
                    return None
                
                # Find a good place to put chart data (in a hidden area)
                data_start_row = 100  # Use row 100+ to keep chart data out of the way
                chart_col_offset = 0 if revision_type == 'P' else 4  # Offset C chart data more
                
                # Add chart data to hidden area
                data_row = data_start_row
                chart_labels_with_counts = []
                for category, count in chart_grouped_counts.items():
                    label_with_count = f"{category} ({count})"
                    chart_labels_with_counts.append(label_with_count)
                    overall_summary[f'{chr(ord(chart_start_col) + chart_col_offset)}{data_row}'] = label_with_count
                    overall_summary[f'{chr(ord(chart_start_col) + chart_col_offset + 1)}{data_row}'] = count
                    data_row += 1
                
                # Create pie chart
                chart = PieChart()
                chart.title = chart_title
                chart.width = 16  # Make charts wider
                chart.height = 12  # Make charts taller
                
                # Define data range from hidden area
                labels = Reference(overall_summary, 
                                 min_col=ord(chart_start_col) - ord('A') + 1 + chart_col_offset, 
                                 min_row=data_start_row, 
                                 max_row=data_row - 1)
                data = Reference(overall_summary, 
                               min_col=ord(chart_start_col) - ord('A') + 2 + chart_col_offset, 
                               min_row=data_start_row, 
                               max_row=data_row - 1)
                
                chart.add_data(data, titles_from_data=False)
                chart.set_categories(labels)
                
                # Configure data labels to show only percentages (not "Series 1")
                chart.dataLabels = DataLabelList()
                chart.dataLabels.showCatName = False  # Don't show category name on slices
                chart.dataLabels.showVal = False      # Don't show values on slices
                chart.dataLabels.showPercent = True
                chart.dataLabels.showSerName = False  # This should remove "Series 1"
                
                # Set legend to show our custom labels with counts
                chart.legend.position = 'r'  # Position legend to the right
                chart.legend.layout = None  # Let Excel auto-position the legend
                
                # Style the chart colors based on config or fallback to STATUS_STYLES
                # Build colors dictionary from config with chart-optimized colors
                colors = {}
                if config and 'STATUS_MAPPINGS' in config:
                    # Use config-based colors but ensure they're visible in charts
                    for category, mapping in config['STATUS_MAPPINGS'].items():
                        config_color = mapping.get('color', 'FFFFFF')
                        # Convert white/very light colors to chart-visible colors
                        colors[category] = get_chart_safe_color(config_color, category)
                else:
                    # Fallback to hardcoded colors
                    colors = {
                        'Status A': '00B050',  # Green
                        'Status B': 'EDDDA1',  # Beige
                        'Status C': 'ED1111',  # Red
                        'Other': '808080'      # Grey
                    }
                
                # Apply colors to chart slices
                try:
                    if chart.series:
                        series = chart.series[0]
                        for i, (category, count) in enumerate(chart_grouped_counts.items()):
                            # Extract the category name without the count
                            category_name = category.split(' (')[0]
                            if category_name in colors:
                                # Create a data point with the specified color
                                pt = DataPoint(idx=i)
                                # Set the fill color
                                pt.graphicalProperties.solidFill = colors[category_name]
                                series.dPt.append(pt)
                except Exception as e:
                    # If coloring fails, continue without custom colors
                    print(f"Warning: Could not apply chart colors: {e}")
                
                # Alternative approach using series formatting
                try:
                    if chart.series and len(chart.series) > 0:
                        series = chart.series[0]
                        # Clear any existing data points
                        series.dPt = []
                        
                        for i, (category, count) in enumerate(chart_grouped_counts.items()):
                            category_name = category.split(' (')[0]
                            if category_name in colors:
                                # Create data point
                                pt = DataPoint(idx=i)
                                # Create fill properties
                                fill = PatternFillProperties()
                                fill.solidFill = ColorChoice(srgbClr=colors[category_name])
                                pt.graphicalProperties.solidFill = fill.solidFill
                                series.dPt.append(pt)
                except Exception as e:
                    # If this approach fails too, try a simpler method
                    try:
                        for i, (category, count) in enumerate(chart_grouped_counts.items()):
                            category_name = category.split(' (')[0]
                            if category_name in colors and chart.series:
                                pt = DataPoint(idx=i)
                                # Simple color assignment
                                pt.graphicalProperties.solidFill = colors[category_name]
                                if not hasattr(chart.series[0], 'dPt') or chart.series[0].dPt is None:
                                    chart.series[0].dPt = []
                                chart.series[0].dPt.append(pt)
                    except Exception as e2:
                        print(f"Warning: All chart coloring methods failed: {e2}")
                        pass
                
                # Position the chart
                overall_summary.add_chart(chart, chart_position)
                
                return data_row
            
            # Create P revision status chart - position it lower to give more space from title
            create_status_pie_chart('P', 'P Revision Status Distribution', 'G2')
            
            # Create C revision status chart - position it below P chart with spacing
            create_status_pie_chart('C', 'C Revision Status Distribution', 'G27')
            
            # Auto-adjust column widths for Overall Summary
            for column in overall_summary.columns:
                max_length = 0
                column_letter = None
                for cell in column:
                    if hasattr(cell, 'column_letter'):
                        column_letter = cell.column_letter
                        break
                
                if not column_letter:
                    continue  # Skip if we can't get column letter
                
                # Skip the first row for merged columns in Overall Summary
                start_row = 2 if column_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O'] else 1
                for cell in column[start_row-1:]:
                    try:
                        # Skip merged cells
                        if hasattr(cell, 'value') and cell.value is not None:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                    except:
                        pass
                # Set minimum width of 8 for empty/narrow columns, otherwise use calculated width
                adjusted_width = max(8, max_length + 2)
                overall_summary.column_dimensions[column_letter].width = adjusted_width
            
            # Auto-adjust column widths for all other sheets
            for sheet_name in ['Summary Data', 'Changes', 'Latest Data']:
                if sheet_name in book.sheetnames:
                    sheet = book[sheet_name]
                    for column in sheet.columns:
                        max_length = 0
                        # Get column letter from first non-merged cell
                        column_letter = None
                        for cell in column:
                            if hasattr(cell, 'column_letter'):
                                column_letter = cell.column_letter
                                break
                        
                        if not column_letter:
                            continue  # Skip if we can't get column letter
                        
                        for cell in column:
                            try:
                                # Skip merged cells
                                if hasattr(cell, 'value') and cell.value is not None:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = (max_length + 2)
                        sheet.column_dimensions[column_letter].width = adjusted_width
            
            # FINAL ATTEMPT: Set alignment as the very last operation
            overall_summary['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Final save with the Overall Summary sheet
            book.save(output_file)
            return True
            
        except PermissionError:
            if attempt < max_retries - 1:
                print(f"File {output_file} is in use. Waiting before retry...")
                time.sleep(2)
            else:
                print(f"Could not save to {output_file} - file is in use.")
                return False
    return False

