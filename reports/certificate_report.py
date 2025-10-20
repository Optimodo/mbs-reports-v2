"""Certificate report generation module."""

import warnings
import pandas as pd
from datetime import datetime
from pathlib import Path
import time
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.chart import PieChart, Reference, BarChart
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.fill import ColorChoice, PatternFillProperties, SolidColorFillProperties

from styles.formatting import (
    OVERALL_SUMMARY_STYLES,
    apply_status_style
)
from utils.status_mapping import (
    get_status_category,
    get_status_color,
    get_grouped_status_counts,
    get_status_display_order
)
from analyzers.document_tracker import (
    categorize_documents,
    get_apartment_certificate_summary,
    get_uncategorized_certificates_in_blocks
)

# Suppress openpyxl warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')


def get_chart_safe_color(config_color, category):
    """Convert config colors to chart-safe colors.
    
    Ensures that white/very light colors are converted to visible colors for charts.
    
    Args:
        config_color: Color from config (hex string)
        category: Status category name for fallback mapping
        
    Returns:
        str: Chart-safe hex color
    """
    # Define chart-safe colors for common white/light statuses
    chart_safe_mapping = {
        'Other': 'C0C0C0',        # Light gray
        'Information': 'D3D3D3',   # Light gray
        'Review': 'E6E6FA',       # Lavender
        'IFC-pending': 'F0E68C',  # Khaki
        'Under Review': 'DDA0DD', # Plum
        'Shared': 'E0F090',       # Light yellow-green
        'Published': '90EE90'     # Light green
    }
    
    # If config color is white or very light, use chart-safe color
    if config_color and config_color.upper() in ['FFFFFF', 'FFF', 'FFFFFFFF']:
        return chart_safe_mapping.get(category, 'C0C0C0')
    
    # Check if color is very light (all RGB components > 240)
    if config_color and len(config_color) >= 6:
        try:
            r = int(config_color[0:2], 16)
            g = int(config_color[2:4], 16)
            b = int(config_color[4:6], 16)
            if r > 240 and g > 240 and b > 240:
                return chart_safe_mapping.get(category, 'C0C0C0')
        except ValueError:
            pass
    
    return config_color


def add_apartment_certificate_tracking(ws, latest_data, config, start_row=5, max_blocks_per_phase=7):
    """
    Add apartment certificate tracking section with progress bars.
    
    Args:
        ws: Worksheet to add tracking to
        latest_data: DataFrame with latest certificate data
        config: Project configuration
        start_row: Starting row for the tracking section
        
    Returns:
        int: Next available row after this section
    """
    # Check if certificate tracking is configured
    cert_tracking = config.get('CERTIFICATE_TRACKING', {})
    if not cert_tracking:
        return start_row
    
    apartment_certs = cert_tracking.get('apartment_certificates', {})
    if not apartment_certs:
        return start_row
    
    # Check if accommodation data is available
    accom_data = config.get('ACCOMMODATION_DATA', {})
    has_accom_data = bool(accom_data and accom_data.get('apartment_lookup'))
    
    # Categorize certificates
    categorized = categorize_documents(latest_data, apartment_certs, cert_tracking)
    
    # Get accommodation data for accurate counts
    accom_data = config.get('ACCOMMODATION_DATA', {})
    
    # Get summary statistics (will use accommodation data if available)
    summary = get_apartment_certificate_summary(categorized, apartment_certs, cert_tracking, accom_data)
    
    # Calculate end column based on max blocks (E + max_blocks_per_phase + 2 for spacing)
    end_col = 5 + max_blocks_per_phase + 2  # Start at E (5), add blocks, add 2 for spacing
    end_col_letter = ws.cell(row=1, column=end_col).column_letter
    
    # Section header - dynamic based on max blocks
    ws[f'A{start_row}'] = 'APARTMENT CERTIFICATE TRACKING'
    ws[f'A{start_row}'].font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    ws[f'A{start_row}'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws[f'A{start_row}'].alignment = Alignment(horizontal='left', vertical='center')
    # Merge header across all columns that will be used
    ws.merge_cells(f'A{start_row}:{end_col_letter}{start_row}')
    
    start_row += 2
    
    # Overall progress bar
    overall = summary['overall_progress']
    total_possible = overall['total_max_apartments'] * len(apartment_certs)
    
    ws[f'A{start_row}'] = 'Overall Certificate Progress'
    ws[f'A{start_row}'].font = Font(name='Calibri', size=12, bold=True)
    
    ws[f'B{start_row}'] = f"{overall['total_apartments_with_docs']}/{total_possible}"
    ws[f'C{start_row}'] = f"{overall['overall_progress_percentage']}%"
    ws[f'D{start_row}'] = overall['total_documents']
    ws[f'D{start_row}'].alignment = Alignment(horizontal='center')
    
    # Add visual progress bar in merged cells - dynamic based on max blocks
    progress_pct = overall['overall_progress_percentage']
    # Calculate end column based on max blocks (E + max_blocks_per_phase + 2 for spacing)
    end_col = 5 + max_blocks_per_phase + 2  # Start at E (5), add blocks, add 2 for spacing
    end_col_letter = ws.cell(row=1, column=end_col).column_letter
    
    ws.merge_cells(f'E{start_row}:{end_col_letter}{start_row}')
    
    # Calculate progress bar length based on available space (more blocks = longer bar)
    bar_length = min(60, max(30, max_blocks_per_phase * 10))  # Much longer bars: scale with blocks, min 30, max 60
    filled_blocks = int(progress_pct / 100 * bar_length)
    empty_blocks = bar_length - filled_blocks
    
    ws[f'E{start_row}'] = f"{'█' * filled_blocks}{'░' * empty_blocks} {progress_pct}%"
    ws[f'E{start_row}'].font = Font(name='Courier New', size=10)
    ws[f'E{start_row}'].alignment = Alignment(horizontal='left')
    
    start_row += 2
    
    # Column headers - dynamic based on max blocks
    headers = ['Certificate Type', 'Apartments', 'Progress %', 'Documents', 'Progress Bar']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx)
        cell.value = header
        cell.font = Font(name='Calibri', size=11, bold=True)
        cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Merge the progress bar header across the remaining columns
    if max_blocks_per_phase > 0:
        progress_bar_start_col = 5  # Column E
        progress_bar_end_col_letter = ws.cell(row=1, column=end_col).column_letter
        ws.merge_cells(f'{ws.cell(row=1, column=progress_bar_start_col).column_letter}{start_row}:{progress_bar_end_col_letter}{start_row}')
        # Update the merged cell content
        progress_bar_cell = ws.cell(row=start_row, column=progress_bar_start_col)
        progress_bar_cell.value = 'Progress Bar'
        progress_bar_cell.font = Font(name='Calibri', size=11, bold=True)
        progress_bar_cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        progress_bar_cell.alignment = Alignment(horizontal='center', vertical='center')
        progress_bar_cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    start_row += 1
    
    # Add progress for each certificate type
    progress_stats = summary['progress_stats']
    
    for category_key, stats in progress_stats.items():
        display_name = apartment_certs[category_key].get('display_name', category_key)
        apartments_with_docs = stats['apartments_with_docs']
        max_apartments = stats['max_apartments']
        progress_pct = stats['progress_percentage']
        doc_count = stats['documents_count']
        
        # Certificate type name
        ws[f'A{start_row}'] = display_name
        ws[f'A{start_row}'].font = Font(name='Calibri', size=10)
        ws[f'A{start_row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        # Apartments completed
        ws[f'B{start_row}'] = f"{apartments_with_docs}/{max_apartments}"
        ws[f'B{start_row}'].font = Font(name='Calibri', size=10)
        ws[f'B{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Progress percentage
        ws[f'C{start_row}'] = f"{progress_pct}%"
        ws[f'C{start_row}'].font = Font(name='Calibri', size=10, bold=True)
        ws[f'C{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Color code based on progress
        if progress_pct >= 80:
            fill_color = '25E82C'  # Green
        elif progress_pct >= 50:
            fill_color = 'EDDDA1'  # Yellow
        elif progress_pct >= 25:
            fill_color = 'FFA500'  # Orange
        else:
            fill_color = 'ED1111'  # Red
        
        ws[f'C{start_row}'].fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        
        # Document count
        ws[f'D{start_row}'] = doc_count
        ws[f'D{start_row}'].font = Font(name='Calibri', size=10)
        ws[f'D{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Progress bar (horizontal bar using block characters) - dynamic width
        # Scale bar length with available space (more blocks = longer bar)
        bar_length = min(60, max(30, max_blocks_per_phase * 10))  # Much longer bars: scale with blocks, min 30, max 60
        filled_blocks = int(progress_pct / 100 * bar_length)
        empty_blocks = bar_length - filled_blocks
        
        # Merge progress bar across all available columns
        progress_bar_end_col_letter = ws.cell(row=1, column=end_col).column_letter
        ws.merge_cells(f'E{start_row}:{progress_bar_end_col_letter}{start_row}')
        ws[f'E{start_row}'] = f"{'█' * filled_blocks}{'░' * empty_blocks} {progress_pct}%"
        ws[f'E{start_row}'].font = Font(name='Courier New', size=10)
        ws[f'E{start_row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        # Add borders
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{start_row}'].border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        start_row += 1
    
    start_row += 1
    
    # Add phase/block breakdown if available
    phase_block_progress = summary.get('phase_block_progress', {})
    if phase_block_progress:
        ws[f'A{start_row}'] = 'PROGRESS BY PHASE & BLOCK'
        ws[f'A{start_row}'].font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        ws[f'A{start_row}'].fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
        # Merge header across all columns that will be used
        end_col_letter = ws.cell(row=1, column=end_col).column_letter
        ws.merge_cells(f'A{start_row}:{end_col_letter}{start_row}')
        start_row += 2
        
        for phase_id, phase_data in phase_block_progress.items():
            phase_display = phase_data['display_name']
            
            # Phase header - dynamic width
            ws[f'A{start_row}'] = phase_display
            ws[f'A{start_row}'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            ws[f'A{start_row}'].fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
            ws.merge_cells(f'A{start_row}:{end_col_letter}{start_row}')
            start_row += 1
            
            # Phase certificate breakdown - show ALL certificate types (same order as overall section)
            phase_stats = phase_data['phase_stats']
            block_stats = phase_data['block_stats']
            
            # Get sorted block list for this phase
            sorted_blocks = sorted(block_stats.keys()) if block_stats else []
            
            # Iterate through categories in the same order as the overall section
            for cert_key in progress_stats.keys():
                if cert_key not in phase_stats:
                    continue
                
                stats = phase_stats[cert_key]
                display_name = apartment_certs[cert_key].get('display_name', cert_key)
                apts = stats['apartments_with_docs']
                max_apts = stats['max_apartments']
                pct = stats['progress_percentage']
                
                # Certificate name in column A
                ws[f'A{start_row}'] = display_name
                ws[f'A{start_row}'].font = Font(name='Calibri', size=10)
                
                # Progress info in column B
                ws[f'B{start_row}'] = f"{apts}/{max_apts} ({pct}%)"
                ws[f'B{start_row}'].font = Font(name='Calibri', size=10)
                ws[f'B{start_row}'].alignment = Alignment(horizontal='center')
                
                # Mini progress bar in column C
                filled = int(pct / 10)  # 10 blocks = 100%
                ws[f'C{start_row}'] = f"{'█' * filled}{'░' * (10 - filled)}"
                ws[f'C{start_row}'].font = Font(name='Courier New', size=9)
                
                # Dynamic block breakdown starting from column D
                if sorted_blocks:
                    block_col = 4  # Column D (4th column)
                    for block_id in sorted_blocks:
                        block_data = block_stats[block_id]
                        if cert_key in block_data:
                            block_cert_data = block_data[cert_key]
                            block_apts = block_cert_data['apartments_with_docs']
                            
                            # Use available columns (D onwards, up to reasonable limit)
                            if block_col <= 10:  # Columns D through J
                                cell = ws.cell(row=start_row, column=block_col)
                                cell.value = f"{block_id}:{block_apts}"
                                cell.font = Font(name='Calibri', size=9)
                                cell.alignment = Alignment(horizontal='center')
                                block_col += 1
                
                start_row += 1
            
            start_row += 1
    
    return start_row


def add_data_quality_section(ws, latest_data, config, start_row=5):
    """
    Add data quality section showing rejected and uncategorized certificates.
    
    Args:
        ws: Worksheet to add section to
        latest_data: DataFrame with latest certificate data
        config: Project configuration
        start_row: Starting row for the section
        
    Returns:
        int: Next available row after this section
    """
    # Section header
    ws[f'A{start_row}'] = 'DATA QUALITY REPORT'
    ws[f'A{start_row}'].font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    ws[f'A{start_row}'].fill = PatternFill(start_color='FF6600', end_color='FF6600', fill_type='solid')
    ws[f'A{start_row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(f'A{start_row}:D{start_row}')
    start_row += 2
    
    # Get certificate tracking config
    cert_tracking = config.get('CERTIFICATE_TRACKING', {})
    apartment_certs = cert_tracking.get('apartment_certificates', {})
    
    # === REJECTED CERTIFICATES SUMMARY ===
    ws[f'A{start_row}'] = '1. REJECTED CERTIFICATES (Status C)'
    ws[f'A{start_row}'].font = Font(name='Calibri', size=11, bold=True, color='ED1111')
    start_row += 1
    
    # Get rejected certificates (Status C)
    status_mappings = config.get('STATUS_MAPPINGS', {})
    status_c_values = []
    
    if 'Status C' in status_mappings:
        status_c_values = status_mappings['Status C'].get('statuses', [])
    
    # Filter for rejected certificates using config
    if status_c_values and 'Status' in latest_data.columns:
        rejected = latest_data[latest_data['Status'].isin(status_c_values)]
    else:
        # Fallback: look for common rejected status terms
        rejected = latest_data[latest_data['Status'].str.contains('reject|Reject|REJECT|C-', case=False, na=False)]
    
    if rejected.empty:
        ws[f'A{start_row}'] = '  ✓ No rejected certificates'
        ws[f'A{start_row}'].font = Font(name='Calibri', size=10, italic=True, color='25E82C')
        start_row += 2
    else:
        ws[f'A{start_row}'] = f'  Total Rejected:'
        ws[f'B{start_row}'] = len(rejected)
        ws[f'A{start_row}'].font = Font(name='Calibri', size=10, bold=True)
        ws[f'B{start_row}'].font = Font(name='Calibri', size=10, bold=True, color='ED1111')
        start_row += 1
        
        if apartment_certs:
            # Categorize rejected certificates
            categorized_rejected = categorize_documents(rejected, apartment_certs, cert_tracking)
            
            # Show rejected count by certificate type
            for cert_key, cert_config in apartment_certs.items():
                display_name = cert_config.get('display_name', cert_key)
                rejected_for_type = categorized_rejected[categorized_rejected['category'] == cert_key]
                
                if len(rejected_for_type) > 0:
                    ws[f'A{start_row}'] = f'    {display_name}:'
                    ws[f'B{start_row}'] = len(rejected_for_type)
                    ws[f'A{start_row}'].font = Font(name='Calibri', size=9)
                    ws[f'B{start_row}'].font = Font(name='Calibri', size=9, color='ED1111')
                    ws[f'B{start_row}'].alignment = Alignment(horizontal='left')
                    start_row += 1
        
        start_row += 1
    
    # === UNCATEGORIZED CERTIFICATES (Missing Plot Numbers) ===
    ws[f'A{start_row}'] = '2. UNCATEGORIZED CERTIFICATES (Potential Naming Issues)'
    ws[f'A{start_row}'].font = Font(name='Calibri', size=11, bold=True, color='C00000')
    start_row += 1
    
    if apartment_certs:
        categorized = categorize_documents(latest_data, apartment_certs, cert_tracking)
        uncategorized = get_uncategorized_certificates_in_blocks(latest_data, categorized)
        
        if uncategorized.empty:
            ws[f'A{start_row}'] = '  ✓ All certificates in block folders are properly categorized'
            ws[f'A{start_row}'].font = Font(name='Calibri', size=10, italic=True, color='25E82C')
            start_row += 2
        else:
            ws[f'A{start_row}'] = '  Certificates in block folders but missing valid plot numbers (or certificate title/description):'
            ws[f'A{start_row}'].font = Font(name='Calibri', size=10, italic=True)
            start_row += 1
            
            # Count by block
            block_counts = uncategorized['extracted_block'].value_counts().sort_index()
            ws[f'A{start_row}'] = f'  Total Uncategorized:'
            ws[f'B{start_row}'] = len(uncategorized)
            ws[f'A{start_row}'].font = Font(name='Calibri', size=10, bold=True)
            ws[f'B{start_row}'].font = Font(name='Calibri', size=10, bold=True, color='C00000')
            start_row += 1
            
            for block, count in block_counts.items():
                ws[f'A{start_row}'] = f'    Block {block}:'
                ws[f'B{start_row}'] = count
                ws[f'A{start_row}'].font = Font(name='Calibri', size=9)
                ws[f'B{start_row}'].font = Font(name='Calibri', size=9, color='C00000')
                start_row += 1
            
            start_row += 1
            
            # Note about detailed uncategorized report
            ws[f'A{start_row}'] = '  Detailed uncategorized certificates moved to separate tab for analysis'
            ws[f'A{start_row}'].font = Font(name='Calibri', size=9, italic=True, color='666666')
            ws.merge_cells(f'A{start_row}:D{start_row}')
            start_row += 1
    else:
        ws[f'A{start_row}'] = '  Certificate tracking not configured for this project'
        ws[f'A{start_row}'].font = Font(name='Calibri', size=10, italic=True)
        start_row += 1
    
    start_row += 2
    return start_row


def add_uncategorized_detailed_tab(wb, latest_data, config):
    """Add a detailed tab for uncategorized certificates analysis"""
    # Get certificate tracking config
    cert_tracking = config.get('CERTIFICATE_TRACKING', {})
    apartment_certs = cert_tracking.get('apartment_certificates', {})
    
    if not apartment_certs:
        return  # No apartment tracking configured
    
    # Create detailed uncategorized tab
    ws = wb.create_sheet("Uncategorized Analysis")
    
    # Title
    ws['A1'] = 'UNCATEGORIZED CERTIFICATES - DETAILED ANALYSIS'
    ws['A1'].font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:H1')
    
    # Filter for certificates
    from utils.document_filters import filter_certificates
    cert_data = filter_certificates(latest_data, config)
    
    # Categorize documents
    from analyzers.document_tracker import categorize_documents, get_uncategorized_certificates_in_blocks
    categorized = categorize_documents(cert_data, apartment_certs, cert_tracking)
    uncategorized = get_uncategorized_certificates_in_blocks(cert_data, categorized)
    
    if uncategorized.empty:
        ws['A3'] = '✓ All certificates in block folders are properly categorized'
        ws['A3'].font = Font(name='Calibri', size=12, italic=True, color='25E82C')
        return
    
    # Headers
    headers = ['Block', 'Doc Ref', 'Doc Title', 'Doc Path', 'Publisher', 'File Type', 'Status', 'Rev']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    row = 6
    for idx, cert in uncategorized.iterrows():
        ws.cell(row=row, column=1, value=cert.get('extracted_block', '?'))
        ws.cell(row=row, column=2, value=cert.get('Doc Ref', ''))
        ws.cell(row=row, column=3, value=cert.get('Doc Title', ''))
        
        # Truncate doc path to show only relevant part
        doc_path = cert.get('Doc Path', '')
        if doc_path:
            # Extract the relevant part starting from "16. Testing & Commissioning" or similar
            # Example: "...\16. Testing & Commissioning\18.03\Block - F\Electrical Certs"
            path_parts = doc_path.split('\\')
            if len(path_parts) > 1:
                # Find the "16. Testing & Commissioning" section
                relevant_start = -1
                for i, part in enumerate(path_parts):
                    if part and ('Testing & Commissioning' in part or '16.' in part):
                        relevant_start = i
                        break
                
                if relevant_start >= 0:
                    # Start from the testing section, but add dots before it
                    truncated_path = '...\\' + '\\'.join(path_parts[relevant_start:])
                else:
                    # Fallback: show last 4 parts if testing section not found
                    truncated_path = '...\\' + '\\'.join(path_parts[-4:])
            else:
                truncated_path = doc_path
        else:
            truncated_path = ''
        
        ws.cell(row=row, column=4, value=truncated_path)
        ws.cell(row=row, column=5, value=cert.get('Publisher', ''))
        ws.cell(row=row, column=6, value=cert.get('File Type', ''))
        ws.cell(row=row, column=7, value=cert.get('Status', ''))
        ws.cell(row=row, column=8, value=cert.get('Rev', ''))
        
        # Alternate row colors for readability
        if row % 2 == 0:
            for col in range(1, 9):
                ws.cell(row=row, column=col).fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        row += 1
    
    # Summary - merge across all columns and center-align
    ws['A3'] = f'Total Uncategorized Certificates: {len(uncategorized)}'
    ws['A3'].font = Font(name='Calibri', size=12, bold=True, color='C00000')
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A3:H3')  # Merge across all columns (A through H)
    
    # Auto-adjust column widths AFTER merging row 3
    for column in ws.columns:
        max_length = 0
        column_letter = None
        for cell in column:
            try:
                # Skip merged cells - they don't have column_letter attribute
                if hasattr(cell, 'column_letter'):
                    if column_letter is None:
                        column_letter = cell.column_letter
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except:
                pass
        
        if column_letter is not None:
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 to prevent extremely wide columns
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Manually set column A (Block) width since it only contains single letters (A-G)
    ws.column_dimensions['A'].width = 8
    
    # Center-align columns A (Block) and H (Rev) for better presentation
    for row_num in range(5, row + 1):  # From headers to last data row
        if ws[f'A{row_num}'].value is not None:
            ws[f'A{row_num}'].alignment = Alignment(horizontal='center', vertical='center')
        if ws[f'H{row_num}'].value is not None:
            ws[f'H{row_num}'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Freeze panes
    ws.freeze_panes = 'A6'
    
    # Analysis section removed to prevent column A from being too wide


def save_certificate_report(summary_df, latest_data, output_file, config):
    """
    Save a comprehensive certificate report to Excel.
    
    Args:
        summary_df: DataFrame with summary data (revision/status counts over time)
        latest_data: DataFrame with the latest certificate data
        output_file: Path to the output Excel file
        config: Project configuration dictionary
        
    Returns:
        bool: True if successful, False otherwise
    """
    try:
        # Create Excel writer
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Write Summary Data sheet
            summary_df.to_excel(writer, sheet_name='Summary Data', index=False)
            
            # Write Latest Certificate Data sheet
            latest_data.to_excel(writer, sheet_name='Latest Certificate Data', index=False)
        
        # Load the workbook to add the Overall Summary sheet
        wb = load_workbook(output_file)
        
        # Create Overall Summary sheet at the beginning
        if 'Overall Summary' in wb.sheetnames:
            del wb['Overall Summary']
        overall_summary = wb.create_sheet('Overall Summary', 0)
        
        # Set up page layout for A4 landscape
        overall_summary.page_setup.orientation = 'landscape'
        overall_summary.page_setup.paperSize = 9  # A4
        overall_summary.page_setup.fitToWidth = 1
        overall_summary.page_setup.fitToHeight = 0
        overall_summary.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)
        
        # Add title and project info
        project_title = config.get('PROJECT_TITLE', 'Project')
        cert_settings = config.get('CERTIFICATE_SETTINGS', {})
        
        overall_summary['A1'] = f"{project_title} - Certificate Report"
        overall_summary['A1'].font = Font(name='Calibri', size=16, bold=True)
        overall_summary['A1'].alignment = Alignment(horizontal='left', vertical='center')
        
        # Get latest row from summary
        latest_row = summary_df.iloc[-1] if not summary_df.empty else {}
        latest_date = latest_row.get('Date', 'N/A')
        latest_time = latest_row.get('Time', 'N/A')
        
        overall_summary['A2'] = f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        overall_summary['A2'].font = Font(name='Calibri', size=10, italic=True)
        
        overall_summary['A3'] = f"Latest Data: {latest_date} {latest_time}"
        overall_summary['A3'].font = Font(name='Calibri', size=10, italic=True)
        
        # Add apartment certificate tracking section if configured
        cert_tracking = config.get('CERTIFICATE_TRACKING', {})
        apartment_certs = cert_tracking.get('apartment_certificates', {})
        
        if apartment_certs:
            # Calculate max blocks per phase for dynamic column sizing
            accom_data = config.get('ACCOMMODATION_DATA', {})
            max_blocks_per_phase = 0
            if 'phases' in accom_data:
                for phase_config in accom_data['phases'].values():
                    phase_blocks = len(phase_config.get('blocks', []))
                    max_blocks_per_phase = max(max_blocks_per_phase, phase_blocks)
            
            # Add apartment certificate tracking with progress bars
            next_row = add_apartment_certificate_tracking(overall_summary, latest_data, config, start_row=5, max_blocks_per_phase=max_blocks_per_phase)
            
            # Add separator
            overall_summary[f'A{next_row}'] = ''
            next_row += 1
            
            # Add "Landlord/Communal Certificates" section header
            overall_summary[f'A{next_row}'] = 'LANDLORD/COMMUNAL CERTIFICATES'
            overall_summary[f'A{next_row}'].font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
            overall_summary[f'A{next_row}'].fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
            overall_summary[f'A{next_row}'].alignment = Alignment(horizontal='left', vertical='center')
            overall_summary.merge_cells(f'A{next_row}:D{next_row}')
            next_row += 2
            
            # Count landlord/communal certificates across all blocks
            landlord_certs_mask = latest_data['Doc Path'].fillna('').astype(str).str.contains(r'\\Landlords\\', case=False, na=False, regex=True)
            landlord_certs = latest_data[landlord_certs_mask]
            
            # Show count of landlord/communal certificates
            overall_summary[f'A{next_row}'] = 'Landlord/Communal Certificates:'
            overall_summary[f'B{next_row}'] = len(landlord_certs)
            overall_summary[f'A{next_row}'].font = OVERALL_SUMMARY_STYLES['total_cell']['font']
            overall_summary[f'B{next_row}'].font = OVERALL_SUMMARY_STYLES['total_cell']['font']
            overall_summary[f'B{next_row}'].alignment = Alignment(horizontal='center')
            next_row += 2
            
            # Add rejected certificates section
            current_row = add_data_quality_section(overall_summary, latest_data, config, next_row)
        else:
            # No apartment tracking - simplified format
            # Add total certificate count
            total_certs = len(latest_data)
            overall_summary['A5'] = 'Total Certificates:'
            overall_summary['B5'] = total_certs
            overall_summary['A5'].font = OVERALL_SUMMARY_STYLES['total_cell']['font']
            overall_summary['A5'].alignment = OVERALL_SUMMARY_STYLES['total_cell']['alignment']
            overall_summary['A5'].fill = OVERALL_SUMMARY_STYLES['total_cell']['fill']
            overall_summary['B5'].font = OVERALL_SUMMARY_STYLES['total_cell']['font']
            overall_summary['B5'].alignment = OVERALL_SUMMARY_STYLES['total_cell']['alignment']
            overall_summary['B5'].fill = OVERALL_SUMMARY_STYLES['total_cell']['fill']
            
            # Add rejected certificates section
            current_row = add_data_quality_section(overall_summary, latest_data, config, 7)
        
        # Revision/status summaries removed - apartment tracking and rejected certificates are more valuable
        # The detailed revision/status data is still available in the "Summary Data" sheet if needed
        
        # Function to add revision and status summary
        def add_certificate_revision_summary(start_row, rev_columns, title):
            """Add a certificate revision summary section with status breakdown"""
            if not rev_columns:
                return start_row
            
            # Add section header
            overall_summary[f'A{start_row}'] = title
            overall_summary[f'A{start_row}'].font = OVERALL_SUMMARY_STYLES['section_header']['font']
            overall_summary[f'A{start_row}'].alignment = OVERALL_SUMMARY_STYLES['section_header']['alignment']
            overall_summary[f'A{start_row}'].fill = OVERALL_SUMMARY_STYLES['section_header']['fill']
            
            # Add headers
            overall_summary[f'A{start_row + 1}'] = 'Revision'
            overall_summary[f'B{start_row + 1}'] = 'Count'
            overall_summary[f'C{start_row + 1}'] = 'Status'
            overall_summary[f'D{start_row + 1}'] = 'Count'
            
            # Style headers
            for col in ['A', 'B', 'C', 'D']:
                overall_summary[f'{col}{start_row + 1}'].font = OVERALL_SUMMARY_STYLES['column_header']['font']
                overall_summary[f'{col}{start_row + 1}'].alignment = OVERALL_SUMMARY_STYLES['column_header']['alignment']
                overall_summary[f'{col}{start_row + 1}'].fill = OVERALL_SUMMARY_STYLES['column_header']['fill']
                overall_summary[f'{col}{start_row + 1}'].border = OVERALL_SUMMARY_STYLES['border']
            
            # Get status counts for all revisions in this group
            status_counts = {}
            total_count = 0
            
            for rev_col in rev_columns:
                rev_name = rev_col.replace('Rev_', '')
                
                # Get count from summary data
                count = latest_row.get(rev_col, 0)
                # Handle NaN values from pandas
                total_count += 0 if pd.isna(count) else count
                
                # Count statuses for this revision
                rev_data = latest_data[latest_data['Rev'] == rev_name]
                
                # Use get_grouped_status_counts to properly group raw status values
                if not rev_data.empty:
                    grouped_counts = get_grouped_status_counts(rev_data['Status'], config)
                    for status, status_count in grouped_counts.items():
                        status_counts[status] = status_counts.get(status, 0) + status_count
            
            # Add revision data
            row = start_row + 2
            for rev_col in rev_columns:
                rev_name = rev_col.replace('Rev_', '')
                
                # Get the count from the summary data
                count = latest_row.get(rev_col, 0)
                # Handle NaN values from pandas
                count = 0 if pd.isna(count) else count
                
                # Only add revision row if there are actually documents with this revision
                if count > 0:
                    # Add revision name and count
                    overall_summary[f'A{row}'] = rev_name
                    overall_summary[f'B{row}'] = count
                    overall_summary[f'A{row}'].font = OVERALL_SUMMARY_STYLES['data_cell']['font']
                    overall_summary[f'A{row}'].alignment = OVERALL_SUMMARY_STYLES['data_cell']['alignment']
                    overall_summary[f'A{row}'].fill = OVERALL_SUMMARY_STYLES['data_cell']['fill']
                    overall_summary[f'A{row}'].border = OVERALL_SUMMARY_STYLES['border']
                    overall_summary[f'B{row}'].font = OVERALL_SUMMARY_STYLES['data_cell']['font']
                    overall_summary[f'B{row}'].alignment = OVERALL_SUMMARY_STYLES['data_cell']['alignment']
                    overall_summary[f'B{row}'].fill = OVERALL_SUMMARY_STYLES['data_cell']['fill']
                    overall_summary[f'B{row}'].border = OVERALL_SUMMARY_STYLES['border']
                    row += 1
            
            # Add total row for revisions
            overall_summary[f'A{row}'] = 'Total'
            overall_summary[f'B{row}'] = total_count
            overall_summary[f'A{row}'].font = OVERALL_SUMMARY_STYLES['total_cell']['font']
            overall_summary[f'A{row}'].alignment = OVERALL_SUMMARY_STYLES['total_cell']['alignment']
            overall_summary[f'A{row}'].fill = OVERALL_SUMMARY_STYLES['total_cell']['fill']
            overall_summary[f'A{row}'].border = OVERALL_SUMMARY_STYLES['border']
            overall_summary[f'B{row}'].font = OVERALL_SUMMARY_STYLES['total_cell']['font']
            overall_summary[f'B{row}'].alignment = OVERALL_SUMMARY_STYLES['total_cell']['alignment']
            overall_summary[f'B{row}'].fill = OVERALL_SUMMARY_STYLES['total_cell']['fill']
            overall_summary[f'B{row}'].border = OVERALL_SUMMARY_STYLES['border']
            row += 1
            
            # Add status summary
            status_row = start_row + 2
            total_status_count = 0
            
            # Create ordered list using project-specific categories
            ordered_statuses = []
            
            # Build ordered list using project-specific display order
            if config and 'STATUS_DISPLAY_ORDER' in config:
                display_order = config['STATUS_DISPLAY_ORDER']
                
                # Add statuses in the order defined by STATUS_DISPLAY_ORDER
                for category in display_order:
                    if category in status_counts:
                        ordered_statuses.append((category, status_counts[category]))
                
                # Add any remaining statuses that weren't in display order
                for status, count in status_counts.items():
                    if status not in display_order:
                        ordered_statuses.append((status, count))
            else:
                # Fallback to alphabetical order if no display order defined
                ordered_statuses = sorted(status_counts.items())
            
            for status, count in ordered_statuses:
                total_status_count += count
                
                # Add status name with conditional formatting
                status_cell = overall_summary[f'C{status_row}']
                status_cell.value = status
                style = apply_status_style(status_cell, status, config)
                status_cell.alignment = OVERALL_SUMMARY_STYLES['data_cell']['alignment']
                status_cell.border = OVERALL_SUMMARY_STYLES['border']
                
                # Add count with matching style
                count_cell = overall_summary[f'D{status_row}']
                count_cell.value = count
                count_cell.font = Font(
                    name=style['font'].name,
                    size=style['font'].size,
                    bold=style['font'].bold,
                    italic=style['font'].italic,
                    color=style['font'].color
                )
                count_cell.fill = style['fill']
                count_cell.alignment = OVERALL_SUMMARY_STYLES['data_cell']['alignment']
                count_cell.border = OVERALL_SUMMARY_STYLES['border']
                
                status_row += 1
            
            # Add total row for status counts
            overall_summary[f'C{status_row}'] = 'Total'
            overall_summary[f'D{status_row}'] = total_status_count
            overall_summary[f'C{status_row}'].font = OVERALL_SUMMARY_STYLES['total_cell']['font']
            overall_summary[f'C{status_row}'].alignment = OVERALL_SUMMARY_STYLES['total_cell']['alignment']
            overall_summary[f'C{status_row}'].fill = OVERALL_SUMMARY_STYLES['total_cell']['fill']
            overall_summary[f'C{status_row}'].border = OVERALL_SUMMARY_STYLES['border']
            overall_summary[f'D{status_row}'].font = OVERALL_SUMMARY_STYLES['total_cell']['font']
            overall_summary[f'D{status_row}'].alignment = OVERALL_SUMMARY_STYLES['total_cell']['alignment']
            overall_summary[f'D{status_row}'].fill = OVERALL_SUMMARY_STYLES['total_cell']['fill']
            overall_summary[f'D{status_row}'].border = OVERALL_SUMMARY_STYLES['border']
            
            return max(row, status_row) + 2
        
        # Revision/status summaries and pie chart removed
        # Data quality section provides more actionable insights
        # Historical revision/status data remains available in "Summary Data" sheet
        
        # Adjust column widths
        overall_summary.column_dimensions['A'].width = 30  # Certificate type / Revision
        overall_summary.column_dimensions['B'].width = 15  # Apartments / Count
        overall_summary.column_dimensions['C'].width = 12  # Progress % / Status
        overall_summary.column_dimensions['D'].width = 12  # Documents / Count
        overall_summary.column_dimensions['E'].width = 8   # Progress bar part 1
        overall_summary.column_dimensions['F'].width = 8   # Progress bar part 2
        overall_summary.column_dimensions['G'].width = 8   # Progress bar part 3
        
        # Add detailed uncategorized analysis tab if apartment tracking is enabled
        apartment_certs = config.get('CERTIFICATE_TRACKING', {}).get('apartment_certificates', {})
        if apartment_certs:
            add_uncategorized_detailed_tab(wb, latest_data, config)
        
        # Save the workbook
        wb.save(output_file)
        return True
        
    except Exception as e:
        print(f"Error generating certificate report: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


def save_certificate_report_with_retry(summary_df, latest_data, output_file, config, max_retries=3, retry_delay=1):
    """
    Save certificate report with retry logic for file access issues.
    
    Args:
        summary_df: DataFrame with summary data
        latest_data: DataFrame with the latest certificate data
        output_file: Path to the output Excel file
        config: Project configuration dictionary
        max_retries: Maximum number of retry attempts
        retry_delay: Delay between retries in seconds
        
    Returns:
        bool: True if successful, False otherwise
    """
    for attempt in range(max_retries):
        if save_certificate_report(summary_df, latest_data, output_file, config):
            return True
        
        if attempt < max_retries - 1:
            print(f"Retry {attempt + 1}/{max_retries - 1} in {retry_delay} seconds...")
            time.sleep(retry_delay)
    
    return False

