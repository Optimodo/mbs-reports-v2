"""Progression report generation module."""

import warnings
import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.page import PageMargins

from styles.formatting import PROGRESSION_STATUS_ORDER
from utils.status_mapping import (
    get_status_category,
    get_status_display_order,
    get_grouped_status_counts
)

# Suppress openpyxl warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')


def detect_new_revision_types(sheet, new_revisions, revision_type):
    """Detect if there are new revision types that don't exist in the current progression report.
    
    Args:
        sheet: openpyxl worksheet object
        new_revisions: List of new revision column names
        revision_type: 'P', 'C', or other revision type prefix
        
    Returns:
        list: Sorted list of new revision names not found in the sheet
    """
    if not sheet or not new_revisions:
        return []
    
    # Get existing revision labels from column A
    existing_revisions = set()
    section_start_row = None
    section_end_row = None
    
    # Find the section for this revision type
    section_title = f'{revision_type} Revision Progression'
    section_found = False
    
    for row in range(1, sheet.max_row + 1):
        cell_value = sheet[f'A{row}'].value
        if cell_value == section_title:
            section_found = True
            section_start_row = row + 2  # Skip header and column header rows
            continue
        
        # If we've found the section, collect revision labels until we hit the next section or total
        if section_found and cell_value:
            if 'Total' in str(cell_value) or 'Progression' in str(cell_value):
                section_end_row = row - 1
                break
            # Extract revision name from the label (remove any prefix)
            revision_name = str(cell_value).replace('Rev_', '').replace('Status_', '')
            if revision_name and not revision_name.startswith('P Revisions') and not revision_name.startswith('C Revisions'):
                existing_revisions.add(revision_name)
    
    # Compare with new revisions
    new_revision_names = {rev.replace('Rev_', '') for rev in new_revisions}
    missing_revisions = new_revision_names - existing_revisions
    
    return sorted(list(missing_revisions))


def fill_empty_cells_with_zeros_in_file(progression_report_path):
    """Open the progression report, fill empty cells in tables with zeros, and save.
    
    Args:
        progression_report_path: Path to the progression report Excel file
    """
    print(f"Filling empty cells in {progression_report_path}...")
    wb = load_workbook(progression_report_path)
    if 'Progression Report' not in wb.sheetnames:
        print("No 'Progression Report' sheet found.")
        return
    sheet = wb['Progression Report']
    
    # Find all date headers and their column numbers
    date_columns = set()
    for row_num in range(1, sheet.max_row + 1):
        for col_num in range(2, sheet.max_column + 1):
            col_letter = chr(ord('A') + col_num - 1)
            date_cell = sheet[f'{col_letter}{row_num}']
            if date_cell.value and '-' in str(date_cell.value) and len(str(date_cell.value)) > 5:
                date_columns.add(col_num)
    
    print(f"  Found date columns: {sorted(date_columns)}")
    
    # Find all table sections
    date_headers = []
    for row_num in range(1, sheet.max_row + 1):
        for col_num in range(2, sheet.max_column + 1):
            col_letter = chr(ord('A') + col_num - 1)
            date_cell = sheet[f'{col_letter}{row_num}']
            if date_cell.value and '-' in str(date_cell.value) and len(str(date_cell.value)) > 5:
                # This looks like a date header, find the next total row
                for next_row in range(row_num + 1, sheet.max_row + 1):
                    next_cell = sheet[f'A{next_row}']
                    if next_cell.value and 'Total' in str(next_cell.value):
                        date_headers.append((row_num, next_row))
                        break
                break
    
    # For each table section, fill empty cells with zeros only in date columns
    total_cells_filled = 0
    for start_row, end_row in date_headers:
        for row in range(start_row + 1, end_row):  # Skip the date header row
            for col_num in date_columns:  # Only process columns with date headers
                col_letter = chr(ord('A') + col_num - 1)
                cell = sheet[f'{col_letter}{row}']
                if cell.value is None or cell.value == '':
                    sheet[f'{col_letter}{row}'] = 0
                    # Apply consistent formatting with other data cells
                    sheet[f'{col_letter}{row}'].font = Font(name='Calibri', size=11)
                    sheet[f'{col_letter}{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    total_cells_filled += 1
    
    wb.save(progression_report_path)
    print(f"Filled {total_cells_filled} empty cells in {progression_report_path}.")


def generate_progression_report(summary_df, output_file, config, latest_data_df=None):
    """Generate a report showing the progression of revisions and statuses over time.
    
    This function creates a comprehensive progression report that tracks:
    - P revision progression over time
    - P revision status distribution
    - C revision progression over time
    - C revision status distribution
    
    The report can be incrementally updated by adding new columns for each time period.
    
    Args:
        summary_df: DataFrame with summary data for a single time period
        output_file: Path to output Excel file (will be created or updated)
        config: Project configuration dictionary
        latest_data_df: Optional DataFrame with detailed latest document data for filtering
        
    Returns:
        bool: True if successful, False otherwise
    """
    try:
        print(f"Debug: Starting progression report generation for {output_file}")
        print(f"Debug: summary_df has {len(summary_df)} rows")
        
        # Create a new workbook or load existing
        if os.path.exists(output_file):
            wb = load_workbook(output_file)
            print(f"Debug: Loaded existing progression report")
        else:
            wb = Workbook()
            # Remove the default 'Sheet' worksheet
            wb.remove(wb['Sheet'])
            print(f"Debug: Created new progression report")
        
        # Get or create the Progression Report sheet
        if 'Progression Report' in wb.sheetnames:
            sheet = wb['Progression Report']
            # Get the last used column
            last_col = 1  # Start with column A
            for col in sheet.columns:
                if any(cell.value for cell in col):
                    last_col = max(last_col, col[0].column)
            next_col = last_col + 1
            print(f"Debug: Existing sheet found, next column will be {next_col}")
        else:
            sheet = wb.create_sheet('Progression Report')
            next_col = 2  # Start with column B (A is for labels)
            print(f"Debug: Created new sheet, starting with column {next_col}")
        
        # Set up the sheet if it's new
        if next_col == 2:
            sheet.page_setup.fitToWidth = 1
            sheet.page_setup.fitToHeight = 0
            sheet.page_setup.horizontalCentered = True
            sheet.page_setup.verticalCentered = True
            sheet.page_margins = PageMargins(left=0.25, right=0.25, top=0.75, bottom=0.75, header=0.3, footer=0.3)
            
            # Add title
            sheet.merge_cells('A1:Z1')
            sheet['A1'] = f"{config.get('PROJECT_TITLE', '')} Document Register Progression Report"
            sheet['A1'].font = Font(name='Calibri', size=14, bold=True)
            sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Get all revision and status columns
        all_columns = summary_df.columns.tolist()
        rev_columns = [col for col in all_columns if col.startswith('Rev_')]
        status_columns = [col for col in all_columns if col.startswith('Status_')]
        
        # Sort revisions by type (P, C, other)
        p_revs = sorted([col for col in rev_columns if col.startswith('Rev_P')], 
                       key=lambda x: int(x.split('_')[1][1:]) if x.split('_')[1][1:].isdigit() else float('inf'))
        c_revs = sorted([col for col in rev_columns if col.startswith('Rev_C')],
                       key=lambda x: int(x.split('_')[1][1:]) if x.split('_')[1][1:].isdigit() else float('inf'))
        other_revs = sorted([col for col in rev_columns if not (col.startswith('Rev_P') or col.startswith('Rev_C'))])
        
        # Detect new revision types if this is an existing report
        if next_col > 2:  # Existing report
            new_p_revs = detect_new_revision_types(sheet, p_revs, 'P')
            new_c_revs = detect_new_revision_types(sheet, c_revs, 'C')
            
            if new_p_revs:
                print(f"WARNING: New P revision types detected: {', '.join(new_p_revs)}")
                print("These revisions will not be properly aligned in the progression report.")
                print("Consider regenerating the progression report to include proper row structure.")
            
            if new_c_revs:
                print(f"WARNING: New C revision types detected: {', '.join(new_c_revs)}")
                print("These revisions will not be properly aligned in the progression report.")
                print("Consider regenerating the progression report to include proper row structure.")
        
        # Function to get status count for a specific status group (all documents)
        def get_status_count(status_group):
            total = 0
            
            # Use config-based mappings if available
            if config and 'STATUS_MAPPINGS' in config:
                status_terms = config['STATUS_MAPPINGS'].get(status_group, {}).get('statuses', [])
            else:
                # Fallback to hardcoded PROGRESSION_STATUS_ORDER
                status_terms = PROGRESSION_STATUS_ORDER.get(status_group, {}).get('status_terms', [])
            
            for col in status_columns:
                # Extract the actual status name from the column name
                status_name = col.replace('Status_', '')
                if status_name in status_terms:
                    total += summary_df.iloc[-1].get(col, 0)
            return total
        
        # Function to get status count for a specific status group filtered by revision type
        def get_filtered_status_count(status_group, revision_type):
            if latest_data_df is None:
                # Fallback to unfiltered count if no latest data available
                return get_status_count(status_group)
            
            total = 0
            
            # Use config-based mappings if available
            if config and 'STATUS_MAPPINGS' in config:
                status_terms = config['STATUS_MAPPINGS'].get(status_group, {}).get('statuses', [])
            else:
                # Fallback to hardcoded PROGRESSION_STATUS_ORDER
                status_terms = PROGRESSION_STATUS_ORDER.get(status_group, {}).get('status_terms', [])
            
            # Filter data by revision type
            if revision_type == 'P':
                filtered_data = latest_data_df[latest_data_df['Rev'].str.startswith('P', na=False)]
            elif revision_type == 'C':
                filtered_data = latest_data_df[latest_data_df['Rev'].str.startswith('C', na=False)]
            else:
                # For other revision types, return 0 or handle as needed
                return 0
            
            # Count documents with statuses in the status_terms list
            for status_term in status_terms:
                count = len(filtered_data[filtered_data['Status'] == status_term])
                total += count
            
            return total
        
        # Function to get count of uncategorized statuses for a revision type
        def get_other_status_count(revision_type):
            if latest_data_df is None:
                return 0
            
            # Filter data by revision type
            if revision_type == 'P':
                filtered_data = latest_data_df[latest_data_df['Rev'].str.startswith('P', na=False)]
            elif revision_type == 'C':
                filtered_data = latest_data_df[latest_data_df['Rev'].str.startswith('C', na=False)]
            else:
                return 0
            
            # Get all defined status terms from config or fallback
            all_defined_statuses = set()
            
            if config and 'STATUS_MAPPINGS' in config:
                # If there's an 'Other' category in the mapping, it's already counted
                # So we return 0 to avoid double counting
                if 'Other' in config['STATUS_MAPPINGS']:
                    return 0
                
                # Otherwise, collect all defined statuses from config
                for status_group, mapping in config['STATUS_MAPPINGS'].items():
                    all_defined_statuses.update(mapping.get('statuses', []))
            else:
                # Fallback to hardcoded PROGRESSION_STATUS_ORDER
                for status_group in PROGRESSION_STATUS_ORDER.values():
                    all_defined_statuses.update(status_group.get('status_terms', []))
            
            # Count documents with statuses not in the defined list
            other_count = 0
            for status in filtered_data['Status'].unique():
                if status not in all_defined_statuses:
                    count = len(filtered_data[filtered_data['Status'] == status])
                    other_count += count
            
            return other_count
        
        # Start row for data
        current_row = 3
        
        # Track total row positions for each section
        total_row_positions = {
            'P Revisions Total': None,
            'P Status Total': None,
            'C Revisions Total': None,
            'C Status Total': None
        }
        
        # Function to fill empty cells with zeros in table sections
        def fill_empty_cells_with_zeros():
            """Fill any empty cells in table sections with zeros"""
            print("  Debug: Starting fill_empty_cells_with_zeros function")
            
            # Find all date headers (column headers) to identify table sections
            date_headers = []
            for row_num in range(1, sheet.max_row + 1):
                cell = sheet[f'A{row_num}']
                # Look for date headers - they should be in columns B onwards and contain date-like patterns
                for col_num in range(2, sheet.max_column + 1):
                    col_letter = chr(ord('A') + col_num - 1)
                    date_cell = sheet[f'{col_letter}{row_num}']
                    if date_cell.value and '-' in str(date_cell.value) and len(str(date_cell.value)) > 5:
                        # This looks like a date header, find the next total row
                        for next_row in range(row_num + 1, sheet.max_row + 1):
                            next_cell = sheet[f'A{next_row}']
                            if next_cell.value and 'Total' in str(next_cell.value):
                                date_headers.append((row_num, next_row))
                                print(f"  Debug: Found table section from row {row_num} to {next_row}")
                                break
                        break
            
            print(f"  Debug: Found {len(date_headers)} table sections")
            
            # For each table section, fill empty cells with zeros
            for start_row, end_row in date_headers:
                # Get the column range (from B to the last column with data)
                # Always include the current column being added (next_col)
                max_col = next_col - 1  # The current column being added
                # Also check existing columns for any that might have data
                for col in sheet.columns:
                    if any(cell.value for cell in col):
                        max_col = max(max_col, col[0].column)
                
                print(f"  Debug: Processing section rows {start_row + 1} to {end_row}, columns B to {chr(ord('A') + max_col - 1)}")
                
                # Fill empty cells in this section
                cells_filled = 0
                for row in range(start_row + 1, end_row):  # Skip the date header row
                    for col_num in range(2, max_col + 1):  # Start from column B
                        col_letter = chr(ord('A') + col_num - 1)
                        cell = sheet[f'{col_letter}{row}']
                        if cell.value is None or cell.value == '':
                            sheet[f'{col_letter}{row}'] = 0
                            sheet[f'{col_letter}{row}'].font = Font(name='Calibri', size=11)
                            sheet[f'{col_letter}{row}'].alignment = Alignment(horizontal='center', vertical='center')
                            cells_filled += 1
                
                print(f"  Debug: Filled {cells_filled} empty cells in this section")
            
            print("  Debug: Completed fill_empty_cells_with_zeros function")
        
        # Function to add a section header
        def add_section_header(title, row):
            if next_col == 2:  # Only add headers for new sheets
                sheet.merge_cells(f'A{row}:Z{row}')
                sheet[f'A{row}'] = title
                sheet[f'A{row}'].font = Font(name='Calibri', size=12, bold=True)
                sheet[f'A{row}'].fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
            else:
                # For existing sheets, find the section header by name
                section_row_found = None
                for row_num in range(1, sheet.max_row + 1):
                    # Check if this cell is part of a merged range
                    cell = sheet[f'A{row_num}']
                    if hasattr(cell, 'coordinate') and cell.coordinate in sheet.merged_cells:
                        # For merged cells, get the value from the top-left cell of the merged range
                        for merged_range in sheet.merged_cells.ranges:
                            if cell.coordinate in merged_range:
                                # Get the top-left cell of the merged range
                                top_left_cell = sheet[f'A{merged_range.min_row}']
                                cell_value = top_left_cell.value
                                break
                    else:
                        cell_value = cell.value
                    
                    if cell_value and title in str(cell_value):
                        section_row_found = row_num
                        break
                
                if section_row_found:
                    # Use the found row position
                    row = section_row_found
            
            return row + 1
        
        # Function to add column headers
        def add_column_headers(row):
            # Add date header for the new column
            date_col = chr(ord('A') + next_col - 1)
            latest_data = summary_df.iloc[-1]  # Get the latest data
            sheet[f'{date_col}{row}'] = latest_data['Date']  # Only use the date
            sheet[f'{date_col}{row}'].font = Font(name='Calibri', size=11, bold=True)
            sheet[f'{date_col}{row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            return row + 1
        
        # Function to add data rows with proper row matching
        def add_data_rows(columns, start_row, title_prefix=''):
            row = start_row
            latest_data = summary_df.iloc[-1]  # Get the latest data
            rows_inserted = 0  # Track how many rows we've inserted
            
            if next_col == 2:
                # New sheet - create row headers and add data
                for col in columns:
                    # Add row header
                    sheet[f'A{row}'] = f"{title_prefix}{col.replace('Rev_', '').replace('Status_', '')}"
                    sheet[f'A{row}'].font = Font(name='Calibri', size=11)
                    sheet[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
                    
                    # Add data for the new column
                    date_col = chr(ord('A') + next_col - 1)
                    value = latest_data.get(col, 0)
                    sheet[f'{date_col}{row}'] = value
                    sheet[f'{date_col}{row}'].font = Font(name='Calibri', size=11)
                    sheet[f'{date_col}{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    row += 1
            else:
                # Existing sheet - match data to existing row headers
                # First, read existing row headers from column A
                existing_headers = []
                current_row_num = start_row
                
                # Read headers until we hit a total row or next section
                while current_row_num <= sheet.max_row:
                    cell_value = sheet[f'A{current_row_num}'].value
                    if not cell_value or 'Total' in str(cell_value) or 'Progression' in str(cell_value):
                        break
                    existing_headers.append((current_row_num, str(cell_value)))
                    current_row_num += 1
                
                # Create a mapping of revision names to their data values
                revision_data = {}
                for col in columns:
                    revision_name = col.replace('Rev_', '').replace('Status_', '')
                    revision_data[revision_name] = latest_data.get(col, 0)
                
                # Add data to existing rows and track which revisions we've handled
                handled_revisions = set()
                for header_row, header_text in existing_headers:
                    # Extract revision name from header (remove any prefix)
                    header_revision = header_text.replace('Rev_', '').replace('Status_', '')
                    
                    # Find matching revision data
                    if header_revision in revision_data:
                        # Add data to this row
                        date_col = chr(ord('A') + next_col - 1)
                        value = revision_data[header_revision]
                        
                        # Check if the cell is a merged cell before trying to edit it
                        cell = sheet[f'{date_col}{header_row}']
                        if not hasattr(cell, 'coordinate') or not cell.coordinate in sheet.merged_cells:
                            sheet[f'{date_col}{header_row}'] = value
                            sheet[f'{date_col}{header_row}'].font = Font(name='Calibri', size=11)
                            sheet[f'{date_col}{header_row}'].alignment = Alignment(horizontal='center', vertical='center')
                        else:
                            print(f"  Warning: Skipping merged cell at {date_col}{header_row}")
                        
                        handled_revisions.add(header_revision)
                    else:
                        # This row doesn't have data for this revision - add 0
                        date_col = chr(ord('A') + next_col - 1)
                        
                        # Check if the cell is a merged cell before trying to edit it
                        cell = sheet[f'{date_col}{header_row}']
                        if not hasattr(cell, 'coordinate') or not cell.coordinate in sheet.merged_cells:
                            sheet[f'{date_col}{header_row}'] = 0
                            sheet[f'{date_col}{header_row}'].font = Font(name='Calibri', size=11)
                            sheet[f'{date_col}{header_row}'].alignment = Alignment(horizontal='center', vertical='center')
                        else:
                            print(f"  Warning: Skipping merged cell at {date_col}{header_row}")
                
                # Find revisions that weren't handled (new revisions)
                new_revisions = set(revision_data.keys()) - handled_revisions
                
                if new_revisions:
                    print(f"INFO: Adding {len(new_revisions)} new revision(s) to existing progression report")
                    
                    # Sort new revisions properly (P01, P02, P10, P11, etc.)
                    def sort_revision_key(rev):
                        if rev.startswith('P') or rev.startswith('C'):
                            # Extract the number part
                            prefix = rev[0]  # P or C
                            number_part = rev[1:]
                            
                            # Handle special cases like P_Certificates
                            if '_' in number_part:
                                # For special cases like P_Certificates, put them at the end
                                return (prefix, float('inf'))
                            
                            try:
                                # Convert to int for proper numeric sorting
                                return (prefix, int(number_part))
                            except ValueError:
                                # If it's not a number, sort as string but after numeric ones
                                return (prefix, float('inf'))
                        else:
                            # For other revisions, sort as string
                            return ('Z', rev)  # Put at end
                    
                    sorted_new_revisions = sorted(new_revisions, key=sort_revision_key)
                    
                    # For each new revision, insert it at the appropriate position
                    for new_rev in sorted_new_revisions:
                        # Find the correct insertion position based on sorting
                        insert_position = None
                        
                        # Get all existing revisions in this section for comparison
                        section_revisions = []
                        for header_row, header_text in existing_headers:
                            if 'Total' not in header_text:
                                rev_name = header_text.replace('Rev_', '').replace('Status_', '')
                                section_revisions.append((header_row, rev_name))
                        
                        # Sort existing revisions
                        section_revisions.sort(key=lambda x: sort_revision_key(x[1]))
                        
                        # Find where to insert the new revision
                        for i, (header_row, rev_name) in enumerate(section_revisions):
                            if sort_revision_key(new_rev) < sort_revision_key(rev_name):
                                insert_position = header_row
                                break
                        
                        # If no position found, insert at the end (before total)
                        if insert_position is None:
                            if section_revisions:
                                insert_position = max(row for row, _ in section_revisions) + 1
                            else:
                                insert_position = start_row
                        
                        # Insert new row at the correct position
                        sheet.insert_rows(insert_position)
                        rows_inserted += 1  # Track the insertion
                        
                        # Update merged cells that are affected by the row insertion
                        # Get all merged ranges and update those that start at or after the insertion point
                        merged_ranges_to_update = []
                        for merged_range in list(sheet.merged_cells.ranges):
                            start_row = merged_range.min_row
                            end_row = merged_range.max_row
                            
                            # If the merged range starts at or after the insertion point, it needs to be updated
                            if start_row >= insert_position:
                                # Unmerge the current range
                                sheet.unmerge_cells(str(merged_range))
                                # Store the new range coordinates
                                new_start_row = start_row + 1
                                new_end_row = end_row + 1
                                merged_ranges_to_update.append((merged_range, new_start_row, new_end_row))
                        
                        # Re-merge the updated ranges
                        for old_range, new_start_row, new_end_row in merged_ranges_to_update:
                            # Create new range string
                            start_col = old_range.min_col
                            end_col = old_range.max_col
                            new_range = f"{chr(ord('A') + start_col - 1)}{new_start_row}:{chr(ord('A') + end_col - 1)}{new_end_row}"
                            sheet.merge_cells(new_range)
                        
                        # Update all row numbers in existing_headers that are >= insert_position
                        updated_headers = []
                        for header_row, header_text in existing_headers:
                            if header_row >= insert_position:
                                updated_headers.append((header_row + 1, header_text))
                            else:
                                updated_headers.append((header_row, header_text))
                        existing_headers = updated_headers
                        
                        # Update total row positions that are >= insert_position
                        for total_label, total_row in total_row_positions.items():
                            if total_row is not None and total_row >= insert_position:
                                total_row_positions[total_label] = total_row + 1
                        
                        # Add the new revision header
                        sheet[f'A{insert_position}'] = f"{title_prefix}{new_rev}"
                        sheet[f'A{insert_position}'].font = Font(name='Calibri', size=11)
                        sheet[f'A{insert_position}'].alignment = Alignment(horizontal='left', vertical='center')
                        
                        # Add the data
                        date_col = chr(ord('A') + next_col - 1)
                        value = revision_data[new_rev]
                        sheet[f'{date_col}{insert_position}'] = value
                        sheet[f'{date_col}{insert_position}'].font = Font(name='Calibri', size=11)
                        sheet[f'{date_col}{insert_position}'].alignment = Alignment(horizontal='center', vertical='center')
                        
                        # Update existing_headers to include the new row
                        existing_headers.append((insert_position, f"{title_prefix}{new_rev}"))
                        
                        print(f"  Added new revision: {new_rev} at row {insert_position}")
                
                # Return the row after the last handled row, adjusted for inserted rows
                if existing_headers:
                    row = max(header_row for header_row, _ in existing_headers) + 1
                else:
                    row = start_row
            
            return row
        
        # Function to add status data rows with revision filtering
        def add_status_data_rows(start_row, revision_type):
            row = start_row
            latest_data = summary_df.iloc[-1]  # Get the latest data
            
            # Get status display order from config or fallback
            if config and 'STATUS_DISPLAY_ORDER' in config:
                status_order = config['STATUS_DISPLAY_ORDER']
                status_mappings = config.get('STATUS_MAPPINGS', {})
            else:
                # Fallback to hardcoded order
                status_order = list(PROGRESSION_STATUS_ORDER.keys())
                status_mappings = PROGRESSION_STATUS_ORDER
            
            # Add each status in the defined order
            for status_group in status_order:
                # Add row header if it's a new sheet
                if next_col == 2:
                    # Get display name from config or fallback
                    if config and 'STATUS_MAPPINGS' in config:
                        display_name = status_mappings.get(status_group, {}).get('display_name', status_group)
                    else:
                        display_name = status_mappings.get(status_group, {}).get('display_name', status_group)
                    
                    sheet[f'A{row}'] = display_name
                    sheet[f'A{row}'].font = Font(name='Calibri', size=11)
                    sheet[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
                
                # Add filtered data for the new column
                date_col = chr(ord('A') + next_col - 1)
                value = get_filtered_status_count(status_group, revision_type)
                sheet[f'{date_col}{row}'] = value
                sheet[f'{date_col}{row}'].font = Font(name='Calibri', size=11)
                sheet[f'{date_col}{row}'].alignment = Alignment(horizontal='center', vertical='center')
                
                row += 1
            
            # Check if 'Other' is already included in status groups
            has_other_in_config = config and 'STATUS_MAPPINGS' in config and 'Other' in config['STATUS_MAPPINGS']
            
            # Only add "Other Status" row if it's not already in the config
            if not has_other_in_config:
                # Add row header if it's a new sheet
                if next_col == 2:
                    sheet[f'A{row}'] = 'Other Status'
                    sheet[f'A{row}'].font = Font(name='Calibri', size=11)
                    sheet[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
                
                # Add filtered data for the new column (always add, even if 0)
                date_col = chr(ord('A') + next_col - 1)
                other_count = get_other_status_count(revision_type)
                sheet[f'{date_col}{row}'] = other_count
                sheet[f'{date_col}{row}'].font = Font(name='Calibri', size=11)
                sheet[f'{date_col}{row}'].alignment = Alignment(horizontal='center', vertical='center')
                
                row += 1
            
            return row
        
        # Function to calculate total for a set of columns
        def calculate_total(columns):
            total = 0
            latest_data = summary_df.iloc[-1]
            for col in columns:
                value = latest_data.get(col, 0)
                # Handle nan values - treat them as 0
                if pd.isna(value):
                    value = 0
                total += value
            return total
        
        # Function to add a total row
        def add_total_row(start_row, total_value, label='Total'):
            # Use tracked position if available, otherwise find by name
            if next_col > 2:
                # Check if we have a tracked position for this total row
                if label in total_row_positions and total_row_positions[label] is not None:
                    actual_row = total_row_positions[label]
                else:
                    # Fallback: Search for the total row by name in column A
                    total_row_found = None
                    for row_num in range(1, sheet.max_row + 1):
                        # Check if this cell is part of a merged range
                        cell = sheet[f'A{row_num}']
                        if hasattr(cell, 'coordinate') and cell.coordinate in sheet.merged_cells:
                            # For merged cells, get the value from the top-left cell of the merged range
                            for merged_range in sheet.merged_cells.ranges:
                                if cell.coordinate in merged_range:
                                    # Get the top-left cell of the merged range
                                    top_left_cell = sheet[f'A{merged_range.min_row}']
                                    cell_value = top_left_cell.value
                                    break
                        else:
                            cell_value = cell.value
                        
                        if cell_value and label in str(cell_value):
                            total_row_found = row_num
                            break
                    
                    if total_row_found:
                        # Use the found row position and store it for future use
                        actual_row = total_row_found
                        total_row_positions[label] = actual_row
                    else:
                        # If not found, use the provided start_row
                        actual_row = start_row
                        total_row_positions[label] = actual_row
            else:
                # New sheet - add row header
                actual_row = start_row
                sheet[f'A{actual_row}'] = label
                sheet[f'A{actual_row}'].font = Font(name='Calibri', size=11, bold=True)
                sheet[f'A{actual_row}'].alignment = Alignment(horizontal='left', vertical='center')
                sheet[f'A{actual_row}'].fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
                # Store the position for future use
                total_row_positions[label] = actual_row
            
            # Add total data for the new column
            date_col = chr(ord('A') + next_col - 1)
            
            # Check if the cell is a merged cell before trying to edit it
            cell = sheet[f'{date_col}{actual_row}']
            if not hasattr(cell, 'coordinate') or not cell.coordinate in sheet.merged_cells:
                sheet[f'{date_col}{actual_row}'] = total_value
                sheet[f'{date_col}{actual_row}'].font = Font(name='Calibri', size=11, bold=True)
                sheet[f'{date_col}{actual_row}'].alignment = Alignment(horizontal='center', vertical='center')
                sheet[f'{date_col}{actual_row}'].fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
                print(f"  Added total {total_value} to {date_col}{actual_row} for {label}")
            else:
                print(f"  Warning: Skipping merged cell at {date_col}{actual_row} for total row {label}")
                # Try to find the correct cell in the merged range
                for merged_range in sheet.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        # Find the correct column within the merged range
                        if merged_range.min_col <= ord(date_col) - ord('A') + 1 <= merged_range.max_col:
                            # This column is within the merged range, try to write to the top-left cell
                            top_left_cell = sheet[f'{date_col}{merged_range.min_row}']
                            if not hasattr(top_left_cell, 'coordinate') or not top_left_cell.coordinate in sheet.merged_cells:
                                sheet[f'{date_col}{merged_range.min_row}'] = total_value
                                sheet[f'{date_col}{merged_range.min_row}'].font = Font(name='Calibri', size=11, bold=True)
                                sheet[f'{date_col}{merged_range.min_row}'].alignment = Alignment(horizontal='center', vertical='center')
                                sheet[f'{date_col}{merged_range.min_row}'].fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
                                print(f"  Added total {total_value} to {date_col}{merged_range.min_row} (merged range) for {label}")
                                break
                        else:
                            print(f"  Column {date_col} is not within merged range {merged_range}")
                        break
            
            return actual_row + 1
        
        # Add P Revision section
        current_row = add_section_header('P Revision Progression', current_row)
        current_row = add_column_headers(current_row)
        current_row = add_data_rows(p_revs, current_row)
        # Add total row for P revisions
        p_total = calculate_total(p_revs)
        current_row = add_total_row(current_row, p_total, 'P Revisions Total')
        current_row += 1  # Add spacing
        
        # Add P Revision Status section
        current_row = add_section_header('P Revision Status Progression', current_row)
        current_row = add_column_headers(current_row)
        current_row = add_status_data_rows(current_row, 'P')
        # Add total row for P revision statuses
        # Get status order from config or fallback
        if config and 'STATUS_DISPLAY_ORDER' in config:
            status_order = config['STATUS_DISPLAY_ORDER']
        else:
            status_order = list(PROGRESSION_STATUS_ORDER.keys())
        p_status_total = sum(get_filtered_status_count(status_group, 'P') for status_group in status_order) + get_other_status_count('P')
        current_row = add_total_row(current_row, p_status_total, 'P Status Total')
        current_row += 1  # Add spacing
        
        # Add C Revision section
        current_row = add_section_header('C Revision Progression', current_row)
        current_row = add_column_headers(current_row)
        current_row = add_data_rows(c_revs, current_row)
        # Add total row for C revisions
        c_total = calculate_total(c_revs)
        current_row = add_total_row(current_row, c_total, 'C Revisions Total')
        current_row += 1  # Add spacing
        
        # Add C Revision Status section
        current_row = add_section_header('C Revision Status Progression', current_row)
        current_row = add_column_headers(current_row)
        current_row = add_status_data_rows(current_row, 'C')
        # Add total row for C revision statuses
        c_status_total = sum(get_filtered_status_count(status_group, 'C') for status_group in status_order) + get_other_status_count('C')
        current_row = add_total_row(current_row, c_status_total, 'C Status Total')
        
        # Note: fill_empty_cells_with_zeros is now called once at the end of all processing
        # to avoid running it multiple times per file

        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                if hasattr(cell, 'column_letter'):
                    column_letter = cell.column_letter
                    break
            
            if not column_letter:
                continue
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = max(8, max_length + 2)
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook
        wb.save(output_file)
        return True
        
    except Exception as e:
        print(f"Error generating progression report: {str(e)}")
        return False

